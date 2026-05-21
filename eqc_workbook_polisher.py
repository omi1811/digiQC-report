from __future__ import annotations

import os
import re
import tempfile

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def normalize_location_key(col_name: str) -> tuple[str | None, str | None]:
    """
    Extract a normalized location key from a building-stage column.

    Logical identifiers such as A, B, 1, 2, and A1 are kept. Noisy location
    names such as Ramp, Columns, Near, and SW14 are routed to UNKNOWN.
    """
    if " - " not in str(col_name):
        return None, None

    loc_str, stage_str = str(col_name).rsplit(" - ", 1)
    loc_str = loc_str.strip().upper()
    stage_str = stage_str.strip()

    for special in ["ALL", "MLCP", "DEVELOPMENT", "UNKNOWN"]:
        if special in loc_str:
            return special, stage_str

    tokens = re.findall(r"\b[A-Z0-9]+\b", loc_str)
    noise_words = {"BUILDING", "TOWER", "WING", "BLOCK", "CLUSTER", "ZONE", "PHASE", "NEAR"}
    core_tokens = [token for token in tokens if token not in noise_words]

    if not core_tokens:
        return "UNKNOWN", stage_str

    core = core_tokens[-1]
    if re.match(r"^[A-Z0-9]{1,2}$", core):
        return core, stage_str

    return "UNKNOWN", stage_str


def _safe_sheet_title(title: str, used_titles: set[str]) -> str:
    base = re.sub(r"[\[\]:*?/\\]", " ", str(title)).strip() or "EQC report"
    base = base[:31]
    candidate = base
    suffix = 1
    while candidate in used_titles:
        token = f" {suffix}"
        candidate = f"{base[:31 - len(token)]}{token}"
        suffix += 1
    used_titles.add(candidate)
    return candidate


def _read_data_sheet(raw_file_path: str, sheet_name: str) -> pd.DataFrame:
    df = pd.read_excel(raw_file_path, sheet_name=sheet_name, header=0)
    df.columns = [str(c).strip() for c in df.columns]
    if "EQC Checklist" in df.columns and "Checklists" not in df.columns:
        df = df.rename(columns={"EQC Checklist": "Checklists"})
    if "Checklists" in df.columns:
        labels = df["Checklists"].astype(str).str.strip().str.upper()
        df = df[labels.ne("TOTAL")].copy()
    return df


def process_eqc_report(raw_file_path: str, output_file_path: str) -> None:
    with pd.ExcelFile(raw_file_path) as xls:
        data_sheets = [s for s in xls.sheet_names if "cumulative" in s.lower()]

    wb = Workbook()
    ws_index = wb.active
    ws_index.title = "Summary or Index"
    ws_index["A1"] = "Sheet"
    ws_index["A1"].font = Font(bold=True)

    header_row = 4
    header_font = Font(bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side("thin"),
        right=Side("thin"),
        top=Side("thin"),
        bottom=Side("thin"),
    )
    total_fill = PatternFill("solid", fgColor="C6EFCE")
    pct_fill = PatternFill("solid", fgColor="D9E1F2")
    all_fill = PatternFill("solid", fgColor="C6EFCE")
    drop_fill = PatternFill("solid", fgColor="FFFF00")
    used_titles = {"Summary or Index"}

    for sheet_name in data_sheets:
        df = _read_data_sheet(raw_file_path, sheet_name)
        if "Checklists" not in df.columns or df.empty:
            continue

        col_groups: dict[str, dict[str, list[str]]] = {}
        key_order: list[str] = []

        for col in df.columns:
            if col == "Checklists":
                continue
            key, stage = normalize_location_key(col)
            if key is None or stage not in {"Pre", "During", "Post"}:
                continue
            if key not in col_groups:
                col_groups[key] = {"Pre": [], "During": [], "Post": []}
                key_order.append(key)
            col_groups[key][stage].append(col)

        special_keys = {"UNKNOWN", "ALL", "MLCP", "DEVELOPMENT"}
        regular = list(dict.fromkeys([k for k in key_order if k not in special_keys]))
        special = list(dict.fromkeys([k for k in key_order if k in special_keys]))
        ordered_keys = regular + special
        if not ordered_keys:
            continue

        valid_cols = ["Checklists"]
        col_mapping: list[tuple[str, str]] = []
        for key in ordered_keys:
            for stage in ["Pre", "During", "Post"]:
                valid_cols.append(f"{key} - {stage}")
                col_mapping.append((key, stage))

        agg_rows = []
        for _, row in df.iterrows():
            new_row = {"Checklists": row["Checklists"]}
            for key, stage in col_mapping:
                src_cols = col_groups[key][stage]
                if src_cols:
                    vals = pd.to_numeric(row[src_cols], errors="coerce").fillna(0)
                    val = int(vals.sum())
                else:
                    val = 0
                new_row[f"{key} - {stage}"] = val
            agg_rows.append(new_row)

        new_df = pd.DataFrame(agg_rows)

        final_cols = ["Checklists"]
        for key in ordered_keys:
            loc_total = sum(new_df[f"{key} - {s}"].sum() for s in ["Pre", "During", "Post"])
            if loc_total > 0:
                final_cols.extend([f"{key} - {s}" for s in ["Pre", "During", "Post"]])
        new_df = new_df[final_cols]

        totals = new_df.iloc[:, 1:].sum(numeric_only=True)
        total_row = pd.DataFrame([["TOTAL"] + totals.tolist()], columns=new_df.columns)

        pct_vals = []
        for col in final_cols[1:]:
            key, stage = col.rsplit(" - ", 1)
            if stage == "Pre":
                pct_vals.append(100.0)
            else:
                pre_val = totals.get(f"{key} - Pre", 0)
                cur_val = totals.get(col, 0)
                pct_vals.append(round((cur_val / pre_val * 100) if pre_val else 0.0, 2))
        pct_row = pd.DataFrame([["Percentage"] + pct_vals], columns=new_df.columns)

        final_df = pd.concat([new_df, total_row, pct_row], ignore_index=True)

        highlight_cells: set[tuple[int, str]] = set()
        for key in ordered_keys:
            pre_col = f"{key} - Pre"
            dur_col = f"{key} - During"
            post_col = f"{key} - Post"
            if all(col in new_df.columns for col in [pre_col, dur_col, post_col]):
                for row_idx, checklist_row in new_df.iterrows():
                    pre_val = checklist_row[pre_col]
                    dur_val = checklist_row[dur_col]
                    post_val = checklist_row[post_col]
                    if (
                        pd.notna(pre_val)
                        and pd.notna(dur_val)
                        and (pre_val - dur_val) > 8
                    ) or (
                        pd.notna(dur_val)
                        and pd.notna(post_val)
                        and (dur_val - post_val) > 8
                    ):
                        highlight_cells.add((row_idx, key))

        out_sheet_name = _safe_sheet_title(sheet_name, used_titles)
        ws = wb.create_sheet(title=out_sheet_name)

        proj_title = re.sub(r"\s*Cumulative\s+rep(?:ort)?\s*$", "", sheet_name, flags=re.I).strip()
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(final_cols))
        ws.cell(1, 1).value = f"{proj_title} Pre-During-Post"
        ws.cell(1, 1).font = Font(bold=True, size=12)
        ws.cell(1, 1).alignment = header_align

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(final_cols))
        ws.cell(2, 1).value = "DD-MM-YYYY"
        ws.cell(2, 1).font = Font(italic=True, color="808080")
        ws.cell(2, 1).alignment = header_align

        for i, col in enumerate(final_cols, 1):
            cell = ws.cell(header_row, i, col)
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border

        for r_idx, row in final_df.iterrows():
            row_num = header_row + 1 + r_idx
            ws.cell(row_num, 1, row["Checklists"]).border = thin_border
            ws.cell(row_num, 1).alignment = Alignment(vertical="center", wrap_text=True)

            for c_idx, col_name in enumerate(final_cols[1:], 2):
                key = col_name.rsplit(" - ", 1)[0]
                val = row[col_name]

                display_val = val
                if r_idx < len(new_df) and val == 0:
                    pre_val = row[f"{key} - Pre"]
                    dur_val = row[f"{key} - During"]
                    post_val = row[f"{key} - Post"]
                    if pre_val == 0 and dur_val == 0 and post_val == 0:
                        display_val = ""

                cell = ws.cell(row_num, c_idx, display_val)
                cell.alignment = data_align
                cell.border = thin_border
                if isinstance(val, (int, float)) and r_idx < len(new_df):
                    cell.number_format = "#,##0"

                if key == "ALL" and r_idx < len(new_df):
                    cell.fill = all_fill
                elif (r_idx, key) in highlight_cells:
                    cell.fill = drop_fill

            if r_idx == len(new_df):
                for c in range(1, len(final_cols) + 1):
                    ws.cell(row_num, c).fill = total_fill
                    ws.cell(row_num, c).font = Font(bold=True)

            if r_idx == len(new_df) + 1:
                for c in range(1, len(final_cols) + 1):
                    cell = ws.cell(row_num, c)
                    cell.fill = pct_fill
                    if c > 1:
                        cell.number_format = "0.00"

        ws.column_dimensions[get_column_letter(1)].width = 45
        for c in range(2, len(final_cols) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 14

        ws_index.cell(ws_index.max_row + 1, 1).value = out_sheet_name

    os.makedirs(os.path.dirname(os.path.abspath(output_file_path)) or ".", exist_ok=True)
    wb.save(output_file_path)


def polish_eqc_workbook_in_place(xlsx_path: str) -> None:
    """Rewrite an EQC workbook with the polished cumulative sheet layout."""
    abs_path = os.path.abspath(xlsx_path)
    directory = os.path.dirname(abs_path) or "."
    fd, tmp_path = tempfile.mkstemp(prefix="eqc_polished_", suffix=".xlsx", dir=directory)
    os.close(fd)
    try:
        process_eqc_report(abs_path, tmp_path)
        os.replace(tmp_path, abs_path)
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
