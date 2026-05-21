#!/usr/bin/env python3
from __future__ import annotations

import io
import os
import tempfile
import uuid
from datetime import datetime
from typing import Dict, List, Tuple

from flask import Flask, render_template, request, redirect, url_for, send_file
import pandas as pd

from project_utils import canonical_project_from_row
import analysis_eqc as EQC
import building_normalizer as bn
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment

app = Flask(__name__)

# --- Helpers ---
# RCC/Structural checklists - apply at FLOOR/POUR level, not flat level
RCC_CATEGORIES: List[Tuple[str, List[str]]] = [
    (
        "RCC / Structural",
        [
            r"\brcc\b|reinforc|shuttering|\bstruct|slab|column|beam|footing",
        ],
    ),
]

# Aluform checklists - structural (floor-level)
ALUFORM_CATEGORIES: List[Tuple[str, List[str]]] = [
    (
        "Aluform",
        [
            r"aluform",
        ],
    ),
]

# Handover checklists - can be at multiple levels
HANDOVER_CATEGORIES: List[Tuple[str, List[str]]] = [
    (
        "Internal Handover",
        [
            r"internal\s*handover|handover",
        ],
    ),
]

# Finishing checklists - apply at FLAT level
FINISHING_CATEGORIES: List[Tuple[str, List[str]]] = [
    (
        "Masonry & Surface Preparation",
        [
            r"aac|autoclaved|block\s*masonry|fly\s*ash|brick\s*work",
            r"internal\s*plaster|plaster\s*works|gypsum",
        ],
    ),
    (
        "Waterproofing",
        [
            r"waterproof.*toilet|skirting|boxtype|toilet\s*&\s*skirting",
            r"brick\s*bat",
            r"pu\s*waterproof",
        ],
    ),
    (
        "Flooring & Finishes",
        [
            r"tiling.*dado|dado\s*\(toilet|kitchen\)|kitchen\s*dado",
            r"tiling.*floor|flooring",
            r"kitchen\s*platform|sink",
            r"granite.*(frame|door|window)",
            r"staircase.*skirting",
        ],
    ),
    (
        "Carpentry & Metal Works",
        [
            r"carpentry.*(door\s*frame|shutters?)",
            r"(door\s*frame|shutters?)",
            r"aluminium.*sliding|aluminum.*sliding|sliding\s*door|sliding\s*window",
            r"railings?\s*&?\s*grills?|ms\s*grill",
            r"ss\s*railing",
        ],
    ),
    (
        "MEP (Mechanical, Electrical & Plumbing)",
        [
            r"internal\s*plumbing|plumbing\s*works",
            r"drainage.*pvc|bathroom.*drain|kitchen.*drain",
            r"electrical.*conduit|wall.*conduiting|wiring",
            r"fire\s*fight|sprinkler",
            r"mep.*before.*casting|sleeve|opening",
        ],
    ),
    (
        "Painting & Finishing",
        [
            r"painting.*internal",
            r"painting.*(railing|grills).*oil",
        ],
    ),
    (
        "Final Cleaning & Handover",
        [
            r"cleaning.*acid|deep\s*clean",
        ],
    ),
]

UNCATEGORIZED_LABEL = "Uncategorized / Other"


def is_rcc_checklist(name: str) -> bool:
    """Check if a checklist is RCC/structural (floor-level) vs Finishing (flat-level)."""
    import re as _re
    s = (name or "").lower()
    for _, patterns in RCC_CATEGORIES:
        for pat in patterns:
            try:
                if _re.search(pat, s, _re.I):
                    return True
            except Exception:
                continue
    return False


def is_aluform_checklist(name: str) -> bool:
    """Check if a checklist is Aluform (floor-level)."""
    import re as _re
    s = (name or "").lower()
    # Keep Aluform RCC/Internal handover separate from plain Aluform checklists.
    if "handover" in s:
        return False
    for _, patterns in ALUFORM_CATEGORIES:
        for pat in patterns:
            try:
                if _re.search(pat, s, _re.I):
                    return True
            except Exception:
                continue
    return False


def is_handover_checklist(name: str) -> bool:
    """Check if a checklist is Handover."""
    import re as _re
    s = (name or "").lower()
    for _, patterns in HANDOVER_CATEGORIES:
        for pat in patterns:
            try:
                if _re.search(pat, s, _re.I):
                    return True
            except Exception:
                continue
    return False


def base_name(n: str) -> str:
    """Standardize checklist names by removing contractor suffixes (matches Weekly_report.py logic)."""
    import re
    if not isinstance(n, str):
        return str(n)
    s = n.strip()
    s_norm = s.lower()
    
    # Handle "Checklist for ..." names - preserve full name with Title Case
    if s_norm.startswith('checklist for '):
        # Remove contractor suffix if present (double space + name pattern)
        cleaned = re.sub(r'\s{2,}[A-Za-z]+\s*$', '', s)
        cleaned = re.sub(r'\s+\([^)]+\)\s*$', '', cleaned)  # Remove (Contractor Name)
        return cleaned.strip().title()
    
    # Known standardized names
    fixes = {
        r'painting.*internal': 'Painting Works : Internal',
        r'painting.*external': 'Painting Works : External',
        r'painting\s*works\s*:\s*external\s*texture': 'Painting Works : External Texture',
        r'waterproof.*boxtype': 'Waterproofing Works: Toilet and Skirting',
        r'waterproof.*toilet': 'Waterproofing Works: Toilet and Skirting',
        r'waterproof.*skirting': 'Waterproofing Works: Toilet and Skirting',
        r'tiling.*kitchen.*platform': 'Tiling - Kitchen Platform',
        r'tiling.*kitchen.*sink': 'Tiling - Kitchen Platform',
        r'tiling[-\s]*toilet.*dado': 'Tiling - Toilet Dado',
        r'tiling.*flooring': 'Tiling - Flooring Work',
        r'kitchen\s*dado': 'Kitchen Dado',
        r'gypsum\s*plaster': 'Gypsum Plaster Works',
        r'aac\s*block': 'AAC Block Work',
        r'brick\s*work': 'Brick Work',
        r'external\s*plaster': 'External Plaster Works',
        r'internal\s*plaster': 'Internal Plaster Works',
        r'\brcc\b.*footing': 'RCC Footing Work',
        r'\brcc\b.*column': 'RCC Column Work',
        r'\brcc\b.*beam': 'RCC Beam Work',
        r'\brcc\b.*slab': 'RCC Slab Work',
        r'\brcc\b\s*work': 'RCC Works',
        r'm\s*s\s*railing': 'MS Railing',
        r's\s*s\s*railing': 'SS Railing',
        r'alluminium\s*window': 'Aluminium Window and Door Work',
        r'aluminium\s*window': 'Aluminium Window and Door Work',
        r'carpentry\s*work': 'Carpentry Work : Door Frame & Shutters',
        r'plumbing\s*work': 'Plumbing Work - Internal',
        r'electrical\s*work': 'Electrical Works',
        r'false\s*ceiling': 'False Ceiling Work',
        r'trimix': 'Trimix Work',
        r'aluform.*rcc.*handover.*checklist': 'Aluform RCC Handover Checklist',
        r'aluform(?!.*handover).*checklist': 'Aluform Checklist',
        r'fire\s*fighting': 'Fire Fighting Work',
    }
    for pat, name in fixes.items():
        if re.search(pat, s_norm):
            return name
    
    # Remove contractor name at the end (common patterns)
    # E.g., "GYPSUM PLASTER WORKS  Faujan" -> "GYPSUM PLASTER WORKS"
    s = re.sub(r'\s{2,}[A-Za-z]+\s*$', '', s)  # Double space followed by word
    s = re.sub(r'\s+[A-Z][a-z]+\s+[A-Z][a-z]+\s*$', '', s)  # "Om Enterprises" pattern
    s = re.sub(r'\s*:\s*$', '', s)  # Trailing colon
    s = re.sub(r'\s+$', '', s)  # Trailing spaces
    
    if '.' in s:
        return s.split('.', 1)[0].strip()
    if ':' in s and s.split(':', 1)[1].strip():
        return s  # Keep if there's text after colon
    elif ':' in s:
        return s.split(':', 1)[0].strip()
    
    return s.strip()


def map_checklist_to_category(name: str) -> Tuple[int, str]:
    """Map checklist to category (for backwards compatibility)."""
    import re as _re
    s = (name or "").lower()
    # Check RCC first
    for idx, (label, patterns) in enumerate(RCC_CATEGORIES):
        for pat in patterns:
            try:
                if _re.search(pat, s, _re.I):
                    return idx, label
            except Exception:
                continue
    # Check Handover
    handover_offset = len(RCC_CATEGORIES) + len(ALUFORM_CATEGORIES)
    for idx, (label, patterns) in enumerate(HANDOVER_CATEGORIES):
        for pat in patterns:
            try:
                if _re.search(pat, s, _re.I):
                    return handover_offset + idx, label
            except Exception:
                continue
    # Check Aluform (exclude handover variants)
    aluform_offset = len(RCC_CATEGORIES)
    for idx, (label, patterns) in enumerate(ALUFORM_CATEGORIES):
        for pat in patterns:
            try:
                if "handover" in s:
                    continue
                if _re.search(pat, s, _re.I):
                    return aluform_offset + idx, label
            except Exception:
                continue
    # Check Finishing
    finishing_offset = len(RCC_CATEGORIES) + len(ALUFORM_CATEGORIES) + len(HANDOVER_CATEGORIES)
    for idx, (label, patterns) in enumerate(FINISHING_CATEGORIES):
        for pat in patterns:
            try:
                if _re.search(pat, s, _re.I):
                    return finishing_offset + idx, label
            except Exception:
                continue
    return finishing_offset + len(FINISHING_CATEGORIES), UNCATEGORIZED_LABEL


def robust_read_eqc(path: str) -> pd.DataFrame:
    import warnings
    print(f"[DEBUG robust_read_eqc] Reading CSV from: {path}")

    def _looks_like_eqc(df: pd.DataFrame) -> bool:
        if not isinstance(df, pd.DataFrame):
            return False
        if df.shape[1] <= 1:
            return False
        colset = {str(c).strip().lower() for c in df.columns}
        expected = {
            "date", "project", "eqc", "eqc type", "eqc_type", "eqc stage",
            "location l0", "location l1", "location l2", "location l3", "status", "url"
        }
        return len(colset.intersection(expected)) >= 3
    
    last_err = None
    for sep in (",", "\t", None):
        try:
            # Try with on_bad_lines parameter (pandas >= 1.3.0)
            # Suppress ParserWarnings since we're handling bad lines
            with warnings.catch_warnings():
                warnings.filterwarnings('ignore', category=pd.errors.ParserWarning)
                try:
                    if sep is None:
                        result = pd.read_csv(path, dtype=str, keep_default_na=False, sep=None, engine="python", on_bad_lines='skip')
                    else:
                        result = pd.read_csv(path, dtype=str, keep_default_na=False, sep=sep, engine="python", on_bad_lines='skip')
                    if _looks_like_eqc(result):
                        print(f"[DEBUG robust_read_eqc] Successfully read {len(result)} rows with sep={repr(sep)}")
                        # Normalize building names
                        result = bn.normalize_dataframe(result)
                        return result
                    print(f"[DEBUG robust_read_eqc] Parsed with sep={repr(sep)} but schema mismatch, trying next...")
                except TypeError:
                    # Fall back for older pandas versions
                    if sep is None:
                        result = pd.read_csv(path, dtype=str, keep_default_na=False, sep=None, engine="python", error_bad_lines=False, warn_bad_lines=False)
                    else:
                        result = pd.read_csv(path, dtype=str, keep_default_na=False, sep=sep, engine="python", error_bad_lines=False, warn_bad_lines=False)
                    if _looks_like_eqc(result):
                        print(f"[DEBUG robust_read_eqc] Successfully read {len(result)} rows with sep={repr(sep)} (old pandas)")
                        # Normalize building names
                        result = bn.normalize_dataframe(result)
                        return result
                    print(f"[DEBUG robust_read_eqc] Parsed with sep={repr(sep)} (old pandas) but schema mismatch, trying next...")
        except Exception as e:
            last_err = e
            print(f"[DEBUG robust_read_eqc] Failed with sep={repr(sep)}: {e}")
    # Final fallback: permissive read with latin-1
    try:
        result = pd.read_csv(path, dtype=str, keep_default_na=False, encoding='latin-1', on_bad_lines='skip')
        if _looks_like_eqc(result):
            print(f"[DEBUG robust_read_eqc] Fallback latin-1 read succeeded with {len(result)} rows")
            # Normalize building names
            result = bn.normalize_dataframe(result)
            return result
    except Exception as e:
        last_err = e

    raise last_err if last_err else RuntimeError("Failed to read EQC file")


def prepare_frame(df: pd.DataFrame) -> pd.DataFrame:
    # Work on a copy to avoid SettingWithCopy warnings
    df = df.copy()
    
    print(f"[DEBUG prepare_frame] Input rows: {len(df)}")
    
    # Drop DEMO rows
    cols = [c for c in ("Project", "Project Name", "Location L0") if c in df.columns]
    if cols:
        mask = pd.Series(False, index=df.index)
        for c in cols:
            mask |= df[c].astype(str).str.contains("DEMO", case=False, na=False)
        df = df[~mask]
    
    print(f"[DEBUG prepare_frame] After DEMO filter: {len(df)}")
    # Stage normalization & dates
    if "Stage" in df.columns:
        checklist_cols = ["Eqc Type", "EQC Type", "Checklist", "Checklist Name", "__Checklist"]
        present_checklist_cols = [c for c in checklist_cols if c in df.columns]
        if present_checklist_cols:
            def _normalize_stage_row(row: pd.Series) -> str:
                checklist_name = ""
                for c in present_checklist_cols:
                    value = str(row.get(c, "") or "").strip()
                    if value:
                        checklist_name = value
                        break
                return EQC.normalize_stage(row.get("Stage", ""), checklist_name)

            df["__Stage"] = df.apply(_normalize_stage_row, axis=1)
        else:
            df["__Stage"] = df["Stage"].map(EQC.normalize_stage)
    if "Date" in df.columns:
        df["__Date"] = df["Date"].astype(str).map(EQC._parse_date_safe)
    else:
        df["__Date"] = None
    # Canonical project
    df["__Project"] = df.apply(canonical_project_from_row, axis=1)

    def _clean_loc_value(value: object) -> str:
        s = str(value or "").strip()
        if not s:
            return ""
        if s.lower() in {"nan", "none", "null", "na", "n/a", "unknown", "unknwn", "-", "--"}:
            return ""
        return s

    def _extract_location_from_eqc(eqc_value: object, project_value: object = "") -> Tuple[str, str]:
        import re as _re

        eqc = _clean_loc_value(eqc_value)
        project = _clean_loc_value(project_value)
        if not eqc:
            return "", ""

        raw_tokens = [_clean_loc_value(t) for t in _re.split(r"[>/|]+", eqc)]
        tokens = [t for t in raw_tokens if t]
        if project:
            project_norm = _re.sub(r"[^a-z0-9]+", "", project.lower())
            tokens = [
                t for t in tokens
                if _re.sub(r"[^a-z0-9]+", "", t.lower()) != project_norm
            ]

        building = ""
        floor = ""
        building_keywords = r"wing|tower|building|block|bldg|blk|buliding|biilding|bellding|bldng"
        building_patterns = [
            _re.compile(rf"\b({building_keywords})\b[\s\-]*([A-Za-z0-9]{{1,8}})\b", _re.I),
            _re.compile(rf"\b([A-Za-z0-9]{{1,8}})\b[\s\-]*({building_keywords})\b", _re.I),
        ]
        floor_patterns = [
            _re.compile(r"\b(ground\s*floor|ground|stilt|terrace|development|podium\s*\d*|parking\s*\d*|lobby|lift\s*room|stair\s*case|shear\s*wall|lower\s*ground|upper\s*ground|mezzanine|basement)\b", _re.I),
            _re.compile(r"\b(floor|fl|level|lvl)\s*(\d{1,2})\b", _re.I),
            _re.compile(r"\b(\d{1,2})(?:st|nd|rd|th)?\s*(floor|fl|level|lvl)\b", _re.I),
            _re.compile(r"\b(\d{1,2})\s*(?:flor|loor|flr)\b", _re.I),
        ]

        def _format_floor_token(token: str) -> str:
            token = _clean_loc_value(token)
            if not token:
                return ""
            special = _re.search(r"\b(ground\s*floor|ground|stilt|terrace|development|podium\s*\d*|parking\s*\d*|lobby|lift\s*room|stair\s*case|shear\s*wall|lower\s*ground|upper\s*ground|mezzanine|basement)\b", token, _re.I)
            if special:
                val = " ".join(w.capitalize() for w in special.group(1).split())
                return "Ground Floor" if val.lower() == "ground" else val
            m = _re.search(r"\b(floor|fl|level|lvl)\s*(\d{1,2})\b", token, _re.I) or _re.search(r"\b(\d{1,2})(?:st|nd|rd|th)?\s*(floor|fl|level|lvl)\b", token, _re.I) or _re.search(r"\b(\d{1,2})\s*(?:flor|loor|flr)\b", token, _re.I)
            if m:
                num = next((g for g in m.groups() if g and str(g).isdigit()), m.group(1))
                return f"Floor {int(num)}"
            m = _re.search(r"\b(?:flat|shop|unit|apt|apartment|room)\s*([A-Z]?\d{3,4}[A-Z]?)\b", token, _re.I) or _re.fullmatch(r"[A-Z]?\d{3,4}[A-Z]?", token)
            if m:
                code = m.group(1) if m.lastindex else m.group(0)
                digits = "".join(ch for ch in code if ch.isdigit())
                if len(digits) >= 3:
                    return f"Floor {int(digits[:-2])}"
            return ""

        for token in tokens:
            if not building:
                for pat in building_patterns:
                    m = pat.search(token)
                    if m:
                        g1, g2 = m.groups()
                        if g1.lower() in {"wing", "tower", "building", "block", "bldg", "blk", "buliding", "biilding", "bellding", "bldng"}:
                            kind = g1
                            ident = g2
                            kind_label = {"bldg": "Building", "blk": "Block", "buliding": "Building", "biilding": "Building", "bellding": "Building", "bldng": "Building"}.get(kind.lower(), kind.title())
                            building = f"{kind_label} {str(ident).strip()}"
                        else:
                            kind = g2
                            ident = g1
                            kind_label = {"bldg": "building", "blk": "block", "buliding": "building", "biilding": "building", "bellding": "building", "bldng": "building"}.get(kind.lower(), kind.lower())
                            building = f"{str(ident).strip()} {kind_label}"
                        break
            if not floor:
                for pat in floor_patterns:
                    m = pat.search(token)
                    if m:
                        floor = _format_floor_token(m.group(0))
                        break
                if not floor:
                    floor = _format_floor_token(token)
            if building and floor:
                break

        return building, floor

    if "EQC" in df.columns:
        extracted_locations = df.apply(
            lambda row: pd.Series(
                _extract_location_from_eqc(row.get("EQC", ""), row.get("__Project", "")),
                index=["__EQC_L1", "__EQC_L2"],
            ),
            axis=1,
        )
        df["Location L1"] = (
            df["Location L1"] if "Location L1" in df.columns else pd.Series("", index=df.index, dtype="object")
        )
        df["Location L2"] = (
            df["Location L2"] if "Location L2" in df.columns else pd.Series("", index=df.index, dtype="object")
        )
        l1_blank = df["Location L1"].map(_clean_loc_value).eq("")
        l2_blank = df["Location L2"].map(_clean_loc_value).eq("")
        df.loc[l1_blank, "Location L1"] = extracted_locations.loc[l1_blank, "__EQC_L1"]
        df.loc[l2_blank, "Location L2"] = extracted_locations.loc[l2_blank, "__EQC_L2"]

    mapping_debug = os.environ.get("DIGIQC_MAPPING_DEBUG", "0").strip().lower() in {"1", "true", "yes", "on"}
    eqc_only_mapping = os.environ.get("DIGIQC_EQC_ONLY_MAPPING", "0").strip().lower() in {"1", "true", "yes", "on"}
    mapping_stats = {
        "rows": 0,
        "b_from_location": 0,
        "f_from_location": 0,
        "fl_from_location": 0,
        "b_from_eqc": 0,
        "f_from_eqc": 0,
        "fl_from_eqc": 0,
        "b_unknown": 0,
        "f_unknown": 0,
        "fl_unknown": 0,
    }
    # Build a mapping of (Project, Building Letter) -> preferred label from CSV (preserve wording like Wing/Tower/Building)
    import re as _re_map
    bldg_pref_map: Dict[Tuple[str, str], str] = {}
    pref_counts: Dict[Tuple[str, str, str], int] = {}
    if "Location L1" in df.columns:
        for _, r in df.iterrows():
            proj = str(r.get("__Project", "") or "").strip()
            l1 = str(r.get("Location L1", "") or "").strip()
            if not proj or not l1:
                continue
            m = _re_map.search(r"\b(wing|tower|building|block|bldg|blk)\b[\s\-]*([A-Za-z0-9])\b", l1, _re_map.I) or \
                _re_map.search(r"\b([A-Za-z0-9])\b[\s\-]*\b(wing|tower|building|block|bldg|blk)\b", l1, _re_map.I)
            if m:
                g = m.groups()
                # extract the single-letter/number token regardless of order
                letter = g[1] if g[0].lower() in ("wing","tower","building","block","bldg","blk") else g[0]
                key = (proj, str(letter).upper())
                pref_counts[(proj, str(letter).upper(), l1)] = pref_counts.get((proj, str(letter).upper(), l1), 0) + 1
        # choose most frequent label for each (project, letter)
        for (proj, letter, label), cnt in pref_counts.items():
            cur = bldg_pref_map.get((proj, letter))
            if cur is None or cnt > pref_counts.get((proj, letter, cur), 0):
                bldg_pref_map[(proj, letter)] = label

    # Infer Building/Floor/Flat: prefer Location L1/L2/L3, but fall back to EQC fields (e.g., Eqc Type) if missing
    def infer_bld_floor_flat(row: pd.Series) -> Tuple[str, str, str]:
        import re as _re
        def _clean_loc(value: object) -> str:
            s = str(value or "").strip()
            if not s:
                return ""
            if s.lower() in {"nan", "none", "null", "na", "n/a", "unknown", "unknwn", "-", "--"}:
                return ""
            return s

        b_loc = _clean_loc(row.get("Location L1", ""))
        f_loc = _clean_loc(row.get("Location L2", ""))
        fl_loc = _clean_loc(row.get("Location L3", ""))

        if eqc_only_mapping:
            b = ""
            f = ""
            fl = ""
        else:
            b = b_loc
            f = f_loc
            fl = fl_loc
        # Candidate source strings to search for fallbacks
        sources = []
        for c in ("EQC", "Eqc", "Eqc Type", "EQC Type", "Stage", "Status", "Location", "Location Variable"):
            if c in row:
                sv = _clean_loc(row[c])
                if sv:
                    sources.append(sv)
        # Also include any column that looks like EQC path
        for c in row.index:
            sc = str(c).strip().lower()
            if ("eqc" in sc or sc in {"location", "location variable"}):
                sv = _clean_loc(row[c])
                if sv:
                    sources.append(sv)
        # Preserve order while removing duplicates.
        sources = list(dict.fromkeys(sources))
        # Tokenize sources for targeted heuristics
        tokens: List[str] = []
        for s in sources:
            tokens.extend([t.strip() for t in _re.split(r"[>/|,;:\-]+", str(s)) if _clean_loc(t)])
        # Building fallback patterns (A wing, Building A, Tower B, Block C)
        if not b:
            patt_building = [
                _re.compile(r"\b(wing|tower|building|block|bldg|blk|buliding|biilding|bellding|bldng)[\s\-]*([A-Za-z0-9]{1,8})\b", _re.I),
                _re.compile(r"\b([A-Za-z0-9]{1,8})[\s\-]*(wing|tower|building|block|bldg|blk|buliding|biilding|bellding|bldng)\b", _re.I),
                _re.compile(r"\b(bldg|blk)\.?[\s\-]*([A-Za-z0-9]{1,8})\b", _re.I),
                # do not use single-letter fallback unless no other hint
            ]
            proj = str(row.get("__Project", "") or "").strip()
            letter_found = None
            type_found = None
            for s in sources + tokens:
                for pat in patt_building:
                    m = pat.search(s)
                    if m:
                        # Normalize to 'A wing'
                        if len(m.groups()) == 2:
                            g1, g2 = m.groups()
                            type_token = g1 if g1.lower() in ("wing","tower","building","block","bldg","blk","buliding","biilding","bellding","bldng") else g2
                            letter_token = g2 if type_token == g1 else g1
                            letter_found = str(letter_token).strip().upper()
                            type_found = str(type_token).strip().lower()
                            break
                if letter_found:
                    break
            if not b and letter_found:
                # Prefer exact label used elsewhere in this project for this letter
                pref = bldg_pref_map.get((proj, letter_found))
                if pref:
                    b = pref
                elif type_found:
                    # Expand type token to a nice word
                    type_map = {"bldg": "Building", "blk": "Block", "buliding": "Building", "biilding": "Building", "bellding": "Building", "bldng": "Building"}
                    t = type_map.get(type_found, type_found).title()
                    b = f"{t} {letter_found}"
        # Floor from Location fields first (L0-L4), else fallback
        if not f:
            # Try Location L2 or L3 text for floor-like tokens
            for loc_field in ("Location L2", "Location L3", "Location L4", "Location L1", "Location L0"):
                sv = str(row.get(loc_field, "") or "").strip()
                if not sv:
                    continue
                # Recognize special non-flat structures
                if _re.search(r"\bshear\s*wall\b|stair\s*case|lobby|podium\s*\d*|development|terrace|parking\s*\d*|lift\s*room|ground\b", sv, _re.I):
                    f = "Ground Floor" if _re.search(r"^\s*ground\s*$", sv, _re.I) else sv
                    break
                m = _re.search(r"\b(floor|fl|level)[\s\-]*(\d{1,2})\b", sv, _re.I) or _re.search(r"\b(\d{1,2})[\s\-]*(floor|fl|level|flor|loor|flr)\b", sv, _re.I) or _re.search(r"\b(?:f|fl|lvl|l)[\s\-]*(\d{1,2})\b", sv, _re.I) or _re.search(r"\b(\d{1,2})(?:st|nd|rd|th)?\s*(?:floor|flor|loor|flr)?\b", sv, _re.I)
                if m:
                    num = next((g for g in m.groups() if g and str(g).isdigit()), m.group(m.lastindex or 1))
                    f = f"Floor {num}"
                    break
        # Floor fallback patterns (Floor 10, 10th Floor, Level 5) from EQC if still missing
        if not f:
            patt_floor = [
                _re.compile(r"\b(floor|fl|level)[\s\-]*(\d{1,2})\b", _re.I),
                _re.compile(r"\b(\d{1,2})[\s\-]*(floor|fl|level|flor|loor|flr)\b", _re.I),
                _re.compile(r"\b(?:f|fl|lvl|l)[\s\-]*(\d{1,2})\b", _re.I),
                _re.compile(r"\b(\d{1,2})(?:st|nd|rd|th)?\s*(?:floor|flor|loor|flr)?\b", _re.I),
            ]
            for s in sources + tokens:
                for pat in patt_floor:
                    m = pat.search(s)
                    if m:
                        # pick the last numeric group found
                        num = next((g for g in m.groups() if g and str(g).isdigit()), m.group(m.lastindex or 1))
                        f = f"Floor {num}"
                        break
                if f:
                    break
        # Flat fallback patterns (Flat 101, Shop 065, Unit 902A, Apartment 12)
        if not fl:
            # Require explicit keywords to avoid misreading product codes like "Samshield XL 1500" as a flat
            patt_flat = [
                _re.compile(r"\b(flat|unit|apt|apartment|shop)\s*(?:no\.?)?[\s\-]*([A-Z]?\d{1,4}[A-Z]?)\b", _re.I),
                _re.compile(r"\b(room)[\s\-]*(\d{1,4}[A-Z]?)\b", _re.I),
            ]
            for s in sources + tokens:
                for pat in patt_flat:
                    m = pat.search(s)
                    if m:
                        val = m.group(2) if m.lastindex and m.lastindex >= 2 else m.group(1)
                        # Normalize leading zeros: 065 -> 65
                        core = _re.sub(r"^([A-Z]?)(0+)(\d)", r"\1\3", val)
                        # Respect type if specified (Flat vs Shop)
                        typ = m.group(1).strip().title() if (m.lastindex and m.lastindex >= 2) else "Flat"
                        if typ.lower() not in ("flat","unit","apt","apartment","shop"):
                            typ = "Flat"
                        # Unit/Apt/Apartment -> Flat
                        if typ.lower() in ("unit","apt","apartment"):
                            typ = "Flat"
                        fl = f"{typ} {core}"
                        break
                if fl:
                    break
        # Standardize labels
        def _std_building(x: str) -> str:
            # Preserve CSV wording exactly if it came from Location L1
            return _clean_loc(x)

        def _std_floor(x: str) -> str:
            s = _clean_loc(x)
            if not s:
                return ""
            m = _re.search(r"(\d{1,2})", s)
            if m and _re.search(r"floor|fl|level|lvl|^\d{1,2}$", s, _re.I):
                return f"Floor {m.group(1)}"
            return s

        def _std_flat(x: str) -> str:
            s = _clean_loc(x)
            if not s:
                return ""
            if _re.search(r"^flat\s+", s, _re.I):
                return _re.sub(r"^flat\s+", "Flat ", s, flags=_re.I)
            m = _re.search(r"([A-Z]?\d{1,4}[A-Z]?)", s)
            if m:
                # Keep Shop vs Flat if present
                t = "Flat"
                m2 = _re.search(r"^(shop|flat)\b", s, _re.I)
                if m2:
                    t = m2.group(1).title()
                # strip leading zeros
                core = _re.sub(r"^([A-Z]?)(0+)(\d)", r"\1\3", m.group(1))
                return f"{t} {core}"
            return s

        b_final = _std_building(b)
        f_final = _std_floor(f)
        fl_final = _std_flat(fl)

        mapping_stats["rows"] += 1
        if (not eqc_only_mapping) and b_loc:
            mapping_stats["b_from_location"] += 1
        elif b_final:
            mapping_stats["b_from_eqc"] += 1
        else:
            mapping_stats["b_unknown"] += 1

        if (not eqc_only_mapping) and f_loc:
            mapping_stats["f_from_location"] += 1
        elif f_final:
            mapping_stats["f_from_eqc"] += 1
        else:
            mapping_stats["f_unknown"] += 1

        if (not eqc_only_mapping) and fl_loc:
            mapping_stats["fl_from_location"] += 1
        elif fl_final:
            mapping_stats["fl_from_eqc"] += 1
        else:
            mapping_stats["fl_unknown"] += 1

        return b_final, f_final, fl_final

    if len(df) > 0:
        bff = df.apply(infer_bld_floor_flat, axis=1, result_type="expand")
        if isinstance(bff, pd.DataFrame) and len(bff.columns) >= 3:
            df["__Building"] = bff[0].astype(str)
            df["__Floor"] = bff[1].astype(str)
            df["__Flat"] = bff[2].astype(str)
        else:
            df["__Building"] = ""
            df["__Floor"] = ""
            df["__Flat"] = ""
    else:
        df["__Building"] = ""
        df["__Floor"] = ""
        df["__Flat"] = ""
    
    print(f"[DEBUG prepare_frame] Unique buildings extracted: {sorted(df['__Building'].unique().tolist()[:10])}")
    print(f"[DEBUG prepare_frame] Unique floors extracted: {sorted(df['__Floor'].unique().tolist()[:10])}")
    
    # Extract Pour information from Location L3 if it contains "pour"
    def extract_pour(row: pd.Series) -> str:
        import re as _re
        for loc_field in ("Location L3", "Location L4", "Location Variable"):
            val = str(row.get(loc_field, "") or "").strip()
            if val and _re.search(r"\bpour\b", val, _re.I):
                # Extract pour number/name (e.g., "Pour 1", "Pour 2")
                m = _re.search(r"\bpour[\s\-]*(\d+|[A-Za-z0-9]+)\b", val, _re.I)
                if m:
                    return f"Pour {m.group(1)}"
                return val
        return ""
    df["__Pour"] = df.apply(extract_pour, axis=1)
    # Checklist name lives in 'Eqc Type' in our extracts
    if "Eqc Type" in df.columns:
        df["__Checklist"] = df["Eqc Type"].astype(str)
    elif "EQC Type" in df.columns:
        df["__Checklist"] = df["EQC Type"].astype(str)
    else:
        df["__Checklist"] = ""
    # Optional link
    if "URL" in df.columns:
        df["__URL"] = df["URL"].astype(str)
    else:
        df["__URL"] = ""
    
    print(f"[DEBUG prepare_frame] Final output rows: {len(df)}")
    print(f"[DEBUG prepare_frame] Sample buildings: {df['__Building'].value_counts().head(5).to_dict()}")

    if mapping_debug:
        print(
            "[DEBUG mapping] mode={} rows={} | Building loc/eqc/unknown={}/{}/{} | Floor loc/eqc/unknown={}/{}/{} | Flat loc/eqc/unknown={}/{}/{}".format(
                "EQC_ONLY" if eqc_only_mapping else "LOCATION_THEN_EQC",
                mapping_stats["rows"],
                mapping_stats["b_from_location"],
                mapping_stats["b_from_eqc"],
                mapping_stats["b_unknown"],
                mapping_stats["f_from_location"],
                mapping_stats["f_from_eqc"],
                mapping_stats["f_unknown"],
                mapping_stats["fl_from_location"],
                mapping_stats["fl_from_eqc"],
                mapping_stats["fl_unknown"],
            )
        )
    
    return df


@app.get("/flat-report")
def flat_report_index():
    # Local test reads from Combined_EQC.csv by default; allow path override
    path = request.args.get("path") or "Combined_EQC.csv"
    if not os.path.exists(path):
        return f"EQC file not found: {path}", 404
    df = prepare_frame(robust_read_eqc(path))
    projects = sorted([p for p in df["__Project"].astype(str).str.strip().unique() if p])
    proj = request.args.get("project") or (projects[0] if projects else None)
    dfp = df[df["__Project"].astype(str).str.strip().eq(proj)] if proj else df.iloc[0:0]

    buildings_all = sorted({(b if b else "UNKNOWN") for b in dfp["__Building"].astype(str).unique()})
    # Prefer a real building over UNKNOWN by default
    default_building = next((b for b in buildings_all if str(b).upper() != "UNKNOWN"), (buildings_all[0] if buildings_all else None))
    building = request.args.get("building") or default_building

    dfl = dfp[dfp["__Building"].astype(str).fillna("").replace("", "UNKNOWN").eq(building or "")]
    # Natural sort floors (Floor 1 .. Floor 21, Parking 1..)
    def _floor_key(v: str):
        import re as _re
        s = str(v or "").strip()
        m = _re.search(r"(floor|fl|level)\s*(\d+)", s, _re.I)
        if m:
            return (0, int(m.group(2)))
        m = _re.search(r"parking\s*(\d+)", s, _re.I)
        if m:
            return (1, int(m.group(1)))
        order = {"terrace": (2, 0), "development": (3, 0), "shops": (4, 0), "unknown": (9, 0)}
        key = s.lower()
        return order.get(key, (8, 0))
    floors_set = {(f if f else "UNKNOWN") for f in dfl["__Floor"].astype(str).unique()}
    # Fallback: if floors are missing or only UNKNOWN, derive from Location L0-L4 or EQC text
    def _derive_floor_from_row(row: pd.Series) -> str:
        import re as _re
        specials = r"shear\s*wall|stair\s*case|lobby|podium|development|terrace|parking|upper\s*ground|lower\s*ground|mezzanine|basement|pod"
        for loc_field in ("Location L2", "Location L3", "Location L4", "Location L1", "Location L0"):
            sv = str(row.get(loc_field, "") or "").strip()
            if not sv:
                continue
            if _re.search(specials, sv, _re.I):
                return sv
            m = _re.search(r"\b(floor|fl|level)[\s\-]*(\d{1,2})\b", sv, _re.I) or _re.search(r"\b(\d{1,2})[\s\-]*(floor|fl|level)\b", sv, _re.I) or _re.search(r"\b(?:f|fl|lvl|l)[\s\-]*(\d{1,2})\b", sv, _re.I) or _re.search(r"\b(\d{1,2})(?:st|nd|rd|th)?\s*(?:floor)?\b", sv, _re.I)
            if m:
                num = m.group(m.lastindex or 1)
                return f"Floor {num}"
        # Fallback to EQC path
        eqc = str(row.get("EQC", "") or "").strip()
        if eqc:
            # split tokens and reuse same patterns
            import re as _re2
            toks = [t.strip() for t in _re2.split(r"[>/|,;:\\-–—]+", eqc) if t.strip()]
            for sv in [eqc] + toks:
                if _re2.search(specials, sv, _re2.I):
                    return sv
                m = _re2.search(r"\b(floor|fl|level)[\s\-]*(\d{1,2})\b", sv, _re2.I) or _re2.search(r"\b(\d{1,2})[\s\-]*(floor|fl|level)\b", sv, _re2.I) or _re2.search(r"\b(?:f|fl|lvl|l)[\s\-]*(\d{1,2})\b", sv, _re2.I) or _re2.search(r"\b(\d{1,2})(?:st|nd|rd|th)?\s*(?:floor)?\b", sv, _re2.I)
                if m:
                    num = m.group(m.lastindex or 1)
                    return f"Floor {num}"
        return "UNKNOWN"

    if not floors_set or all(str(x).upper() == "UNKNOWN" for x in floors_set):
        alt = set()
        for _, r in dfl.iterrows():
            val = _derive_floor_from_row(r)
            if val:
                alt.add(val)
        floors_set = alt if alt else floors_set
    floors_all = sorted(floors_set, key=_floor_key)
    # Prefer a numeric Floor N as the default if available, else first non-UNKNOWN
    import re as _re_df
    default_floor = next((f for f in floors_all if _re_df.search(r"\bfloor\s*\d+\b", str(f), _re_df.I)), None)
    if not default_floor:
        default_floor = next((f for f in floors_all if str(f).upper() != "UNKNOWN"), (floors_all[0] if floors_all else None))
    floor = request.args.get("floor") or default_floor

    if floor == "ALL":
        dff = dfl.copy()
    else:
        dff = dfl[dfl["__Floor"].astype(str).fillna("").replace("", "UNKNOWN").eq(floor or "")]
    # Natural sort flats (Flat/Shop 101, 101A etc.)
    def _flat_key(v: str):
        import re as _re
        s = str(v or "").strip()
        m = _re.search(r"(shop|flat|unit|apt|apartment)\s*([A-Z]?\d+)([A-Z]?)", s, _re.I)
        if m:
            num = int(_re.sub(r"\D", "", m.group(2))) if _re.sub(r"\D", "", m.group(2)) else 0
            suf = m.group(3).upper() if m.group(3) else ""
            return (0, num, suf)
        # Fallback: any number in string
        m2 = _re.search(r"(\d+)", s)
        if m2:
            return (1, int(m2.group(1)), "")
        return (9, 0, s.lower())
    flats_set = {(x if x else "UNKNOWN") for x in dff["__Flat"].astype(str).unique()}
    # Prefer only units that look like Flat/Shop for dropdown, but fall back to all when none match
    import re as _re_fl
    flats_filtered = [x for x in flats_set if _re_fl.match(r"^(?:Flat|Shop)\s+", str(x))]
    flats_all = sorted(flats_filtered if flats_filtered else flats_set, key=_flat_key)
    default_flat = next((x for x in flats_all if str(x).upper() != "UNKNOWN"), (flats_all[0] if flats_all else None))
    flat = request.args.get("flat") or ("ALL" if floor == "ALL" else default_flat)

    # Scope for selected flat
    if floor == "ALL":
        # Building-wise: use entire building as scope
        scope_flat = dff.copy()
        scope_floor = dfl.copy()
    else:
        if (flat or "").upper() == "ALL":
            scope_flat = dff.copy()
        else:
            scope_flat = dff[dff["__Flat"].astype(str).fillna("").replace("", "UNKNOWN").eq(flat or "")].copy()
        scope_floor = dfl.copy()

    # Build grouped structure: category -> list of items
    items_by_cat: Dict[Tuple[int, str], List[Dict]] = {}
    # Collect all possible checklist names across this project (for "show all sections")
    all_checklists = set(dfp["__Checklist"].astype(str).str.strip())
    def _latest_per_checklist(frame: pd.DataFrame) -> pd.DataFrame:
        if frame is None or frame.empty:
            return frame.iloc[0:0]
        tmp = frame.copy()
        tmp["__ChecklistKey"] = tmp["__Checklist"].astype(str).str.strip().str.lower()
        tmp["__DateOrd"] = pd.to_datetime(tmp["__Date"]).fillna(pd.Timestamp.min)
        idxs = tmp.sort_values("__DateOrd").groupby("__ChecklistKey").tail(1).index
        return tmp.loc[idxs]

    # When ALL floors are selected, do not populate items (UI should show only export option)
    if floor == "ALL":
        reps_flat = scope_flat.iloc[0:0]
        reps_floor = scope_floor.iloc[0:0]
    else:
        reps_flat = _latest_per_checklist(scope_flat)
        reps_floor = _latest_per_checklist(scope_floor)

    # Populate items: RCC, Aluform, Handover, and Finishing checklists go to separate tables
    rcc_items: List[Dict] = []
    aluform_items: List[Dict] = []
    handover_items: List[Dict] = []
    finishing_items_by_cat: Dict[Tuple[int, str], List[Dict]] = {}
    
    # Combine all reps from both floor and flat level
    all_reps = pd.concat([reps_floor, reps_flat], ignore_index=True) if (not reps_floor.empty and not reps_flat.empty) else (reps_floor if not reps_floor.empty else reps_flat)
    
    if all_reps is not None and not all_reps.empty:
        for _, row in all_reps.iterrows():
            cl = row["__Checklist"]
            # Apply base_name for standardization
            cl_display = base_name(cl)
            url = row.get("__URL", "")
            stage = row.get("__Stage", "")
            status = str(row.get("Status", "")).strip()
            
            item = {
                "checklist": cl_display,
                "checklist_raw": cl,
                "url": url,
                "stage": stage,
                "status": status,
                "date": row.get("__Date"),
            }
            
            # Determine category based on checklist type, not source
            if is_rcc_checklist(cl):
                # RCC checklists (structural/floor-level) go to RCC table
                rcc_items.append(item)
            elif is_aluform_checklist(cl):
                # Aluform checklists (floor-level) go to Aluform table
                aluform_items.append(item)
            elif is_handover_checklist(cl):
                # Handover checklists go to Handover table
                handover_items.append(item)
            else:
                # Finishing checklists (flat-level) go to Finishing categories
                idx, cat = map_checklist_to_category(cl)
                if (idx, cat) not in finishing_items_by_cat:
                    finishing_items_by_cat[(idx, cat)] = []
                finishing_items_by_cat[(idx, cat)].append(item)
    
    # Build ordered keys for all categories
    ordered_rcc_keys: List[Tuple[int, str]] = [(0, label) for label, _ in RCC_CATEGORIES]
    ordered_aluform_keys: List[Tuple[int, str]] = [(len(RCC_CATEGORIES) + i, label) for i, (label, _) in enumerate(ALUFORM_CATEGORIES)]
    ordered_handover_keys: List[Tuple[int, str]] = [(len(RCC_CATEGORIES) + len(ALUFORM_CATEGORIES) + i, label) for i, (label, _) in enumerate(HANDOVER_CATEGORIES)]
    
    finishing_offset = len(RCC_CATEGORIES) + len(ALUFORM_CATEGORIES) + len(HANDOVER_CATEGORIES)
    ordered_finishing_keys: List[Tuple[int, str]] = []
    for i, (label, _) in enumerate(FINISHING_CATEGORIES):
        ordered_finishing_keys.append((finishing_offset + i, label))
    ordered_finishing_keys.append((finishing_offset + len(FINISHING_CATEGORIES), UNCATEGORIZED_LABEL))

    return render_template(
        "flat_report.html",
        path=path,
        projects=projects,
        selected_project=proj,
        buildings=buildings_all,
        selected_building=building,
        floors=floors_all,
        selected_floor=floor,
        flats=flats_all,
        selected_flat=flat,
        rcc_items=rcc_items,
        aluform_items=aluform_items,
        handover_items=handover_items,
        finishing_items_by_cat=finishing_items_by_cat,
        ordered_finishing_keys=ordered_finishing_keys,
        show_only_export=(floor == "ALL"),
    )


@app.get("/flat-report/export")
def flat_report_export():
    # Build an Excel status grid with separate sheets for RCC, Aluform, Handover, and Finishing
    path = request.args.get("path") or "Combined_EQC.csv"
    if not os.path.exists(path):
        return f"EQC file not found: {path}", 404
    df = prepare_frame(robust_read_eqc(path))
    
    # Debug: log total rows read
    print(f"[DEBUG] Total rows read from CSV: {len(df)}")
    
    projects = sorted([p for p in df["__Project"].astype(str).str.strip().unique() if p])
    proj = request.args.get("project") or (projects[0] if projects else None)
    dfp = df[df["__Project"].astype(str).str.strip().eq(proj)] if proj else df.iloc[0:0]
    
    print(f"[DEBUG] Project: {proj}, Rows after project filter: {len(dfp)}")
    
    building = request.args.get("building") or ""
    print(f"[DEBUG] Looking for building: '{building}'")
    print(f"[DEBUG] Available buildings in data: {sorted(dfp['__Building'].unique().tolist())}")
    
    # Try exact match first
    dfl = dfp[dfp["__Building"].astype(str).fillna("").replace("", "UNKNOWN").eq(building or "")]
    
    # If no exact match, try case-insensitive match
    if len(dfl) == 0 and building:
        print(f"[DEBUG] No exact match found, trying case-insensitive match...")
        dfl = dfp[dfp["__Building"].astype(str).str.lower().fillna("").replace("", "unknown").eq(building.lower())]
    
    print(f"[DEBUG] Building: {building}, Rows after building filter: {len(dfl)}")
    
    floor = request.args.get("floor") or ""
    if floor == "ALL":
        dff = dfl.copy()
    else:
        dff = dfl[dfl["__Floor"].astype(str).fillna("").replace("", "UNKNOWN").eq(floor or "")]
    
    print(f"[DEBUG] Floor: {floor}, Rows after floor filter: {len(dff)}")
    
    # If no data found, provide helpful error message
    if len(dff) == 0:
        print(f"[ERROR] No data found after filtering!")
        print(f"[ERROR] Filters: Project='{proj}', Building='{building}', Floor='{floor}'")
        if len(dfp) > 0:
            available_buildings = sorted(dfp["__Building"].unique().tolist())
            available_floors = sorted(dfp["__Floor"].unique().tolist())
            print(f"[ERROR] Available buildings for this project: {available_buildings}")
            print(f"[ERROR] Available floors for this project: {available_floors}")
        return f"No data found for the selected filters. Project: {proj}, Building: {building}, Floor: {floor}. Please check the browser console for available options.", 404

    # Determine flats on this floor
    def _flat_key(v: str):
        import re as _re
        s = str(v or "").strip()
        m = _re.search(r"(shop|flat|unit|apt|apartment)\s*([A-Z]?\d+)([A-Z]?)", s, _re.I)
        if m:
            num = int(_re.sub(r"\D", "", m.group(2))) if _re.sub(r"\D", "", m.group(2)) else 0
            suf = m.group(3).upper() if m.group(3) else ""
            return (0, num, suf)
        m2 = _re.search(r"(\d+)", s)
        if m2:
            return (1, int(m2.group(1)), "")
        return (9, 0, s.lower())
    flats = sorted(set(dff["__Flat"].astype(str).fillna("").replace("", "UNKNOWN")), key=_flat_key)
    
    # Helper function to build floor-level sheet (for RCC, Aluform, Handover)
    def build_floor_status_sheet(tmp_data: pd.DataFrame) -> pd.DataFrame:
        """Build floor-level status list for RCC/Aluform/Handover checklists (not flat-wise)."""
        if tmp_data.empty:
            return pd.DataFrame()
        
        tmp = tmp_data.copy()
        tmp["__ChecklistKey"] = tmp["__Checklist"].astype(str).str.strip().str.lower()
        tmp["__DateOrd"] = pd.to_datetime(tmp["__Date"]).fillna(pd.Timestamp.min)
        # Get latest entry per checklist (floor-level)
        tmp = tmp.sort_values("__DateOrd").groupby("__ChecklistKey", as_index=False).tail(1)

        # Build list with checklist details
        def status_symbol(status: str) -> str:
            up = (status or "").strip().upper()
            if up == "PASSED":
                return "✓ Passed"
            if "PROGRESS" in up:
                return "– In Progress"
            if "REDO" in up:
                return "↻ Redo"
            if "FAIL" in up:
                return "✗ Failed"
            return "– Pending"
        
        def get_category(checklist: str) -> str:
            """Determine category for a checklist."""
            if is_rcc_checklist(checklist):
                return "RCC / Structural"
            elif is_aluform_checklist(checklist):
                return "Aluform"
            elif is_handover_checklist(checklist):
                return "Internal Handover"
            return "Other"

        grid = []
        for _, row in tmp.iterrows():
            cl_raw = str(row["__Checklist"])
            cl = base_name(cl_raw)
            category = get_category(cl_raw)
            status = status_symbol(str(row.get("Status", "")))
            stage = str(row.get("__Stage", ""))
            date = str(row.get("__Date", ""))
            pour = str(row.get("__Pour", "")).strip() if "__Pour" in row and row.get("__Pour") else ""
            
            grid.append({
                "Category": category,
                "Checklist": cl,
                "Pour": pour if pour else "–",
                "Stage": stage,
                "Status": status,
                "Date": date,
            })

        return pd.DataFrame(grid, columns=["Category", "Checklist", "Pour", "Stage", "Status", "Date"]) if grid else pd.DataFrame()
    
    # Helper function to build flat-wise sheet (for Finishing checklists)
    def build_status_sheet(tmp_data: pd.DataFrame) -> pd.DataFrame:
        """Build status matrix for flat-level checklists (Finishing)."""
        if tmp_data.empty:
            return pd.DataFrame()
        
        tmp = tmp_data.copy()
        tmp["__ChecklistKey"] = tmp["__Checklist"].astype(str).str.strip().str.lower()
        tmp["__FlatKey"] = tmp["__Flat"].astype(str).fillna("").replace("", "UNKNOWN")
        tmp["__DateOrd"] = pd.to_datetime(tmp["__Date"]).fillna(pd.Timestamp.min)
        tmp = tmp.sort_values("__DateOrd").groupby(["__FlatKey", "__ChecklistKey"], as_index=False).tail(1)

        # Use standardized, human-readable checklist display names to avoid
        # accidental inclusion of flat identifiers in the column headers.
        tmp["__ChecklistDisplay"] = tmp["__Checklist"].astype(str).map(base_name)

        # Determine checklist columns as union across this floor (use readable names)
        checklist_cols = list(dict.fromkeys(tmp["__ChecklistDisplay"].astype(str).tolist()))

        # Build status matrix
        # Completion is based on stage progression, not EQC pass/fail status.
        def stage_symbol(stage: str) -> str:
            """Determine symbol + stage marker for color-specific flat mapping.
            Handles variant stage text like "Pre Pour", "During Check", etc.
            """
            s = (stage or "").strip().lower()
            if not s:
                return "✗"

            if "post" in s:
                return "✓"
            if "pre" in s:
                return "– PRE"
            if any(token in s for token in ("during", "reinforcement", "reinforc", "shuttering", "shutter")):
                return "– DURING"
            # Any non-empty unknown stage is treated as in-progress.
            return "– DURING"

        # Only include flats that actually have finishing checklist entries
        local_flats_set = list(dict.fromkeys(tmp["__FlatKey"].astype(str).tolist()))
        # Preserve natural ordering using outer _flat_key
        try:
            local_flats = sorted(local_flats_set, key=_flat_key)
        except Exception:
            local_flats = local_flats_set

        grid = []
        for flt in local_flats:
            row = {"Flat": flt}
            sub = tmp[tmp["__FlatKey"].eq(flt)]
            present = set(sub["__ChecklistDisplay"].astype(str))
            any_present = bool(present)
            all_complete = True
            for col in checklist_cols:
                rec = sub[sub["__ChecklistDisplay"].astype(str).eq(col)]
                if rec.empty:
                    row[col] = "✗"
                    all_complete = False
                else:
                    sym = stage_symbol(str(rec.iloc[0].get("__Stage", "")))
                    row[col] = sym
                    if sym != "✓":
                        all_complete = False
            # Overall status per flat
            row["Overall"] = "✓" if all_complete and any_present else ("–" if any_present else "✗")
            grid.append(row)

        return pd.DataFrame(grid, columns=["Flat", "Overall"] + checklist_cols) if grid else pd.DataFrame()
    
    # Separate data by checklist type
    # Combine RCC, Aluform, and Handover into one category (all floor-level)
    structural_data = dff[dff["__Checklist"].apply(lambda x: is_rcc_checklist(x) or is_aluform_checklist(x) or is_handover_checklist(x))]
    finishing_data = dff[~dff["__Checklist"].apply(lambda x: is_rcc_checklist(x) or is_aluform_checklist(x) or is_handover_checklist(x))]
    
    # Data validation: ensure no checklists are lost
    total_checklists = len(dff)
    structural_count = len(structural_data)
    finishing_count = len(finishing_data)
    data_validation_passed = (structural_count + finishing_count == total_checklists)
    
    # Debug logging
    print(f"[DEBUG] Total rows after filter: {total_checklists}")
    print(f"[DEBUG] Structural rows: {structural_count}")
    print(f"[DEBUG] Finishing rows: {finishing_count}")
    if total_checklists > 0:
        print(f"[DEBUG] Sample checklist names: {dff['__Checklist'].head(5).tolist()}")
    
    # Build status DataFrames for each category
    # Structural (RCC, Aluform, Handover combined) - floor-level (not flat-wise)
    structural_df = build_floor_status_sheet(structural_data)
    # Finishing is flat-wise
    finishing_df = build_status_sheet(finishing_data)

    # Write Excel with separate sheets
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import PatternFill, Font, Border, Side
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove the default sheet
    
    # Add Summary sheet first for data validation
    summary_ws = wb.create_sheet(title="Summary", index=0)
    summary_ws.cell(row=1, column=1, value="Data Export Summary")
    summary_ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    summary_ws.cell(row=3, column=1, value="Project:")
    summary_ws.cell(row=3, column=2, value=proj)
    summary_ws.cell(row=4, column=1, value="Building:")
    summary_ws.cell(row=4, column=2, value=building)
    summary_ws.cell(row=5, column=1, value="Floor:")
    summary_ws.cell(row=5, column=2, value=(floor or 'ALL'))
    summary_ws.cell(row=6, column=1, value="Export Date:")
    summary_ws.cell(row=6, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    
    summary_ws.cell(row=8, column=1, value="Data Validation:")
    summary_ws.cell(row=8, column=1).font = Font(bold=True)
    summary_ws.cell(row=9, column=1, value="Total Checklists:")
    summary_ws.cell(row=9, column=2, value=total_checklists)
    summary_ws.cell(row=10, column=1, value="Structural/Handover:")
    summary_ws.cell(row=10, column=2, value=structural_count)
    summary_ws.cell(row=11, column=1, value="Finishing:")
    summary_ws.cell(row=11, column=2, value=finishing_count)
    summary_ws.cell(row=12, column=1, value="Validation Status:")
    if total_checklists == 0:
        summary_ws.cell(row=12, column=2, value="⚠ WARNING: No data found for selected filters!")
        summary_ws.cell(row=12, column=2).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    elif data_validation_passed:
        summary_ws.cell(row=12, column=2, value="✓ All data captured (No missing checklists)")
        summary_ws.cell(row=12, column=2).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    else:
        summary_ws.cell(row=12, column=2, value=f"⚠ Warning: {total_checklists - structural_count - finishing_count} checklists may be missing!")
        summary_ws.cell(row=12, column=2).fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    
    summary_ws.cell(row=14, column=1, value="Sheet Organization:")
    summary_ws.cell(row=14, column=1).font = Font(bold=True)
    summary_ws.cell(row=15, column=1, value="• Structural & Handover:")
    summary_ws.cell(row=15, column=2, value="Floor-level checklists (RCC, Aluform, Internal Handover)")
    summary_ws.cell(row=16, column=1, value="• Finishing:")
    summary_ws.cell(row=16, column=2, value="Flat-wise matrix (Masonry, Plaster, Tiling, Painting, etc.)")
    
    # Apply borders to summary
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in summary_ws.iter_rows(min_row=1, max_row=16, min_col=1, max_col=2):
        for cell in row:
            cell.border = border
    
    # Auto-adjust column widths in summary
    summary_ws.column_dimensions['A'].width = 25
    summary_ws.column_dimensions['B'].width = 60
    
    # Helper to add floor-level sheet (for RCC, Aluform, Handover)
    def add_floor_level_sheet(workbook, sheet_name: str, data_df: pd.DataFrame) -> None:
        """Add a floor-level sheet (not flat-wise breakdown)."""
        if data_df.empty:
            # Create sheet even if empty to show that no data exists
            ws = workbook.create_sheet(title=sheet_name)
            ws.cell(row=1, column=1, value="Building:")
            ws.cell(row=1, column=2, value=building)
            ws.cell(row=1, column=4, value="Floor:")
            ws.cell(row=1, column=5, value=(floor or 'ALL'))
            ws.cell(row=3, column=1, value="No data found for this category on the selected floor.")
            return
        
        ws = workbook.create_sheet(title=sheet_name)
        # Add building name header for clarity
        ws.cell(row=1, column=1, value="Building:")
        ws.cell(row=1, column=2, value=building)
        ws.cell(row=1, column=4, value="Floor:")
        ws.cell(row=1, column=5, value=(floor or 'ALL'))
        
        # Add note explaining floor-level nature
        ws.cell(row=1, column=7, value="Note:")
        ws.cell(row=1, column=8, value="These checklists apply at Floor/Pour level, not per flat")
        
        # Add data count for validation
        ws.cell(row=1, column=11, value="Total Rows:")
        ws.cell(row=1, column=12, value=len(data_df))
        
        # Leave row 2 blank, start data at row 3
        ws.append([])
        start_row = 3
        for r in dataframe_to_rows(data_df, index=False, header=True):
            ws.append(r)
        
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        red = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
        gray = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        # Apply coloring to Status column only (5th column)
        status_col_idx = 5  # Column E (Category, Checklist, Pour, Stage, Status, Date)
        for row in ws.iter_rows(min_row=start_row+1, min_col=status_col_idx, max_row=ws.max_row, max_col=status_col_idx):
            for cell in row:
                val = str(cell.value or "")
                if "✓" in val or "Passed" in val:
                    cell.fill = green
                elif "–" in val or "Progress" in val or "Pending" in val:
                    cell.fill = yellow
                elif "✗" in val or "Failed" in val:
                    cell.fill = red
                elif "↻" in val or "Redo" in val:
                    cell.fill = gray
        
        # Apply borders and wrap text
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        align = Alignment(wrap_text=True, vertical="center")
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                try:
                    cell.border = border
                    cell.alignment = align
                except Exception:
                    pass
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
    
    # Helper to add flat-level sheet (for Finishing)
    def add_flat_level_sheet(workbook, sheet_name: str, data_df: pd.DataFrame) -> None:
        """Add a flat-wise sheet (for Finishing checklists)."""
        if data_df.empty:
            # Create sheet even if empty to show that no data exists
            ws = workbook.create_sheet(title=sheet_name)
            ws.cell(row=1, column=1, value="Building:")
            ws.cell(row=1, column=2, value=building)
            ws.cell(row=1, column=4, value="Floor:")
            ws.cell(row=1, column=5, value=(floor or 'ALL'))
            ws.cell(row=3, column=1, value="No data found for this category on the selected floor.")
            return
        
        ws = workbook.create_sheet(title=sheet_name)
        # Add building name header for clarity
        ws.cell(row=1, column=1, value="Building:")
        ws.cell(row=1, column=2, value=building)
        ws.cell(row=1, column=4, value="Floor:")
        ws.cell(row=1, column=5, value=(floor or 'ALL'))
        
        # Add legend for symbols
        ws.cell(row=1, column=7, value="Legend:")
        ws.cell(row=1, column=8, value="– PRE = In Progress (White)")
        ws.cell(row=1, column=9, value="– DURING = In Progress (Yellow)")
        ws.cell(row=1, column=10, value="✓ = Completed (Green), ✗ = Not Present (Red)")
        
        # Add data count for validation (number of flats shown)
        ws.cell(row=1, column=12, value="Flats:")
        ws.cell(row=1, column=13, value=len(data_df))
        
        # Leave row 2 blank, start data at row 3
        ws.append([])
        start_row = 3
        for r in dataframe_to_rows(data_df, index=False, header=True):
            ws.append(r)
        
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        red = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
        
        # Apply coloring only over data rows (excluding header at start_row)
        for row in ws.iter_rows(min_row=start_row+1, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                val = str(cell.value or "")
                if val == "✓":
                    cell.fill = green
                elif "PRE" in val.upper():
                    cell.fill = white
                    cell.font = Font(color="000000", bold=False)
                elif "DURING" in val.upper() or val == "–":
                    cell.fill = yellow
                    cell.font = Font(color="000000", bold=False)
                elif val == "✗":
                    cell.fill = red
                    cell.font = Font(color="000000", bold=True)
        
        # Apply borders and wrap text
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        align = Alignment(wrap_text=True, vertical="center")
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                try:
                    cell.border = border
                    cell.alignment = align
                except Exception:
                    pass
    
    # Add sheets in order
    # Structural (RCC, Aluform, Handover combined) - floor-level (not flat-wise)
    add_floor_level_sheet(wb, "Structural & Handover", structural_df)
    # Finishing is flat-wise
    add_flat_level_sheet(wb, "Finishing", finishing_df)
    
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    
    # Sanitize building name for filename (extract key identifier)
    import re as _re_fname
    building_clean = building or "All"
    # Try to extract just the letter/number (e.g., "A wing" -> "A", "Building B" -> "B")
    m = _re_fname.search(r'\b([A-Z]\d?|\d+[A-Z]?)\b', building_clean)
    if m:
        building_clean = m.group(1)
    else:
        # Just remove spaces and special chars
        building_clean = _re_fname.sub(r'[^\w]', '_', building_clean)
    
    floor_clean = (floor or 'ALL').replace(" ", "_")
    fname = f"Flat_Status_{proj}_{building_clean}_{floor_clean}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    # Remove any remaining problematic characters
    fname = fname.replace("/", "-").replace("\\", "-")
    
    return send_file(bio, as_attachment=True, download_name=fname, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/")
def root_redirect():
    # Convenience: redirect base URL to the flat report page
    return redirect(url_for("flat_report_index"))


@app.get("/favicon.ico")
def favicon_empty():
    # Avoid noisy 404s for favicon during local testing
    from flask import Response
    return Response(status=204)


if __name__ == "__main__":
    # Local run only
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5050)), debug=True)
