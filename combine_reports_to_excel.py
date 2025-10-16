import glob
import os
import pandas as pd


def find_mode_files(mode: str):
    pattern = f"*-{mode}-digiQC-report_EQC_Summary_WithTotals_Building-wise.csv"
    return sorted(glob.glob(pattern))


def detect_project_from_filename(fname: str, mode: str) -> str:
    base = os.path.basename(fname)
    # Expect formats like: FUTURA-Weekly-... or Itrend-Palacio-Weekly-...
    # Split at '-{mode}-' to preserve hyphens within project names
    sep = f"-{mode}-"
    if sep in base:
        proj = base.split(sep, 1)[0]
    else:
        # Fallback: remove trailing parts after first occurrence of mode keyword
        idx = base.lower().find(mode.lower())
        proj = base[:idx].rstrip('-') if idx > 0 else base
    # Normalize underscores to spaces where relevant
    proj = proj.replace('_', ' ')
    # Canonicalize common variants to match app & weekly script
    key = proj.strip().lower().replace('_', ' ')
    mappings = {
        "city life": "Itrend City Life",
        "itrend city life": "Itrend City Life",
        "futura": "Itrend Futura",
        "itrend futura": "Itrend Futura",
        "itrend-palacio": "Itrend Palacio",
        "itrend palacio": "Itrend Palacio",
    }
    return mappings.get(key, proj.strip())


def coerce_numeric_df(df: pd.DataFrame) -> pd.DataFrame:
    """Convert columns that contain only numeric-like values into integer dtype.

    Rules:
    - Skip obvious text columns like 'Checklists' or 'Building'.
    - For other columns, strip commas/spaces and attempt to convert to numeric.
    - If all non-empty cells convert to numeric, cast column to int (fill empty as 0).
    """
    import pandas as _pd
    out = df.copy()
    text_cols = {c.lower() for c in ('checklists', 'building', 'eqc checklist')}
    for col in out.columns:
        if str(col).strip().lower() in text_cols:
            continue
        try:
            s = out[col].astype(str).str.replace(',', '').str.strip()
        except Exception:
            continue
        non_empty_mask = s.replace('', _pd.NA).notna()
        non_empty_count = int(non_empty_mask.sum())
        if non_empty_count == 0:
            continue
        coerced = _pd.to_numeric(s.replace('', _pd.NA), errors='coerce')
        numeric_count = int(coerced.notna().sum())
        if numeric_count == non_empty_count:
            out[col] = coerced.fillna(0).astype(int)
    return out


def load_file(fpath: str) -> pd.DataFrame | None:
    try:
        df = pd.read_csv(fpath, dtype=str)
        return coerce_numeric_df(df)
    except Exception as e:
        print(f"Failed to read {fpath}: {e}")
        return None


def main():
    # Discover all available (project, mode) from files
    modes = ['Weekly', 'Monthly', 'Cumulative']
    files_by_mode = {mode: find_mode_files(mode) for mode in modes}
    pairs: list[tuple[str, str]] = []
    paths: dict[tuple[str, str], str] = {}
    for mode, files in files_by_mode.items():
        for f in files:
            proj = detect_project_from_filename(f, mode)
            # Skip DEMO projects
            if 'demo' in proj.lower():
                continue
            key = (proj, mode)
            paths[key] = f
            pairs.append(key)
    # Stable sheet order: group by project name, then mode order Weekly, Monthly, Cumulative
    project_names = sorted({p for p, _ in pairs})
    sheet_order = []
    for p in project_names:
        for m in modes:
            if (p, m) in paths:
                sheet_order.append((p, m))
    out_xlsx = 'EQC_Weekly_Monthly_Cumulative_AllProjects.xlsx'

    # paths already built above

    with pd.ExcelWriter(out_xlsx, engine='openpyxl') as writer:
        # Index sheet
        index_rows = []
        for proj, mode in sheet_order:
            sheet_name = f"{proj} {mode} report"
            index_rows.append({'Sheet': sheet_name})
        pd.DataFrame(index_rows).to_excel(writer, sheet_name='Summary or Index', index=False)

        any_written = False
        for proj, mode in sheet_order:
            fpath = paths.get((proj, mode))
            if not fpath:
                print(f"Missing file for {proj} {mode}; skipping sheet")
                continue
            df = load_file(fpath)
            if df is None or df.empty:
                print(f"Empty or unreadable data for {proj} {mode}; skipping sheet")
                continue
            sheet_name = f"{proj} {mode} report"
            sheet_name = sheet_name[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"Wrote sheet {sheet_name} with {len(df)} rows (from {os.path.basename(fpath)})")
            any_written = True

        # Apply formatting: borders and wrap text on all data sheets and index
        from openpyxl.styles import Border, Side, Alignment
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for ws_name in ['Summary or Index'] + [f"{proj} {mode} report"[:31] for proj, mode in sheet_order]:
            ws = writer.sheets.get(ws_name)
            if not ws:
                continue
            max_row = ws.max_row
            max_col = ws.max_column
            for r in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                for cell in r:
                    try:
                        cell.border = border
                        cell.alignment = Alignment(wrap_text=True, vertical='center')
                    except Exception:
                        pass
            # Auto-fit basic column widths
            try:
                for col in ws.columns:
                    max_len = 0
                    col_letter = getattr(col[0], 'column_letter', None)
                    for cell in col:
                        try:
                            val = cell.value
                            if val is None:
                                continue
                            l = len(str(val))
                            if l > max_len:
                                max_len = l
                        except Exception:
                            continue
                    if col_letter:
                        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 60)
            except Exception:
                pass

    if any_written:
        print(f"✅ Combined workbook written to {out_xlsx}")
    else:
        print("No sheets were written; nothing to combine.")


if __name__ == '__main__':
    main()
