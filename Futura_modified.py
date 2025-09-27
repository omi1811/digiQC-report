import os
import re
import pandas as pd
from typing import Tuple
import argparse

# --- CLI args / site name ---
parser = argparse.ArgumentParser(description="EQC report generator")
parser.add_argument("--input", "-i", default="Combined_EQC.csv", help="Input EQC CSV file")
args = parser.parse_args()
eqc_file = args.input

# site_name will be derived from the data (Location L0) after we load the file
site_name = None

# --- Load data ---
# The source `abc.csv` is tab-separated. Specify sep='\t' to avoid tokenization errors
df = pd.read_csv(eqc_file, dtype=str, keep_default_na=False, sep='\t')

# Derive site name from Location L0 (first non-empty value). Fallback to cwd or 'site'
loc0 = df.get('Location L0', pd.Series(dtype=str)).astype(str).str.strip()
sites = [s for s in loc0.unique() if s and s.strip()]
if sites:
    site_name = str(sites[0]).replace(' ', '_')
else:
    site_name = os.getenv('SITE') or os.path.basename(os.getcwd()) or 'site'

# Final filenames
wide_filename = f"{site_name}-digiQC-report_EQC_Summary_WithTotals_Building-wise.csv"
alerts_filename = f"{site_name}-digiQC-report_EQC_Alerts.csv"

# Drop unwanted columns if present
drop_cols = ["Approver", "Approved Timestamp", "Time Zone",
             "Team", "Total EQC Stages", "Fail Stages", "% Fail", "URL"]
df = df.drop(columns=[c for c in drop_cols if c in df.columns])

# Convert date (source uses DD-MM-YYYY)
df["Date"] = pd.to_datetime(df.get("Date", ""), dayfirst=True, errors="coerce")

# --- Normalize Stage ---
def normalize_stage(stage: str) -> str:
    s = str(stage or "").lower()
    if "pre" in s:
        return "Pre"
    if "during" in s:
        return "During"
    if "post" in s:
        return "Post"
    if "reinforce" in s:
        return "Reinforcement"
    if "shutter" in s:
        return "Shuttering"
    return "Other"

df["Stage_Norm"] = df.get("Stage", "").apply(normalize_stage)

# --- Define custom stage order ---
stage_order = ["Pre", "During", "Post", "Reinforcement", "Shuttering", "Other"]
df["Stage_Norm"] = pd.Categorical(df["Stage_Norm"], categories=stage_order, ordered=True)

# --- Pivot table (Eqc Type as index, Building + Stage as columns) ---
# Use groupby + unstack to avoid pandas grouper errors
summary = (
    df.groupby(["Eqc Type", "Location L1", "Stage_Norm"])['Eqc Type']
    .count()
    .unstack(level=[1, 2], fill_value=0)
)

# --- Add per-building totals (Pre+During+Post only) ---
for bldg in df.get("Location L1", pd.Series(dtype=str)).unique():
    if not bldg:
        continue
    try:
        if all(((bldg, stage) in summary.columns) for stage in ["Pre", "During", "Post"]):
            summary[(bldg, "Total")] = (
                summary[(bldg, "Pre")] +
                summary[(bldg, "During")] +
                summary[(bldg, "Post")]
            )
    except Exception:
        continue

# --- Add overall totals across buildings ---
totals = (
    df.groupby(["Eqc Type", "Stage_Norm"])['Eqc Type'].count().unstack(fill_value=0)
)
totals.columns = pd.MultiIndex.from_product([["Total"], totals.columns.astype(str)])
summary = pd.concat([summary, totals], axis=1)

summary.index.name = "EQC Checklist"

# --- Add Status summary at the end ---
if "Status" in df.columns:
    status_summary = df.groupby(["Eqc Type", "Status"])['Eqc Type'].count().unstack(fill_value=0)
    if not status_summary.empty:
        status_summary.columns = pd.MultiIndex.from_product([["Status"], status_summary.columns.astype(str)])
        summary = pd.concat([summary, status_summary], axis=1)

# --- Alerts ---
alerts = []
today = pd.Timestamp.today()

# 1. REDO > 3 days — include Location L2 (floor) and L3 (flat) for context
# In our data, REDO appears in the generic 'Status' column; ensure alignment with df.index
redo_col = df.get("Status", pd.Series(["" for _ in range(len(df))], index=df.index)).astype(str).str.upper()
redo_df = df[redo_col == "REDO"]
if not redo_df.empty:
    # group by Eqc Type + Location L1/L2/L3 so we can include floor/flat in alerts
    max_dates = redo_df.groupby(["Eqc Type", "Location L1", "Location L2", "Location L3"])["Date"].max()
    for keys, last_date in max_dates.items():
        eqc = keys[0]
        loc_parts = [k for k in keys[1:] if k and str(k).strip()]
        loc_label = " / ".join(loc_parts) if loc_parts else (keys[1] or "UNKNOWN")
        if pd.notna(last_date) and (today - last_date).days > 3:
            alerts.append(f"ALERT: {eqc} in {loc_label} is REDO for more than 3 days (last: {last_date.date()})")

# 2. IN_PROGRESS > 3 weeks without stage change (include floor/flat)
status_col = df.get("Status", pd.Series(["" for _ in range(len(df))], index=df.index)).astype(str).str.upper()
inprog_df = df[status_col == "IN_PROGRESS"]
if not inprog_df.empty:
    # group by Eqc Type + Location L1/L2/L3 + Stage_Norm
    min_dates = inprog_df.groupby(["Eqc Type", "Location L1", "Location L2", "Location L3", "Stage_Norm"])["Date"].min()
    for keys, first_date in min_dates.items():
        eqc = keys[0]
        stage = keys[4]
        loc_parts = [k for k in keys[1:4] if k and str(k).strip()]
        loc_label = " / ".join(loc_parts) if loc_parts else (keys[1] or "UNKNOWN")
        if pd.notna(first_date) and (today - first_date).days > 21:
            alerts.append(f"ALERT: {eqc} in {loc_label} stuck in {stage} for more than 3 weeks (since: {first_date.date()})")

# Write alerts file (only alerts output requested)
pd.DataFrame(alerts, columns=["Alerts"]).to_csv(alerts_filename, index=False)
print(f"✅ Alerts written to {alerts_filename}")

# --- Save outputs ---
# (No intermediate files) We'll produce only the wide building-wise summary and alerts at the end.
print("Raw entries:", len(df))
print("Summary total count:", summary.select_dtypes(include=["number"]).sum().sum())

# Calculate totals per building
# Ensure columns are tuples
summary.columns = pd.MultiIndex.from_tuples([c if isinstance(c, tuple) else ("", str(c)) for c in summary.columns])
building_totals = {}
for bldg in sorted({c[0] for c in summary.columns if c[0] and c[0] not in ("Total", "Status")}):
    cols = [c for c in summary.columns if c[0] == bldg and c[1] != "Total"]
    if cols:
        building_totals[(bldg, "Total")] = summary.loc[:, cols].sum(axis=1)

if building_totals:
    bt_df = pd.concat(building_totals, axis=1)
    summary = pd.concat([summary, bt_df], axis=1)

# Calculate grand total row (numeric only)
numeric_sum = summary.select_dtypes(include=["number"]).sum()
if not numeric_sum.empty:
    grand_total = numeric_sum.to_frame().T
    grand_total.index = ["TOTAL"]
    summary_final = pd.concat([summary, grand_total], axis=0, sort=False)
else:
    summary_final = summary.copy()

# Sort columns: building, stage order, Totals at end
stage_rank = {s: i for i, s in enumerate(stage_order)}
def sort_key(col: Tuple[str, str]):
    building, stage = col
    if building == "Total":
        return (2, "ZZZ", stage_rank.get(stage, 99), str(stage))
    if building == "":
        return (3, "ZZZ", 99, str(stage))
    return (1, building, stage_rank.get(stage, 99), str(stage))

# If there are duplicate MultiIndex columns, combine them by summing (avoids non-unique MultiIndex)
if isinstance(summary_final.columns, pd.MultiIndex) and not summary_final.columns.is_unique:
    summary_final = summary_final.groupby(level=list(range(summary_final.columns.nlevels)), axis=1).sum()

sorted_cols = sorted(summary_final.columns.tolist(), key=sort_key)
summary_final = summary_final.reindex(columns=sorted_cols)

# --- Save outputs ---
summary_final_reset = summary_final.reset_index()
# do not write intermediate combined CSV; we will produce only the wide building-wise report and alerts
print("✅ Computed combined summary in memory; writing only building-wise wide report next")

# --- Also write EQC_Summary_WithTotals.csv in the flattened tuple-string header format (reference style) ---
def flatten_col(col):
    # col may be tuple (building, stage) or a plain string
    if isinstance(col, tuple):
        return f"('{col[0]}', '{col[1]}')"
    return str(col)

# Work with a copy
out_df = summary_final.copy()

# Reset index to get 'EQC Checklist' as a column
out_df_reset = out_df.reset_index()

# Flatten column names
flat_cols = []
for col in out_df_reset.columns:
    if col == out_df_reset.columns[0]:
        # the index column name (likely 'EQC Checklist') keep as-is
        flat_cols.append(str(col))
    else:
        flat_cols.append(flatten_col(col))

out_df_reset.columns = flat_cols

# Ensure IN_PROGRESS and PASSED columns exist (they may come from ('Status','IN_PROGRESS'))
status_in_prog = None
status_passed = None
for c in out_df_reset.columns:
    if "('Status', 'IN_PROGRESS')" == c:
        status_in_prog = c
    if "('Status', 'PASSED')" == c or "('Status', 'PASS')" == c:
        status_passed = c

# Create clean IN_PROGRESS / PASSED columns (unprefixed) if present
if status_in_prog:
    out_df_reset['IN_PROGRESS'] = out_df_reset[status_in_prog]
if status_passed:
    out_df_reset['PASSED'] = out_df_reset[status_passed]

# Compute Total Count as sum across columns where header starts with "('Total',"
total_cols = [c for c in out_df_reset.columns if c.startswith("('Total',")]
if total_cols:
    out_df_reset['Total Count'] = out_df_reset[total_cols].sum(axis=1)
else:
    # fallback: sum numeric columns
    out_df_reset['Total Count'] = out_df_reset.select_dtypes(include=['number']).sum(axis=1)

# Reorder columns: make the EQC Checklist (index column) the leftmost, then building-stage columns,
# then IN_PROGRESS / PASSED (if present), and finally Total Count.
eqc_col = str(out_df_reset.columns[0])
building_stage_cols = [c for c in out_df_reset.columns if isinstance(c, str) and c.startswith("('")]

ordered = [eqc_col] + building_stage_cols
if 'IN_PROGRESS' in out_df_reset.columns:
    ordered += ['IN_PROGRESS']
if 'PASSED' in out_df_reset.columns:
    ordered += ['PASSED']
ordered += ['Total Count']

# Ensure uniqueness while preserving order and only keep columns that exist
seen = set()
final_cols = [c for c in ordered if c in out_df_reset.columns and (c not in seen and not seen.add(c))]

out_df_final = out_df_reset[final_cols]

# Append stage-total rows (one row per stage with EQC Checklist set to stage name and Total Count set to total)
grand = None
if 'TOTAL' in summary_final.index:
    grand = summary_final.loc['TOTAL']
else:
    grand = summary_final.sum()

stage_rows = []
for stage in stage_order:
    # sum across all columns where second element == stage
    s = 0
    for col in summary_final.columns:
        try:
            if isinstance(col, tuple) and str(col[1]) == str(stage):
                s += grand.get(col, 0)
        except Exception:
            continue
    row = {c: '' for c in out_df_final.columns}
    # put stage name in EQC Checklist column and value in Total Count
    row[eqc_col] = stage
    row['Total Count'] = s
    stage_rows.append(row)

if stage_rows:
    out_df_final = pd.concat([out_df_final, pd.DataFrame(stage_rows)], ignore_index=True, sort=False)

# Save
    # Do not write the flattened reference file; we'll output the wide building-wise CSV instead.

# --- Also create a tidy by-building CSV where 'Building' is a proper column and
# similar contractor-specific checklist names are collapsed into a base name
# (split on the first '.') and summed per building. Back up existing file first.
try:
    src_path = wide_filename
    bak_path = None
    # back up user-edited file if present
    if os.path.exists(src_path) and bak_path:
        # Only attempt to replace if we have a valid backup path
        os.replace(src_path, bak_path)

    # Build a long-form DataFrame from summary_final
    # collect building names from raw data; treat empty Location L1 as 'UNKNOWN' so no rows are dropped
    raw_buildings = list(df.get('Location L1', pd.Series(dtype=str)).unique())
    buildings = sorted({(b if b else 'UNKNOWN') for b in raw_buildings})

    rows = []
    for bldg in buildings:
        # gather columns for this building for each stage
        cols = [(b, s) for (b, s) in summary_final.columns if b == bldg and s in stage_order]
        # ensure we have all stages
        for eqc in summary_final.index:
            if str(eqc).upper() == 'TOTAL':
                continue
            row = {"EQC Checklist": eqc, "Building": bldg}
            for s in stage_order:
                val = 0
                # lookup column: in summary_final the empty building may be stored as '' so map 'UNKNOWN' -> ''
                lookup_building = '' if bldg == 'UNKNOWN' else bldg
                lookup_col = (lookup_building, s)
                if lookup_col in summary_final.columns:
                    try:
                        val = int(summary_final.at[eqc, lookup_col])
                    except Exception:
                        try:
                            val = int(summary_final.loc[eqc, lookup_col])
                        except Exception:
                            val = 0
                else:
                    val = 0
                row[s] = val
            row['Total Count'] = sum(row[s] for s in stage_order)
            rows.append(row)

    long_df = pd.DataFrame(rows)
    if long_df.empty:
        print('No per-building rows to write.')
    else:
        # Collapse contractor-specific names into a base name (split on first '.')
        # smarter base-name extraction: prefer text before '.' or ':'; otherwise
        # group names sharing first-2 tokens and return their longest common token prefix
        def longest_common_prefix_tokens(names):
            if not names:
                return ''
            token_lists = [str(n).split() for n in names]
            prefix = []
            for tokens in zip(*token_lists):
                if all(t == tokens[0] for t in tokens):
                    prefix.append(tokens[0])
                else:
                    break
            return ' '.join(prefix).strip()

        def base_name(n: str) -> str:
            if not isinstance(n, str):
                return str(n)
            s = n.strip()
            # Normalize known variants to canonical checklist names
            s_norm = s.lower()
            # quick pattern fixes for commonly miswritten names
            fixes = {
                r'painting.*internal': 'Painting Works : Internal',
                r'painting.*external': 'Painting Works : External',
                r'painting\s*works\s*:\s*internal': 'Painting Works : Internal',
                r'painting\s*works\s*:\s*external': 'Painting Works : External',
                r'waterproof.*boxtype': 'Waterproofing works: Toilet and Skirting',
                r'waterproof.*toilet': 'Waterproofing works: Toilet and Skirting',
                r'waterproof.*skirting': 'Waterproofing works: Toilet and Skirting',
                r'tiling.*kitchen.*platform': 'Tiling - Kitchen Platform',
                r'tiling.*kitchen.*sink': 'Tiling - Kitchen Platform',
                r'tiling[-\s]*toilet.*dado': 'Tiling - Toilet Dado',
                r'kitchen\s*dado': 'Kitchen Dado Checklist',
            }
            for pat, name in fixes.items():
                if re.search(pat, s_norm):
                    return name
            # if dot or colon present, prefer text before it (common vendor separator)
            if '.' in s:
                return s.split('.', 1)[0].strip()
            if ':' in s:
                return s.split(':', 1)[0].strip()
            # fallback: find other names that share first two tokens
            tokens = s.split()
            if len(tokens) >= 2:
                key = ' '.join(tokens[:2])
                same_prefix = [x for x in summary_final.index.astype(str) if x.startswith(key)]
                if len(same_prefix) > 1:
                    lcp = longest_common_prefix_tokens(same_prefix)
                    return lcp if lcp else key
            # if no grouping, return up to first 4 tokens to avoid very long names
            return ' '.join(tokens[:4]).strip()

    # Use 'Checklists' as the canonical column name (user requested)
    long_df['Checklists'] = long_df['EQC Checklist'].apply(base_name)

    # Aggregate by Checklists + Building
    agg = long_df.groupby(['Checklists', 'Building'])[stage_order].sum().reset_index()

    # Ensure numeric and fill missing
    for s in stage_order:
        agg[s] = pd.to_numeric(agg[s], errors='coerce').fillna(0).astype(int)

    # Apply cumulative logic: Pre = Pre + During + Post; During = During + Post
    if all(col in agg.columns for col in ['Pre', 'During', 'Post']):
        agg['Pre'] = agg['Pre'] + agg['During'] + agg['Post']
        agg['During'] = agg['During'] + agg['Post']

    # Compute Total Count after cumulative transformation
    agg['Total Count'] = agg[stage_order].sum(axis=1)

    # Also add an overall Checklists total (across buildings)
    overall = agg.groupby('Checklists')[stage_order].sum().reset_index()
    overall['Building'] = 'ALL'
    overall['Total Count'] = overall[stage_order].sum(axis=1)

    final_by_building = pd.concat([agg, overall], ignore_index=True, sort=False)

    # Reorder columns
    cols_order = ['Checklists', 'Building'] + stage_order + ['Total Count']
    final_by_building = final_by_building[cols_order]

    # We will not write the intermediate tidy CSV; instead produce a wide building-wise CSV below
    print("✅ Prepared tidy per-building aggregation (in-memory)")
    # Also produce a wide-format CSV with buildings as top-level columns and stages as sub-columns
    try:
        # final_by_building has columns: Checklists, Building, Pre, During, ...
        wide_idx = final_by_building.set_index(['Checklists', 'Building'])[stage_order]
        # unstack buildings -> columns (stage, building)
        wide_un = wide_idx.unstack(level='Building', fill_value=0)
        # swap to (building, stage)
        wide_un.columns = wide_un.columns.swaplevel(0, 1)

        # Add per-building 'Total' sub-column (sum of stages) for every building present
        buildings_set = sorted({c[0] for c in wide_un.columns})
        for b in buildings_set:
            # compute sum across available stages for this building
            stage_cols = [(b, s) for s in stage_order if (b, s) in wide_un.columns]
            if stage_cols:
                wide_un[(b, 'Total')] = wide_un[stage_cols].sum(axis=1)

        # reorder buildings alphabetically but keep 'ALL' at the end if present
        buildings = sorted({c[0] for c in wide_un.columns})
        if 'ALL' in buildings:
            buildings.remove('ALL')
            buildings.append('ALL')

        # reorder columns by buildings and include Total as last stage per building
        ordered_cols = []
        for b in buildings:
            for s in stage_order + ['Total']:
                if (b, s) in wide_un.columns:
                    ordered_cols.append((b, s))

        wide_un = wide_un.reindex(columns=pd.MultiIndex.from_tuples(ordered_cols))

        # Append grand TOTAL row (sum across BaseName rows) so no data is missing
        wide_un.loc['TOTAL'] = wide_un.sum(axis=0)

        # Before writing, remove any per-building 'Total' sub-columns (we want totals only as the final row)
        try:
            if isinstance(wide_un.columns, pd.MultiIndex):
                cols_to_keep = [c for c in wide_un.columns if str(c[1]) != 'Total']
                wide_un = wide_un.loc[:, cols_to_keep]
            else:
                # If columns are flat, drop a plain 'Total' column if present
                if 'Total' in wide_un.columns:
                    wide_un = wide_un.drop(columns=['Total'])
        except Exception:
            # be conservative: if anything goes wrong, keep the original wide_un
            pass

        # Reset index so Checklists is the leftmost column and write CSV (MultiIndex header preserved)
        wide_un_reset = wide_un.reset_index()

        # Remove rows where Checklists is empty or is just a serial number
        if 'Checklists' in wide_un_reset.columns:
            chk = wide_un_reset['Checklists'].astype(str).str.strip()
            mask = (~chk.isna()) & (chk != '') & (~chk.str.match(r'^\d+$'))
            wide_un_reset = wide_un_reset[mask]

        wide_un_reset.to_csv(wide_filename, index=False)
        print(f"✅ Wrote wide-format per-building summary to {wide_filename}")

        # Also write a formatted Excel file (.xlsx) for management-ready view
        xlsx_filename = f"{os.path.splitext(wide_filename)[0]}.xlsx"
        try:
            # Use openpyxl to create formatting (merged top header cells and borders)
            import openpyxl
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Border, Side, Alignment, Font

            with pd.ExcelWriter(xlsx_filename, engine='openpyxl') as writer:
                # write the DataFrame starting 3 rows down so we have 3 blank rows above
                startrow = 3
                # wide_un_reset already has the 'Checklists' leftmost column
                # write with index=True so pandas preserves MultiIndex headers correctly
                # For Excel, flatten MultiIndex columns into single row labels (e.g. Building - Stage)
                excel_df = wide_un_reset.copy()
                # If original wide_un had MultiIndex columns, create flattened labels
                if isinstance(wide_un.columns, pd.MultiIndex):
                    new_cols = []
                    for c in wide_un_reset.columns:
                        if isinstance(c, tuple):
                            new_cols.append(f"{c[0]} - {c[1]}")
                        else:
                            new_cols.append(str(c))
                    excel_df.columns = new_cols
                # write the flattened DataFrame to Excel
                excel_df.to_excel(writer, sheet_name='Summary', startrow=startrow, index=False)
                ws = writer.sheets['Summary']

                # compute header rows and table bounds
                n_header_rows = getattr(wide_un.columns, 'nlevels', 1)
                # pandas writes MultiIndex headers starting at startrow (0-indexed), Excel rows are 1-indexed
                # when index=True pandas inserts index name cells; header_top should account for that
                header_top = startrow + 1
                header_bottom = header_top + n_header_rows - 1
                max_row = ws.max_row
                max_col = ws.max_column

                # Write site name centered in the blank area above the table (row 1)
                try:
                    site_cell_row = 1
                    ws.merge_cells(start_row=site_cell_row, start_column=1, end_row=site_cell_row, end_column=max_col)
                    cell = ws.cell(row=site_cell_row, column=1)
                    cell.value = site_name
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(bold=True)
                except Exception:
                    pass

                # Write current date at top-right (row 2, last column)
                from datetime import date
                try:
                    date_cell_row = 2
                    date_str = date.today().isoformat()
                    ws.cell(row=date_cell_row, column=max_col, value=date_str)
                except Exception:
                    pass

                # Compute overall Pre/During/Post totals from the TOTAL row in wide_un (if present)
                total_pre = total_during = total_post = 0
                try:
                    if 'TOTAL' in wide_un.index:
                        total_row = wide_un.loc['TOTAL']
                    else:
                        total_row = wide_un.sum(axis=0)
                    # total_row is a Series indexed by MultiIndex columns
                    for col in wide_un.columns:
                        if str(col[1]).lower().startswith('pre'):
                            total_pre += int(total_row.get(col, 0))
                        elif str(col[1]).lower().startswith('during'):
                            total_during += int(total_row.get(col, 0))
                        elif str(col[1]).lower().startswith('post'):
                            total_post += int(total_row.get(col, 0))
                except Exception:
                    pass

                total_sum = total_pre + total_during + total_post
                completed_pct = (total_post / total_sum * 100) if total_sum else 0
                inprog_pct = (total_during / total_sum * 100) if total_sum else 0

                # Place percentage info below the date cell (rows 3 and 4, rightmost columns)
                try:
                    pct_label_col = max_col - 1 if max_col >= 2 else max_col
                    ws.cell(row=3, column=pct_label_col, value='Completed %')
                    ws.cell(row=3, column=max_col, value=f"{completed_pct:.1f}%")
                    ws.cell(row=4, column=pct_label_col, value='In Progress %')
                    ws.cell(row=4, column=max_col, value=f"{inprog_pct:.1f}%")
                    # Add thin border to these 2x2 cells
                    thin = Side(border_style="thin", color="000000")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    for rr in (3, 4):
                        for cc in (pct_label_col, max_col):
                            try:
                                ws.cell(row=rr, column=cc).border = border
                            except Exception:
                                pass
                except Exception:
                    pass

                # Merge top-level building headers by parsing flattened labels (format: 'Building - Stage')
                try:
                    cols = list(excel_df.columns)
                    # Excel columns are 1-indexed; leftmost data column is column 1
                    left_offset = 1
                    # ensure the leftmost header cell shows 'Checklists' so it's obvious in the Excel view
                    try:
                        chk_cell = ws.cell(row=header_top, column=1)
                        if not chk_cell.value:
                            chk_cell.value = 'Checklists'
                            chk_cell.font = Font(bold=True)
                            chk_cell.alignment = Alignment(horizontal='center', vertical='center')
                    except Exception:
                        pass

                    run_start = None
                    run_value = None
                    run_end = None
                    for i in range(len(cols)):
                        label = cols[i]
                        # parse building from 'Building - Stage' format; fallback to full label
                        parts = str(label).split(' - ', 1)
                        top = parts[0] if len(parts) > 1 else label
                        excel_col = left_offset + i
                        if run_value is None:
                            run_value = top
                            run_start = excel_col
                            run_end = excel_col
                        elif top == run_value:
                            run_end = excel_col
                        else:
                            if run_end > run_start:
                                ws.merge_cells(start_row=header_top, start_column=run_start, end_row=header_top, end_column=run_end)
                                cell = ws.cell(row=header_top, column=run_start)
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                                cell.font = Font(bold=True)
                            run_value = top
                            run_start = excel_col
                            run_end = excel_col
                    if run_value is not None and run_start is not None and run_end > run_start:
                        ws.merge_cells(start_row=header_top, start_column=run_start, end_row=header_top, end_column=run_end)
                        cell = ws.cell(row=header_top, column=run_start)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font = Font(bold=True)
                except Exception:
                    pass

                # Apply thin border around all used cells to make sheet management-ready
                try:
                    thin = Side(border_style="thin", color="000000")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    for r in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                        for cell in r:
                            try:
                                cell.border = border
                            except Exception:
                                pass
                except Exception:
                    pass

                # Highlight the TOTAL row (look for 'TOTAL' in the first column)
                try:
                    from openpyxl.styles import PatternFill
                    highlight = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    for row in range(header_bottom + 1, max_row + 1):
                        try:
                            val = ws.cell(row=row, column=1).value
                            if isinstance(val, str) and val.strip().upper() == 'TOTAL':
                                for c in range(1, max_col + 1):
                                    cell = ws.cell(row=row, column=c)
                                    try:
                                        cell.fill = highlight
                                        cell.font = Font(bold=True)
                                    except Exception:
                                        pass
                                break
                        except Exception:
                            continue
                except Exception:
                    pass
            print(f"✅ Wrote formatted Excel to {xlsx_filename}")
        except Exception as e:
            print('Failed to write formatted Excel:', str(e))
    except Exception as e:
        print('Failed to write wide-format per-building summary:', str(e))
except Exception as e:
    print('Failed to write per-building collapsed summary:', str(e))
