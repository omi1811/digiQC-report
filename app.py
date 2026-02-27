#!/usr/bin/env python3
from __future__ import annotations

import io
import os
import sys
import tempfile
import uuid
from datetime import date, datetime
from typing import Dict, List, Tuple

from flask import Flask, render_template, request, send_file, redirect, url_for, session, send_from_directory
from flask import jsonify
import pandas as pd
import analysis_eqc as EQC
import analysis_issues as ISS
from project_utils import canonicalize_project_name, canonical_project_from_row
import subprocess
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment

# --- Checklist Categories (for flat-wise export separation) ---
# RCC/Structural checklists - apply at FLOOR/POUR level, not flat level
RCC_CATEGORIES: List[Tuple[str, List[str]]] = [
    (
        "RCC / Structural",
        [
            r"\brcc\b|reinforc|shuttering|\bstruct|slab|column|beam|footing",
        ],
    ),
]

# Aluform checklists - structural (floor-level)
ALUFORM_CATEGORIES: List[Tuple[str, List[str]]] = [
    (
        "Aluform",
        [
            r"aluform",
        ],
    ),
]

# Handover checklists - can be at multiple levels
HANDOVER_CATEGORIES: List[Tuple[str, List[str]]] = [
    (
        "Internal Handover",
        [
            r"internal\s*handover|handover",
        ],
    ),
]

# Finishing checklists - apply at FLAT level
FINISHING_CATEGORIES: List[Tuple[str, List[str]]] = [
    (
        "Masonry & Surface Preparation",
        [
            r"aac|autoclaved|block\s*masonry|fly\s*ash|brick\s*work",
            r"internal\s*plaster|plaster\s*works|gypsum",
        ],
    ),
    (
        "Waterproofing",
        [
            r"waterproof.*toilet|skirting|boxtype|toilet\s*&\s*skirting",
            r"brick\s*bat",
            r"pu\s*waterproof",
        ],
    ),
    (
        "Flooring & Finishes",
        [
            r"tiling.*dado|dado\s*\(toilet|kitchen\)|kitchen\s*dado",
            r"tiling.*floor|flooring",
            r"kitchen\s*platform|sink",
            r"granite.*(frame|door|window)",
            r"staircase.*skirting",
        ],
    ),
    (
        "Carpentry & Metal Works",
        [
            r"carpentry.*(door\s*frame|shutters?)",
            r"(door\s*frame|shutters?)",
            r"aluminium.*sliding|aluminum.*sliding|sliding\s*door|sliding\s*window",
            r"railings?\s*&?\s*grills?|ms\s*grill",
            r"ss\s*railing",
        ],
    ),
    (
        "MEP (Mechanical, Electrical & Plumbing)",
        [
            r"internal\s*plumbing|plumbing\s*works",
            r"drainage.*pvc|bathroom.*drain|kitchen.*drain",
            r"electrical.*conduit|wall.*conduiting|wiring",
            r"fire\s*fight|sprinkler",
            r"mep.*before.*casting|sleeve|opening",
        ],
    ),
    (
        "Painting & Finishing",
        [
            r"painting.*internal",
            r"painting.*(railing|grills).*oil",
        ],
    ),
    (
        "Final Cleaning & Handover",
        [
            r"cleaning.*acid|deep\s*clean",
        ],
    ),
]

UNCATEGORIZED_LABEL = "Uncategorized / Other"

# --- Helpers (schema + stage normalization) ---

def parse_date_safe(s: str) -> date | None:
    s = (s or "").strip()
    if not s:
        return None
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def normalize_stage(stage: str) -> str:
    # Delegate to analysis_eqc for consistency
    return EQC.normalize_stage(stage)


def canonical_project(row: pd.Series) -> str:
    # Use shared helper with fallback order and robust normalization
    return canonical_project_from_row(row)


def is_rcc_checklist(name: str) -> bool:
    """Check if a checklist is RCC/structural (floor-level) vs Finishing (flat-level)."""
    import re as _re
    s = (name or "").lower()
    for _, patterns in RCC_CATEGORIES:
        for pat in patterns:
            try:
                if _re.search(pat, s, _re.I):
                    return True
            except Exception:
                continue
    return False


def is_aluform_checklist(name: str) -> bool:
    """Check if a checklist is Aluform (floor-level)."""
    import re as _re
    s = (name or "").lower()
    for _, patterns in ALUFORM_CATEGORIES:
        for pat in patterns:
            try:
                if _re.search(pat, s, _re.I):
                    return True
            except Exception:
                continue
    return False


def is_handover_checklist(name: str) -> bool:
    """Check if a checklist is Handover."""
    import re as _re
    s = (name or "").lower()
    for _, patterns in HANDOVER_CATEGORIES:
        for pat in patterns:
            try:
                if _re.search(pat, s, _re.I):
                    return True
            except Exception:
                continue
    return False


def base_name(n: str) -> str:
    """Standardize checklist names by removing contractor suffixes (matches Weekly_report.py logic)."""
    import re
    if not isinstance(n, str):
        return str(n)
    s = n.strip()
    s_norm = s.lower()
    
    # Handle "Checklist for ..." names - preserve full name with Title Case
    if s_norm.startswith('checklist for '):
        # Remove contractor suffix if present (double space + name pattern)
        cleaned = re.sub(r'\s{2,}[A-Za-z]+\s*$', '', s)
        cleaned = re.sub(r'\s+\([^)]+\)\s*$', '', cleaned)  # Remove (Contractor Name)
        return cleaned.strip().title()
    
    # Known standardized names
    fixes = {
        r'painting.*internal': 'Painting Works : Internal',
        r'painting.*external': 'Painting Works : External',
        r'painting\s*works\s*:\s*external\s*texture': 'Painting Works : External Texture',
        r'waterproof.*boxtype': 'Waterproofing Works: Toilet and Skirting',
        r'waterproof.*toilet': 'Waterproofing Works: Toilet and Skirting',
        r'waterproof.*skirting': 'Waterproofing Works: Toilet and Skirting',
        r'tiling.*kitchen.*platform': 'Tiling - Kitchen Platform',
        r'tiling.*kitchen.*sink': 'Tiling - Kitchen Platform',
        r'tiling[-\s]*toilet.*dado': 'Tiling - Toilet Dado',
        r'tiling.*flooring': 'Tiling - Flooring Work',
        r'kitchen\s*dado': 'Kitchen Dado',
        r'gypsum\s*plaster': 'Gypsum Plaster Works',
        r'aac\s*block': 'AAC Block Work',
        r'brick\s*work': 'Brick Work',
        r'external\s*plaster': 'External Plaster Works',
        r'internal\s*plaster': 'Internal Plaster Works',
        r'\brcc\b.*footing': 'RCC Footing Work',
        r'\brcc\b.*column': 'RCC Column Work',
        r'\brcc\b.*beam': 'RCC Beam Work',
        r'\brcc\b.*slab': 'RCC Slab Work',
        r'\brcc\b\s*work': 'RCC Works',
        r'm\s*s\s*railing': 'MS Railing',
        r's\s*s\s*railing': 'SS Railing',
        r'alluminium\s*window': 'Aluminium Window and Door Work',
        r'aluminium\s*window': 'Aluminium Window and Door Work',
    }
    
    for pat, replacement in fixes.items():
        if re.search(pat, s_norm, re.I):
            return replacement
    
    # Fallback: capitalize first letter of each word
    return s.title()


def read_eqc_robust(fobj: io.BytesIO) -> pd.DataFrame:
    fobj.seek(0)
    last_err: Exception | None = None
    for sep in ("\t", ",", None):
        try:
            if sep is None:
                df = pd.read_csv(fobj, dtype=str, keep_default_na=False, sep=None, engine="python")
            else:
                fobj.seek(0)
                df = pd.read_csv(fobj, dtype=str, keep_default_na=False, sep=sep, engine="python")
            return df
        except Exception as e:
            last_err = e
    raise last_err if last_err else RuntimeError("Failed to read uploaded file")


# --- Excel formatting helpers (borders + wrap) ---

def _apply_borders_wrap_to_wb(wb) -> None:
    """Apply thin borders and wrap text to all cells in all sheets of an openpyxl Workbook."""
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align = Alignment(wrap_text=True, vertical="center")
    for ws in wb.worksheets:
        max_row = ws.max_row or 1
        max_col = ws.max_column or 1
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                try:
                    cell.border = border
                    cell.alignment = align
                except Exception:
                    # Skip cells that cannot be styled (unlikely)
                    pass


def _format_xlsx_from_path(xlsx_path: str) -> io.BytesIO:
    """Load an existing .xlsx file, apply borders + wrap to all sheets, return BytesIO of the result."""
    wb = load_workbook(xlsx_path)
    _apply_borders_wrap_to_wb(wb)
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# --- Session-scoped CSV management ---

def _ensure_upload_dir() -> str:
    root = os.path.join(tempfile.gettempdir(), "eqc_uploads")
    os.makedirs(root, exist_ok=True)
    return root


def _save_session_file(file_storage) -> tuple[str, str]:
    """Save uploaded file to a temp location and return (path, display_name)."""
    upload_dir = _ensure_upload_dir()
    ext = ".csv"
    name = getattr(file_storage, "filename", "Combined_EQC.csv") or "Combined_EQC.csv"
    try:
        base_ext = os.path.splitext(name)[1]
        if base_ext:
            ext = base_ext
    except Exception:
        pass
    uid = uuid.uuid4().hex
    out_path = os.path.join(upload_dir, f"eqc_{uid}{ext}")
    file_storage.stream.seek(0)
    with open(out_path, "wb") as out:
        out.write(file_storage.read())
    return out_path, name


def _get_session_eqc_path() -> str | None:
    p = session.get("eqc_file_path")
    if p and os.path.exists(p):
        return p
    return None


# Issues file session management
def _ensure_issues_upload_dir() -> str:
    root = os.path.join(tempfile.gettempdir(), "issues_uploads")
    os.makedirs(root, exist_ok=True)
    return root


def _save_session_issues_file(file_storage) -> tuple[str, str]:
    upload_dir = _ensure_issues_upload_dir()
    ext = ".csv"
    name = getattr(file_storage, "filename", "Combined_Instructions.csv") or "Combined_Instructions.csv"
    try:
        base_ext = os.path.splitext(name)[1]
        if base_ext:
            ext = base_ext
    except Exception:
        pass
    uid = uuid.uuid4().hex
    out_path = os.path.join(upload_dir, f"issues_{uid}{ext}")
    file_storage.stream.seek(0)
    with open(out_path, "wb") as out:
        out.write(file_storage.read())
    return out_path, name


def _get_session_issues_path() -> str | None:
    p = session.get("issues_file_path")
    if p and os.path.exists(p):
        return p
    return None


# --- EQC summaries (cumulative/month/daily) ---

def eqc_summaries(df: pd.DataFrame, target: date) -> Dict[str, Dict[str, Dict[str, int]]]:
    df = df.copy()
    # Drop known footer/summary rows like analysis_eqc._read_and_clean
    for footer_col in ["Total EQC Stages", "Fail Stages", "% Fail"]:
        if footer_col in df.columns:
            df = df[df[footer_col].astype(str).str.strip().isin(["", "nan", "NaN", "None", "-"])]
    if "Stage" in df.columns and "Eqc Type" in df.columns:
        mask_empty = df["Stage"].astype(str).str.strip().eq("") & df["Eqc Type"].astype(str).str.strip().eq("")
        if mask_empty.any():
            df = df[~mask_empty]
    # DEMO filter
    for col in ("Project", "Project Name", "Location L0"):
        if col in df.columns:
            mask_demo = df[col].astype(str).str.contains("DEMO", case=False, na=False)
            df = df[~mask_demo]
    # Stages and dates
    if "Stage" not in df.columns:
        # If input is malformed, return empty summary
        return {}
    df["__Stage"] = df["Stage"].map(EQC.normalize_stage)
    dates = df.get("Date", pd.Series([None] * len(df), index=df.index)).astype(str).map(EQC._parse_date_safe)
    # Projects
    df["__ProjectKey"] = df.apply(canonical_project, axis=1)
    projects = [p for p in sorted(df["__ProjectKey"].astype(str).str.strip().unique()) if p]

    def counts_daily_post_plus_other(frame: pd.DataFrame) -> Dict[str, int]:
        # Raw counts with Other added to all columns (single-stage checklists)
        if frame is None or frame.empty:
            return {"Pre": 0, "During": 0, "Post": 0}
        c = frame.groupby("__Stage", dropna=False)["__Stage"].count()
        n_pre = int(c.get("Pre", 0)); n_during = int(c.get("During", 0)); n_post = int(c.get("Post", 0)); n_other = int(c.get("Other", 0))
        return {"Pre": n_pre + n_other, "During": n_during + n_other, "Post": n_post + n_other}

    def counts_raw(frame: pd.DataFrame) -> Dict[str, int]:
        # Raw counts without any cumulative roll-up and without adding Other
        if frame is None or frame.empty:
            return {"Pre": 0, "During": 0, "Post": 0}
        c = frame.groupby("__Stage", dropna=False)["__Stage"].count()
        n_pre = int(c.get("Pre", 0)); n_during = int(c.get("During", 0)); n_post = int(c.get("Post", 0))
        return {"Pre": n_pre, "During": n_during, "Post": n_post}

    def counts_cumulative_weeklylogic(frame: pd.DataFrame) -> Dict[str, int]:
        """Cumulative roll-up using the same mapping as the console (weekly report) and workbook:

        - Pre = total rows (including Other for single-stage checklists)
        - During = During + Post + Other (single-stage checklists counted here)
        - Post = Post + Other (single-stage checklists counted here)

        Where 'Other' are rows whose Stage is not pre/during/post/reinforce/shutter.
        Single-stage checklists (like Paver block, Kerb stone) are counted in all columns.
        """
        if frame is None or frame.empty:
            return {"Pre": 0, "During": 0, "Post": 0}
        s = frame.get("Stage", pd.Series(["" ] * len(frame), index=frame.index)).astype(str).str.lower()
        pre = s.str.contains("pre", na=False)
        during = s.str.contains("during", na=False)
        post = s.str.contains("post", na=False)
        reinf = s.str.contains("reinforce", na=False)
        shut = s.str.contains("shutter", na=False)
        other = ~(pre | during | post | reinf | shut)
        total = int(len(s))
        d_count = int(during.sum())
        p_count = int(post.sum())
        o_count = int(other.sum())
        return {"Pre": total, "During": d_count + p_count + o_count, "Post": p_count + o_count}

    out: Dict[str, Dict[str, Dict[str, int]]] = {}
    for proj in projects:
        sub = df[df["__ProjectKey"].astype(str).str.strip() == proj]
        dates_sub = dates.loc[sub.index]
        month_mask = dates_sub.apply(lambda d: bool(d and d.year == target.year and d.month == target.month))
        today_mask = dates_sub == target
        out[proj] = {
            # All-time: cumulative roll-up to match workbook 'Cumulative' ALL sums (weekly report mapping)
            "all": counts_cumulative_weeklylogic(sub),
            # This month: raw counts with Post += Other
            "month": counts_daily_post_plus_other(sub[month_mask]),
            # Today: raw counts with Post += Other
            "today": counts_daily_post_plus_other(sub[today_mask]),
        }
    return out


# --- Flask app ---
app = Flask(__name__)
# Simple secret key for sessions (override via env SECRET_KEY in production)
app.secret_key = os.environ.get("SECRET_KEY", "change-me-please")

# --- AdSense config (optional) ---
app.config["ADSENSE_CLIENT_ID"] = (os.environ.get("ADSENSE_CLIENT_ID", "").strip())
app.config["ADSENSE_ENABLE"] = (os.environ.get("ADSENSE_ENABLE", "1").strip().lower() in {"1","true","yes","on"}) if app.config["ADSENSE_CLIENT_ID"] else False

@app.context_processor
def _inject_adsense_config():
    return {
        "ADSENSE_CLIENT_ID": app.config.get("ADSENSE_CLIENT_ID", ""),
        "ADSENSE_ENABLE": app.config.get("ADSENSE_ENABLE", False),
    }

# Hardcoded simple credentials (no database)
AUTH_USERNAME = os.environ.get("APP_USERNAME", "admin")
AUTH_PASSWORD = os.environ.get("APP_PASSWORD", "admin123")

# Default external person name for issues (can be overridden via UI). Use empty by default.
DEFAULT_ISSUES_EXTERNAL = os.environ.get("ISSUES_EXTERNAL_NAME", "")

# Public route for ads.txt at domain root
@app.get("/ads.txt")
def ads_txt():
    # Serve the static/ads.txt file as plain text without requiring login
    return send_from_directory(os.path.join(app.root_path, "static"), "ads.txt", mimetype="text/plain")

def login_required(fn):
    from functools import wraps
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not session.get("user"):
            return redirect(url_for("login", next=request.path))
        return fn(*args, **kwargs)
    return wrapper


@app.post("/set-eqc-file")
@login_required
def set_eqc_file():
    f = request.files.get("eqc")
    nxt = request.args.get("next") or request.referrer or url_for("index")
    if not f:
        return redirect(nxt)
    try:
        path, disp = _save_session_file(f)
    except Exception as e:
        return f"Failed to save file: {e}", 400
    session["eqc_file_path"] = path
    session["eqc_file_name"] = disp
    return redirect(nxt)


@app.post("/set-issues-file")
@login_required
def set_issues_file():
    f = request.files.get("issues")
    nxt = request.args.get("next") or request.referrer or url_for("index")
    if not f:
        return redirect(nxt)
    try:
        path, disp = _save_session_issues_file(f)
    except Exception as e:
        return f"Failed to save issues file: {e}", 400
    session["issues_file_path"] = path
    session["issues_file_name"] = disp
    return redirect(nxt)


@app.post("/set-issues-external")
@login_required
def set_issues_external():
    name = (request.form.get("external_name") or "").strip()
    nxt = request.args.get("next") or request.referrer or url_for("index")
    # If blank: set to empty to indicate "no external auditor"
    session["issues_external_name"] = name
    return redirect(nxt)


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        if username == AUTH_USERNAME and password == AUTH_PASSWORD:
            session["user"] = username
            # After login, open uploads modal by default to set session files
            return redirect(url_for("index", openUpload=1))
        return render_template("login.html", error="Invalid credentials")
    if session.get("user"):
        return redirect(url_for("index"))
    return render_template("login.html")

@app.get("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.get("/")
@login_required
def index():
        # Optional helper: suggest common "Raised By" names if Issues file is available
        top_names: List[Tuple[str, int]] = []
        path = _get_session_issues_path()
        if not path or not os.path.exists(path):
            # Autodetect most recent Instruction CSV in workspace
            candidates: list[str] = []
            try:
                for root, _dirs, files in os.walk(os.getcwd()):
                    for f in files:
                        fl = f.lower()
                        if fl.endswith('.csv') and ('instruction' in fl or fl.startswith('csv-instruction-latest-report')):
                            candidates.append(os.path.join(root, f))
            except Exception:
                candidates = []
            if candidates:
                path = max(candidates, key=lambda p: os.path.getmtime(p))
        try:
            if path and os.path.exists(path):
                df_issues = _robust_read_issues_path(path)
                if "Raised By" in df_issues.columns:
                    ser = df_issues["Raised By"].astype(str).str.strip()
                    ser = ser[ser != ""]
                    top = ser.value_counts().head(8)
                    top_names = [(str(idx), int(cnt)) for idx, cnt in top.items()]
        except Exception:
            top_names = []
        return render_template("index.html", top_names=top_names)


@app.get("/api/suggest-raised-by")
@login_required
def api_suggest_raised_by():
    """Return a JSON list of suggested 'Raised By' names matching the query.
    Query params: q (string), limit (int, default 8)
    """
    q = (request.args.get("q") or "").strip()
    try:
        limit = int(request.args.get("limit", 8))
    except Exception:
        limit = 8
    path = _get_session_issues_path()
    if not path or not os.path.exists(path):
        return jsonify([])
    try:
        df = _robust_read_issues_path(path)
        if "Raised By" not in df.columns:
            return jsonify([])
        ser = df["Raised By"].astype(str).str.strip()
        ser = ser[ser != ""]
        if q:
            ser = ser[ser.str.contains(q, case=False, na=False)]
        counts = ser.value_counts()
        out = [{"name": str(idx), "count": int(cnt)} for idx, cnt in counts.head(limit).items()]
        return jsonify(out)
    except Exception:
        return jsonify([])


@app.get("/reports")
@login_required
def reports_page():
    return render_template("reports.html")


@app.get("/dashboards")
@login_required
def dashboards_page():
    return render_template("dashboards.html")


@app.post("/daily-dashboard")
@login_required
def daily_dashboard():
    f = request.files.get("eqc")
    if f:
        buf = io.BytesIO(f.read())
    else:
        p = _get_session_eqc_path() or os.path.join(os.getcwd(), "Combined_EQC.csv")
        if not p or not os.path.exists(p):
            # As a last resort, try to autodetect a likely EQC CSV
            candidates: list[str] = []
            try:
                for root, _dirs, files in os.walk(os.getcwd()):
                    for fn in files:
                        fl = fn.lower()
                        if not fl.endswith('.csv'):
                            continue
                        if fl.startswith('combined_eqc') or 'eqc' in fl:
                            candidates.append(os.path.join(root, fn))
            except Exception:
                candidates = []
            if candidates:
                p = max(candidates, key=lambda q: os.path.getmtime(q))
            else:
                return redirect(url_for("index"))
        try:
            with open(p, "rb") as fh:
                buf = io.BytesIO(fh.read())
        except Exception as e:
            return f"Failed to read stored file: {e}", 400
    try:
        df = read_eqc_robust(buf)
    except Exception as e:
        return f"Failed to read file: {e}", 400
    today = date.today()
    data = eqc_summaries(df, today)
    # Sort projects for stable display
    data_sorted = {k: data[k] for k in sorted(data.keys())}
    return render_template("dashboard.html", data=data_sorted, target=today.strftime("%d-%m-%Y"))


@app.get("/daily-dashboard")
@login_required
def daily_dashboard_get():
    # Support GET to render using session EQC file or sensible fallbacks
    p = _get_session_eqc_path() or os.path.join(os.getcwd(), "Combined_EQC.csv")
    if not p or not os.path.exists(p):
        # Autodetect most recent EQC-like CSV
        candidates: list[str] = []
        try:
            for root, _dirs, files in os.walk(os.getcwd()):
                for fn in files:
                    fl = fn.lower()
                    if not fl.endswith('.csv'):
                        continue
                    if fl.startswith('combined_eqc') or 'eqc' in fl:
                        candidates.append(os.path.join(root, fn))
        except Exception:
            candidates = []
        if candidates:
            p = max(candidates, key=lambda q: os.path.getmtime(q))
        else:
            # Nothing to show; fallback to index
            return redirect(url_for("index"))
    try:
        with open(p, "rb") as fh:
            buf = io.BytesIO(fh.read())
        df = read_eqc_robust(buf)
    except Exception as e:
        return f"Failed to read EQC file: {e}", 400
    today = date.today()
    data = eqc_summaries(df, today)
    data_sorted = {k: data[k] for k in sorted(data.keys())}
    return render_template("dashboard.html", data=data_sorted, target=today.strftime("%d-%m-%Y"))


@app.post("/weekly-report")
@login_required
def weekly_report():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    f = request.files.get("eqc")
    if f:
        # Save to persistent temp file (not auto-deleted directory)
        upload_dir = _ensure_upload_dir()
        uid = uuid.uuid4().hex
        in_path = os.path.join(upload_dir, f"weekly_report_{uid}.csv")
        with open(in_path, "wb") as out:
            out.write(f.read())
        
        try:
            import subprocess, sys as _sys
            cmd = [_sys.executable, "Weekly_report.py", "--input", in_path]
            proc = subprocess.run(cmd, cwd=script_dir, capture_output=True, text=True)
            if proc.returncode != 0:
                return f"Failed to build weekly workbook.\nSTDOUT:\n{proc.stdout}\nSTDERR:\n{proc.stderr}", 500
            out_xlsx = os.path.join(script_dir, "EQC_Weekly_Monthly_Cumulative_AllProjects.xlsx")
            if not os.path.exists(out_xlsx):
                return f"Workbook not found after generation. Expected at: {out_xlsx}\nSTDOUT:\n{proc.stdout}\nSTDERR:\n{proc.stderr}", 500
            # Apply formatting before sending
            content = _format_xlsx_from_path(out_xlsx)
            try:
                os.remove(out_xlsx)
            except Exception:
                pass
            today_str = date.today().strftime("%d-%m-%Y")
            dl_name = f"EQC_Weekly_Monthly_Cumulative_AllProjects_{today_str}.xlsx"
            return send_file(content, as_attachment=True, download_name=dl_name, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        finally:
            # Clean up the temp file after processing
            try:
                if os.path.exists(in_path):
                    os.remove(in_path)
            except Exception:
                pass
    else:
        p = _get_session_eqc_path()
        if not p:
            return redirect(url_for("index"))
        import subprocess, sys as _sys
        cmd = [_sys.executable, "Weekly_report.py", "--input", p]
        proc = subprocess.run(cmd, cwd=script_dir, capture_output=True, text=True)
        if proc.returncode != 0:
            return f"Failed to build weekly workbook.\nSTDOUT:\n{proc.stdout}\nSTDERR:\n{proc.stderr}", 500
        out_xlsx = os.path.join(script_dir, "EQC_Weekly_Monthly_Cumulative_AllProjects.xlsx")
        if not os.path.exists(out_xlsx):
            return f"Workbook not found after generation. Expected at: {out_xlsx}\nSTDOUT:\n{proc.stdout}\nSTDERR:\n{proc.stderr}", 500
        # Apply formatting before sending
        content = _format_xlsx_from_path(out_xlsx)
        try:
            os.remove(out_xlsx)
        except Exception:
            pass
        today_str = date.today().strftime("%d-%m-%Y")
        dl_name = f"EQC_Weekly_Monthly_Cumulative_AllProjects_{today_str}.xlsx"
        return send_file(content, as_attachment=True, download_name=dl_name, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/weekly-report")
@login_required
def weekly_report_get():
    # Avoid 405 if someone opens this URL directly; send them to the form
    return redirect(url_for("index"))


# ---------------- Flat-wise report (Floor/Flat) ----------------

# Category order and mapping adapted from flat_app.py
from typing import Tuple as _Tuple, List as _List

CATEGORY_ORDER: _List[_Tuple[str, _List[str]]] = [
    (
        "1. RCC / Structural Stage ",
        [
            # structural/aluform
            r"\brcc\b|reinforc|shuttering|\bstruct|slab|column|beam|footing",
            r"aluform|internal\s*handover",
        ],
    ),
    (
        "2. Masonry & Surface Preparation",
        [
            r"aac|autoclaved|block\s*masonry|fly\s*ash|brick\s*work",
            r"internal\s*plaster|plaster\s*works|gypsum",
        ],
    ),
    (
        "3. Waterproofing",
        [
            r"waterproof.*toilet|skirting|boxtype|toilet\s*&\s*skirting",
            r"brick\s*bat",
            r"pu\s*waterproof",
        ],
    ),
    (
        "4. Flooring & Finishes",
        [
            r"tiling.*dado|dado\s*\(toilet|kitchen\)|kitchen\s*dado",
            r"tiling.*floor|flooring",
            r"kitchen\s*platform|sink",
            r"granite.*(frame|door|window)",
            r"staircase.*skirting",
        ],
    ),
    (
        "5. Carpentry & Metal Works",
        [
            r"carpentry.*(door\s*frame|shutters?)",
            r"(door\s*frame|shutters?)",
            r"aluminium.*sliding|aluminum.*sliding|sliding\s*door|sliding\s*window",
            r"railings?\s*&?\s*grills?|ms\s*grill",
            r"ss\s*railing",
        ],
    ),
    (
        "6. MEP (Mechanical, Electrical & Plumbing)",
        [
            r"internal\s*plumbing|plumbing\s*works",
            r"drainage.*pvc|bathroom.*drain|kitchen.*drain",
            r"electrical.*conduit|wall.*conduiting|wiring",
            r"fire\s*fight|sprinkler",
            r"mep.*before.*casting|sleeve|opening",
        ],
    ),
    (
        "7. Painting & Finishing",
        [
            r"painting.*internal",
            r"painting.*(railing|grills).*oil",
        ],
    ),
    (
        "8. Final Cleaning & Handover",
        [
            r"cleaning.*acid|deep\s*clean",
        ],
    ),
]

UNCATEGORIZED_LABEL = "Uncategorized / Other"


def _map_checklist_to_category(name: str) -> _Tuple[int, str]:
    import re as _re
    s = (name or "").lower()
    for idx, (label, patterns) in enumerate(CATEGORY_ORDER):
        for pat in patterns:
            try:
                if _re.search(pat, s, _re.I):
                    return idx, label
            except Exception:
                continue
    return len(CATEGORY_ORDER), UNCATEGORIZED_LABEL


def _robust_read_eqc_path(path: str) -> pd.DataFrame:
    import warnings
    last_err = None
    # Try different encodings and separators
    encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
    
    for encoding in encodings:
        for sep in ("\t", ",", None):
            # Try different quoting and error handling options
            for on_bad_lines in ['skip', 'warn']:
                try:
                    with warnings.catch_warnings():
                        warnings.filterwarnings('ignore', category=pd.errors.ParserWarning)
                        if sep is None:
                            df = pd.read_csv(path, dtype=str, keep_default_na=False, sep=None, 
                                             engine="python", encoding=encoding, on_bad_lines=on_bad_lines)
                        else:
                            df = pd.read_csv(path, dtype=str, keep_default_na=False, sep=sep, 
                                             engine="python", encoding=encoding, on_bad_lines=on_bad_lines)
                        if df.shape[0] > 0 and df.shape[1] > 5:  # Ensure we have actual data
                            return df
                except Exception as e:
                    last_err = e
    
    # Last resort: try with c engine and skip bad lines
    try:
        with warnings.catch_warnings():
            warnings.filterwarnings('ignore', category=pd.errors.ParserWarning)
            df = pd.read_csv(path, dtype=str, keep_default_na=False, encoding='latin-1', on_bad_lines='skip')
            if df.shape[0] > 0 and df.shape[1] > 5:
                return df
    except Exception as e:
        last_err = e
    
    raise last_err if last_err else RuntimeError("Failed to read EQC file")


def _robust_read_issues_path(path: str) -> pd.DataFrame:
    # Try normal read via ISS helper first
    try:
        df = ISS._read_instructions(path)
    except Exception:
        df = pd.DataFrame()
    if isinstance(df, pd.DataFrame) and df.shape[1] > 1:
        return df
    # Fallback: attempt header merge if everything got jammed into one column
    try:
        with open(path, 'r', encoding='utf-8', errors='replace') as f:
            text = f.read()
        import re as _re
        m = _re.search(r"(?m)^\s*\d{2,},", text)
        if m:
            pos = m.start()
            header = text[:pos].replace("\r", "").replace("\n", "")
            data = text[pos:]
            new_text = header + "\n" + data
            buf = io.StringIO(new_text)
            df2 = pd.read_csv(buf, dtype=str, keep_default_na=False, engine='python')
            if df2.shape[1] > 1:
                return df2
    except Exception:
        pass
    return df if isinstance(df, pd.DataFrame) else pd.DataFrame()


def _prepare_frame(df: pd.DataFrame) -> pd.DataFrame:
    # Copy & filter DEMO
    if df is None or df.empty:
        return pd.DataFrame()
    
    df = df.copy()
    
    cols = [c for c in ("Project", "Project Name", "Location L0") if c in df.columns]
    if cols:
        mask = pd.Series(False, index=df.index)
        for c in cols:
            mask |= df[c].astype(str).str.contains("DEMO", case=False, na=False)
        df = df[~mask]
    
    if df.empty:
        return df
    # Stage & date
    if "Stage" in df.columns:
        df["__Stage"] = df["Stage"].map(EQC.normalize_stage)
    if "Date" in df.columns:
        df["__Date"] = df["Date"].astype(str).map(EQC._parse_date_safe)
    else:
        df["__Date"] = None
    # Project key
    df["__Project"] = df.apply(canonical_project_from_row, axis=1)

    # Preferred building label per (Project, Letter)
    import re as _re_map
    bldg_pref_map: Dict[_Tuple[str, str], str] = {}
    pref_counts: Dict[_Tuple[str, str, str], int] = {}
    if "Location L1" in df.columns:
        for _, r in df.iterrows():
            proj = str(r.get("__Project", "") or "").strip()
            l1 = str(r.get("Location L1", "") or "").strip()
            if not proj or not l1:
                continue
            m = _re_map.search(r"\b(wing|tower|building|block|bldg|blk)\b[\s\-]*([A-Za-z0-9])\b", l1, _re_map.I) or \
                _re_map.search(r"\b([A-Za-z0-9])\b[\s\-]*\b(wing|tower|building|block|bldg|blk)\b", l1, _re_map.I)
            if m:
                g = m.groups()
                letter = g[1] if g[0].lower() in ("wing","tower","building","block","bldg","blk") else g[0]
                key = (proj, str(letter).upper())
                pref_counts[(proj, str(letter).upper(), l1)] = pref_counts.get((proj, str(letter).upper(), l1), 0) + 1
        for (proj, letter, label), cnt in pref_counts.items():
            cur = bldg_pref_map.get((proj, letter))
            if cur is None or cnt > pref_counts.get((proj, letter, cur), 0):
                bldg_pref_map[(proj, letter)] = label

    def _infer_bff(row: pd.Series) -> _Tuple[str, str, str]:
        import re as _re
        b = str(row.get("Location L1", "") or "").strip()
        f = str(row.get("Location L2", "") or "").strip()
        fl = str(row.get("Location L3", "") or "").strip()
        sources = []
        for c in ("EQC", "Eqc", "Eqc Type", "Stage", "Status"):
            if c in row and str(row[c]).strip():
                sources.append(str(row[c]))
        for c in row.index:
            sc = str(c).strip().lower()
            if sc in {"eqc path", "eqc location", "eqc name"} and str(row[c]).strip():
                sources.append(str(row[c]))
        tokens: _List[str] = []
        for s in sources:
            tokens.extend([t.strip() for t in _re.split(r"[>/|,;:\-–—]+", str(s)) if t.strip()])
        if not b:
            patt_building = [
                _re.compile(r"\b(wing|tower|building|block)[\s\-]*([A-Za-z0-9])\b", _re.I),
                _re.compile(r"\b([A-Za-z0-9])[\s\-]*(wing|tower|building|block)\b", _re.I),
                _re.compile(r"\b(bldg|blk)\.?[\s\-]*([A-Za-z0-9])\b", _re.I),
            ]
            proj = str(row.get("__Project", "") or "").strip()
            letter_found = None
            type_found = None
            for s in sources + tokens:
                for pat in patt_building:
                    m = pat.search(s)
                    if m:
                        g1, g2 = m.groups()
                        type_token = g1 if g1.lower() in ("wing","tower","building","block","bldg","blk") else g2
                        letter_token = g2 if type_token == g1 else g1
                        letter_found = str(letter_token).strip().upper()
                        type_found = str(type_token).strip().lower()
                        break
                if b:
                    break
            if not b and letter_found:
                pref = bldg_pref_map.get((proj, letter_found))
                if pref:
                    b = pref
                elif type_found:
                    type_map = {"bldg": "Building", "blk": "Block"}
                    t = type_map.get(type_found, type_found).title()
                    b = f"{t} {letter_found}"
        if not f:
            for loc_field in ("Location L2", "Location L3", "Location L4", "Location L1", "Location L0"):
                sv = str(row.get(loc_field, "") or "").strip()
                if not sv:
                    continue
                if _re.search(r"\bshear\s*wall\b|stair\s*case|lobby|podium|development|terrace|parking", sv, _re.I):
                    f = sv
                    break
                m = _re.search(r"\b(floor|fl|level)[\s\-]*(\d{1,2})\b", sv, _re.I) or _re.search(r"\b(\d{1,2})[\s\-]*(floor|fl|level)\b", sv, _re.I) or _re.search(r"\b(?:f|fl|lvl|l)[\s\-]*(\d{1,2})\b", sv, _re.I) or _re.search(r"\b(\d{1,2})(?:st|nd|rd|th)?\s*(?:floor)?\b", sv, _re.I)
                if m:
                    num = m.group(m.lastindex or 1)
                    f = f"Floor {num}"
                    break
        if not f:
            patt_floor = [
                _re.compile(r"\b(floor|fl|level)[\s\-]*(\d{1,2})\b", _re.I),
                _re.compile(r"\b(\d{1,2})[\s\-]*(floor|fl|level)\b", _re.I),
                _re.compile(r"\b(?:f|fl|lvl|l)[\s\-]*(\d{1,2})\b", _re.I),
                _re.compile(r"\b(\d{1,2})(?:st|nd|rd|th)?\s*(?:floor)?\b", _re.I),
            ]
            for s in sources + tokens:
                for pat in patt_floor:
                    m = pat.search(s)
                    if m:
                        num = m.group(m.lastindex or 1)
                        f = f"Floor {num}"
                        break
                if f:
                    break
        if not fl:
            # Require an explicit type like Flat/Shop/Unit/Apt/Room to avoid misreading product codes (e.g., "Samshield XL 1500") as flats
            patt_flat = [
                _re.compile(r"\b(flat|unit|apt|apartment|shop)[\s\-]*([A-Z]?\d{1,4}[A-Z]?)\b", _re.I),
                _re.compile(r"\b(room)[\s\-]*(\d{1,4}[A-Z]?)\b", _re.I),
            ]
            for s in sources + tokens:
                for pat in patt_flat:
                    m = pat.search(s)
                    if m:
                        val = m.group(2) if m.lastindex and m.lastindex >= 2 else m.group(1)
                        core = _re.sub(r"^([A-Z]?)(0+)(\d)", r"\1\3", val)
                        typ = m.group(1).strip().title() if (m.lastindex and m.lastindex >= 2) else "Flat"
                        if typ.lower() not in ("flat","unit","apt","apartment","shop"):
                            typ = "Flat"
                        if typ.lower() in ("unit","apt","apartment"):
                            typ = "Flat"
                        fl = f"{typ} {core}"
                        break
                if fl:
                    break
        def _std_floor(x: str) -> str:
            s = (x or "").strip()
            if not s:
                return ""
            m = _re.search(r"(\d{1,2})", s)
            if m and _re.search(r"floor|fl|level|lvl|^\d{1,2}$", s, _re.I):
                return f"Floor {m.group(1)}"
            return s
        def _std_flat(x: str) -> str:
            s = (x or "").strip()
            if not s:
                return ""
            if _re.search(r"^flat\s+", s, _re.I):
                return _re.sub(r"^flat\s+", "Flat ", s, flags=_re.I)
            m = _re.search(r"([A-Z]?\d{1,4}[A-Z]?)", s)
            if m:
                t = "Flat"
                m2 = _re.search(r"^(shop|flat)\b", s, _re.I)
                if m2:
                    t = m2.group(1).title()
                core = _re.sub(r"^([A-Z]?)(0+)(\d)", r"\1\3", m.group(1))
                return f"{t} {core}"
            return s
        return (b or "").strip(), _std_floor(f), _std_flat(fl)

    if len(df) > 0:
        try:
            bff = df.apply(_infer_bff, axis=1, result_type="expand")
            if isinstance(bff, pd.DataFrame) and len(bff.columns) >= 3:
                df["__Building"] = bff[0].astype(str)
                df["__Floor"] = bff[1].astype(str)
                df["__Flat"] = bff[2].astype(str)
            else:
                df["__Building"] = ""
                df["__Floor"] = ""
                df["__Flat"] = ""
        except Exception as e:
            print(f"Error in _infer_bff: {e}")
            df["__Building"] = ""
            df["__Floor"] = ""
            df["__Flat"] = ""
    else:
        df["__Building"] = ""
        df["__Floor"] = ""
        df["__Flat"] = ""
    
    # Extract Pour information from Location L3/L4
    def extract_pour(row: pd.Series) -> str:
        import re as _re
        for loc_field in ("Location L3", "Location L4", "Location Variable"):
            val = str(row.get(loc_field, "") or "").strip()
            if val and _re.search(r"\bpour\b", val, _re.I):
                # Extract pour number/name (e.g., "Pour 1", "Pour 2")
                m = _re.search(r"\bpour[\s\-]*(\d+|[A-Za-z0-9]+)\b", val, _re.I)
                if m:
                    return f"Pour {m.group(1)}"
                return val
        return ""
    
    if len(df) > 0:
        try:
            df["__Pour"] = df.apply(extract_pour, axis=1)
        except Exception as e:
            print(f"Error extracting pour: {e}")
            df["__Pour"] = ""
    else:
        df["__Pour"] = ""
    
    df["__Checklist"] = df["Eqc Type"].astype(str) if "Eqc Type" in df.columns else ""
    df["__URL"] = df["URL"].astype(str) if "URL" in df.columns else ""
    return df


@app.get("/flat-report")
@login_required
def flat_report_index():
    # Default to session CSV if present; allow override via ?path=
    path = request.args.get("path") or _get_session_eqc_path() or "Combined_EQC.csv"
    if not os.path.exists(path):
        return f"EQC file not found: {path}", 404
    
    try:
        df = _prepare_frame(_robust_read_eqc_path(path))
    except Exception as e:
        return f"Error reading EQC file: {e}", 500
    
    if df.empty or "__Project" not in df.columns:
        return f"No valid data found in EQC file: {path}", 400
    
    projects = sorted([p for p in df["__Project"].astype(str).str.strip().unique() if p])
    proj = request.args.get("project") or (projects[0] if projects else None)
    dfp = df[df["__Project"].astype(str).str.strip().eq(proj)] if proj else df.iloc[0:0]

    buildings_all = sorted({(b if b else "UNKNOWN") for b in dfp["__Building"].astype(str).unique()})
    default_building = next((b for b in buildings_all if str(b).upper() != "UNKNOWN"), (buildings_all[0] if buildings_all else None))
    building = request.args.get("building") or default_building

    dfl = dfp[dfp["__Building"].astype(str).fillna("").replace("", "UNKNOWN").eq(building or "")]

    def _floor_key(v: str):
        import re as _re
        s = str(v or "").strip()
        m = _re.search(r"(floor|fl|level)\s*(\d+)", s, _re.I)
        if m:
            return (0, int(m.group(2)))
        m = _re.search(r"parking\s*(\d+)", s, _re.I)
        if m:
            return (1, int(m.group(1)))
        order = {"terrace": (2, 0), "development": (3, 0), "shops": (4, 0), "unknown": (9, 0)}
        key = s.lower()
        return order.get(key, (8, 0))

    floors_set = {(f if f else "UNKNOWN") for f in dfl["__Floor"].astype(str).unique()}
    if not floors_set or all(str(x).upper() == "UNKNOWN" for x in floors_set):
        alt = set()
        def _derive_floor_from_row(row: pd.Series) -> str:
            import re as _re
            specials = r"shear\s*wall|stair\s*case|lobby|podium|development|terrace|parking|upper\s*ground|lower\s*ground|mezzanine|basement|pod"
            for loc_field in ("Location L2", "Location L3", "Location L4", "Location L1", "Location L0"):
                sv = str(row.get(loc_field, "") or "").strip()
                if not sv:
                    continue
                if _re.search(specials, sv, _re.I):
                    return sv
                m = _re.search(r"\b(floor|fl|level)[\s\-]*(\d{1,2})\b", sv, _re.I) or _re.search(r"\b(\d{1,2})[\s\-]*(floor|fl|level)\b", sv, _re.I) or _re.search(r"\b(?:f|fl|lvl|l)[\s\-]*(\d{1,2})\b", sv, _re.I) or _re.search(r"\b(\d{1,2})(?:st|nd|rd|th)?\s*(?:floor)?\b", sv, _re.I)
                if m:
                    num = m.group(m.lastindex or 1)
                    return f"Floor {num}"
            eqc = str(row.get("EQC", "") or "").strip()
            if eqc:
                toks = [t.strip() for t in _re.split(r"[>/|,;:\\-–—]+", eqc) if t.strip()]
                for sv in [eqc] + toks:
                    if _re.search(specials, sv, _re.I):
                        return sv
                    m = _re.search(r"\b(floor|fl|level)[\s\-]*(\d{1,2})\b", sv, _re.I) or _re.search(r"\b(\d{1,2})[\s\-]*(floor|fl|level)\b", sv, _re.I) or _re.search(r"\b(?:f|fl|lvl|l)[\s\-]*(\d{1,2})\b", sv, _re.I) or _re.search(r"\b(\d{1,2})(?:st|nd|rd|th)?\s*(?:floor)?\b", sv, _re.I)
                    if m:
                        num = m.group(m.lastindex or 1)
                        return f"Floor {num}"
            return "UNKNOWN"
        for _, r in dfl.iterrows():
            val = _derive_floor_from_row(r)
            if val:
                alt.add(val)
        floors_set = alt if alt else floors_set
    floors_all = sorted(floors_set, key=_floor_key)
    import re as _re_df
    default_floor = next((f for f in floors_all if _re_df.search(r"\bfloor\s*\d+\b", str(f), _re_df.I)), None)
    if not default_floor:
        default_floor = next((f for f in floors_all if str(f).upper() != "UNKNOWN"), (floors_all[0] if floors_all else None))
    floor = request.args.get("floor") or default_floor

    if floor == "ALL":
        dff = dfl.copy()
    else:
        dff = dfl[dfl["__Floor"].astype(str).fillna("").replace("", "UNKNOWN").eq(floor or "")]

    def _flat_key(v: str):
        import re as _re
        s = str(v or "").strip()
        m = _re.search(r"(shop|flat|unit|apt|apartment)\s*([A-Z]?\d+)([A-Z]?)", s, _re.I)
        if m:
            num = int(_re.sub(r"\D", "", m.group(2))) if _re.sub(r"\D", "", m.group(2)) else 0
            suf = m.group(3).upper() if m.group(3) else ""
            return (0, num, suf)
        m2 = _re.search(r"(\d+)", s)
        if m2:
            return (1, int(m2.group(1)), "")
        return (9, 0, s.lower())

    flats_set = {(x if x else "UNKNOWN") for x in dff["__Flat"].astype(str).unique()}
    import re as _re_fl
    flats_filtered = [x for x in flats_set if _re_fl.match(r"^(?:Flat|Shop)\s+", str(x))]
    flats_all = sorted(flats_filtered if flats_filtered else flats_set, key=_flat_key)
    default_flat = next((x for x in flats_all if str(x).upper() != "UNKNOWN"), (flats_all[0] if flats_all else None))
    flat = request.args.get("flat") or ("ALL" if floor == "ALL" else default_flat)

    if floor == "ALL":
        scope_flat = dff.copy()
        scope_floor = dfl.copy()
    else:
        if (flat or "").upper() == "ALL":
            scope_flat = dff.copy()
        else:
            scope_flat = dff[dff["__Flat"].astype(str).fillna("").replace("", "UNKNOWN").eq(flat or "")].copy()
        scope_floor = dfl.copy()

    def _latest_per_checklist(frame: pd.DataFrame) -> pd.DataFrame:
        if frame is None or frame.empty:
            return frame.iloc[0:0]
        tmp = frame.copy()
        tmp["__ChecklistKey"] = tmp["__Checklist"].astype(str).str.strip().str.lower()
        tmp["__DateOrd"] = pd.to_datetime(tmp["__Date"]).fillna(pd.Timestamp.min)
        idxs = tmp.sort_values("__DateOrd").groupby("__ChecklistKey").tail(1).index
        return tmp.loc[idxs]

    def _is_rcc_checklist(name: str) -> bool:
        import re as _re
        s = (name or "").lower()
        rcc_patterns = [
            r"\brcc\b|reinforc|shuttering|\bstruct|slab|column|beam|footing",
            r"aluform|internal\s*handover|handover",
        ]
        for pat in rcc_patterns:
            try:
                if _re.search(pat, s, _re.I):
                    return True
            except Exception:
                continue
        return False

    def _base_name(n: str) -> str:
        import re as _re
        if not isinstance(n, str):
            return str(n)
        s = n.strip()
        s_norm = s.lower()
        
        # Known standardized names
        fixes = {
            r'painting.*internal': 'Painting Works : Internal',
            r'painting.*external': 'Painting Works : External',
            r'waterproof.*boxtype': 'Waterproofing works: Toilet and Skirting',
            r'waterproof.*toilet': 'Waterproofing works: Toilet and Skirting',
            r'waterproof.*skirting': 'Waterproofing works: Toilet and Skirting',
            r'tiling.*kitchen.*platform': 'Tiling - Kitchen Platform',
            r'tiling.*kitchen.*sink': 'Tiling - Kitchen Platform',
            r'tiling[-\s]*toilet.*dado': 'Tiling - Toilet Dado',
            r'kitchen\s*dado': 'Kitchen Dado Checklist',
            r'gypsum\s*plaster': 'Gypsum Plaster Works',
            r'aac\s*block': 'AAC Block Work',
            r'\brcc\b.*footing': 'RCC Footing',
            r'\brcc\b.*column': 'RCC Column',
            r'\brcc\b.*beam': 'RCC Beam',
            r'\brcc\b.*slab': 'RCC Slab',
        }
        for pat, nm in fixes.items():
            if _re.search(pat, s_norm):
                return nm
        
        # Remove contractor/team name at the end (common patterns)
        # E.g., "GYPSUM PLASTER WORKS  Faujan" -> "GYPSUM PLASTER WORKS"
        s = _re.sub(r'\s{2,}[A-Za-z]+\s*$', '', s)  # Double space followed by word
        s = _re.sub(r'\s+[A-Z][a-z]+\s+[A-Z][a-z]+\s*$', '', s)  # "Om Enterprises" pattern
        s = _re.sub(r'\s*:\s*$', '', s)  # Trailing colon
        s = _re.sub(r'\s+$', '', s)  # Trailing spaces
        
        if '.' in s:
            return s.split('.', 1)[0].strip()
        if ':' in s and s.split(':', 1)[1].strip():
            return s  # Keep if there's text after colon
        elif ':' in s:
            return s.split(':', 1)[0].strip()
        
        return s.strip()

    if floor == "ALL":
        reps_flat = scope_flat.iloc[0:0]
        reps_floor = scope_floor.iloc[0:0]
    else:
        reps_flat = _latest_per_checklist(scope_flat)
        reps_floor = _latest_per_checklist(scope_floor)

    rcc_items: _List[Dict] = []
    finishing_items_by_cat: Dict[_Tuple[int, str], _List[Dict]] = {}

    # Combine all reps from both floor and flat level
    all_reps = pd.concat([reps_floor, reps_flat], ignore_index=True) if (not reps_floor.empty and not reps_flat.empty) else (reps_floor if not reps_floor.empty else reps_flat)
    
    if all_reps is not None and not all_reps.empty:
        for _, row in all_reps.iterrows():
            cl_raw = row["__Checklist"]
            cl = _base_name(cl_raw)
            url = row.get("__URL", "")
            stage = row.get("__Stage", "")
            status = str(row.get("Status", "")).strip()
            
            # Determine category based on checklist type, not source
            if _is_rcc_checklist(cl_raw):
                # RCC checklists (structural/floor-level) go to RCC table
                rcc_items.append({
                    "checklist": cl,
                    "checklist_raw": cl_raw,
                    "url": url,
                    "stage": stage,
                    "status": status,
                    "date": row.get("__Date"),
                })
            else:
                # Finishing checklists (flat-level) go to Finishing categories
                idx, cat = _map_checklist_to_category(cl_raw)
                if (idx, cat) not in finishing_items_by_cat:
                    finishing_items_by_cat[(idx, cat)] = []
                finishing_items_by_cat[(idx, cat)].append({
                    "checklist": cl,
                    "checklist_raw": cl_raw,
                    "url": url,
                    "stage": stage,
                    "status": status,
                    "date": row.get("__Date"),
                })

    ordered_finishing_keys: _List[_Tuple[int, str]] = []
    for i, (label, _) in enumerate(CATEGORY_ORDER):
        if i == 0:
            continue
        ordered_finishing_keys.append((i, label))
    ordered_finishing_keys.append((len(CATEGORY_ORDER), UNCATEGORIZED_LABEL))

    return render_template(
        "flat_report.html",
        path=path,
        projects=projects,
        selected_project=proj,
        buildings=buildings_all,
        selected_building=building,
        floors=floors_all,
        selected_floor=floor,
        flats=flats_all,
        selected_flat=flat,
        rcc_items=rcc_items,
        finishing_items_by_cat=finishing_items_by_cat,
        ordered_finishing_keys=ordered_finishing_keys,
        show_only_export=(floor == "ALL"),
    )


@app.get("/flat-report/export")
@login_required
def flat_report_export():
    # Build an Excel status grid with separate sheets for RCC, Aluform, Handover, and Finishing
    path = request.args.get("path") or _get_session_eqc_path() or "Combined_EQC.csv"
    if not os.path.exists(path):
        return f"EQC file not found: {path}", 404
    df = _prepare_frame(_robust_read_eqc_path(path))
    
    # Debug: log total rows read
    print(f"[DEBUG] Total rows read from CSV: {len(df)}")
    print(f"[DEBUG] CSV columns: {list(df.columns)[:10]}...")  # Show first 10 columns
    
    projects = sorted([p for p in df["__Project"].astype(str).str.strip().unique() if p])
    print(f"[DEBUG] All projects in data: {projects}")
    
    proj = request.args.get("project") or (projects[0] if projects else None)
    print(f"[DEBUG] Project filter requested: '{proj}'")
    
    # Try exact match first
    dfp = df[df["__Project"].astype(str).str.strip().eq(proj)] if proj else df.iloc[0:0]
    
    # If no exact match and project provided, try case-insensitive
    if len(dfp) == 0 and proj:
        print(f"[DEBUG] No exact project match, trying case-insensitive...")
        dfp = df[df["__Project"].astype(str).str.strip().str.lower().eq(proj.lower())]
    
    # If still no match, try partial match (contains)
    if len(dfp) == 0 and proj:
        print(f"[DEBUG] No case-insensitive match, trying partial match...")
        dfp = df[df["__Project"].astype(str).str.strip().str.contains(proj, case=False, na=False)]
    
    print(f"[DEBUG] Project: '{proj}', Rows after project filter: {len(dfp)}")
    if len(dfp) > 0:
        print(f"[DEBUG] Sample __Project values after filter: {dfp['__Project'].head(3).tolist()}")
    
    building = request.args.get("building") or ""
    print(f"[DEBUG] Building filter requested: '{building}'")
    available_buildings = sorted([b for b in dfp['__Building'].unique() if str(b).strip()])
    print(f"[DEBUG] Available buildings in data: {available_buildings}")
    
    if len(dfp) > 0 and len(available_buildings) > 0:
        print(f"[DEBUG] Sample Location L1 values: {dfp['Location L1'].head(3).tolist() if 'Location L1' in dfp.columns else 'N/A'}")
        print(f"[DEBUG] Sample __Building values: {dfp['__Building'].head(3).tolist()}")
    
    # Try exact match first
    dfl = dfp[dfp["__Building"].astype(str).fillna("").replace("", "UNKNOWN").eq(building or "")]
    
    # If no exact match, try case-insensitive match
    if len(dfl) == 0 and building:
        print(f"[DEBUG] No exact match found, trying case-insensitive match...")
        dfl = dfp[dfp["__Building"].astype(str).str.lower().fillna("").replace("", "unknown").eq(building.lower())]
    
    # If still no match, try partial match
    if len(dfl) == 0 and building:
        print(f"[DEBUG] No case-insensitive match, trying partial match...")
        dfl = dfp[dfp["__Building"].astype(str).str.contains(building, case=False, na=False)]
    
    print(f"[DEBUG] Building: '{building}', Rows after building filter: {len(dfl)}")
    if len(dfl) > 0:
        print(f"[DEBUG] Sample __Building values after filter: {dfl['__Building'].head(3).tolist()}")
    
    floor = request.args.get("floor") or ""
    print(f"[DEBUG] Floor filter requested: '{floor}'")
    available_floors = sorted([f for f in dfl['__Floor'].unique() if str(f).strip()])
    print(f"[DEBUG] Available floors in data: {available_floors}")
    
    if len(dfl) > 0 and len(available_floors) > 0:
        print(f"[DEBUG] Sample Location L2 values: {dfl['Location L2'].head(3).tolist() if 'Location L2' in dfl.columns else 'N/A'}")
        print(f"[DEBUG] Sample __Floor values: {dfl['__Floor'].head(3).tolist()}")
    
    if floor == "ALL":
        dff = dfl.copy()
    else:
        dff = dfl[dfl["__Floor"].astype(str).fillna("").replace("", "UNKNOWN").eq(floor or "")]
    
    # If no match and floor specified, try case-insensitive
    if len(dff) == 0 and floor and floor != "ALL":
        print(f"[DEBUG] No exact floor match, trying case-insensitive...")
        dff = dfl[dfl["__Floor"].astype(str).str.lower().fillna("").replace("", "unknown").eq(floor.lower())]
    
    # If still no match, try partial match
    if len(dff) == 0 and floor and floor != "ALL":
        print(f"[DEBUG] No case-insensitive floor match, trying partial match...")
        dff = dfl[dfl["__Floor"].astype(str).str.contains(floor, case=False, na=False)]
    
    print(f"[DEBUG] Floor: '{floor}', Rows after floor filter: {len(dff)}")
    if len(dff) > 0:
        print(f"[DEBUG] Sample __Floor values after filter: {dff['__Floor'].head(3).tolist()}")
        print(f"[DEBUG] Sample __Checklist values: {dff['__Checklist'].head(3).tolist()}")
    
    # If no data found, provide helpful error message
    if len(dff) == 0:
        print(f"[ERROR] No data found after filtering!")
        print(f"[ERROR] Filters: Project='{proj}', Building='{building}', Floor='{floor}'")
        
        error_msg = f"<h3>No data found for the selected filters</h3>"
        error_msg += f"<p><strong>Selected Filters:</strong><br>"
        error_msg += f"Project: {proj}<br>"
        error_msg += f"Building: {building}<br>"
        error_msg += f"Floor: {floor}</p>"
        
        if len(df) > 0:
            all_projects = sorted([p for p in df["__Project"].unique() if str(p).strip()])
            error_msg += f"<p><strong>Available Projects:</strong> {', '.join(all_projects)}</p>"
        
        if len(dfp) > 0:
            available_buildings = sorted([b for b in dfp["__Building"].unique() if str(b).strip()])
            error_msg += f"<p><strong>Available Buildings for '{proj}':</strong> {', '.join(available_buildings)}</p>"
        
        if len(dfl) > 0:
            available_floors = sorted([f for f in dfl["__Floor"].unique() if str(f).strip()])
            error_msg += f"<p><strong>Available Floors for '{building}':</strong> {', '.join(available_floors)}</p>"
        else:
            error_msg += f"<p><strong style='color:red;'>No data found for Building '{building}' in Project '{proj}'</strong></p>"
            if len(dfp) > 0:
                error_msg += f"<p>This could mean:<br>"
                error_msg += f"1. The building name doesn't match exactly<br>"
                error_msg += f"2. No data exists for this building in the CSV file<br>"
                error_msg += f"3. The project filter didn't work correctly</p>"
        
        error_msg += f"<p><em>Check the terminal/console for detailed debug logs</em></p>"
        return error_msg, 404

    # Determine flats on this floor
    def _flat_key(v: str):
        import re as _re
        s = str(v or "").strip()
        m = _re.search(r"(shop|flat|unit|apt|apartment)\s*([A-Z]?\d+)([A-Z]?)", s, _re.I)
        if m:
            num = int(_re.sub(r"\D", "", m.group(2))) if _re.sub(r"\D", "", m.group(2)) else 0
            suf = m.group(3).upper() if m.group(3) else ""
            return (0, num, suf)
        m2 = _re.search(r"(\d+)", s)
        if m2:
            return (1, int(m2.group(1)), "")
        return (9, 0, s.lower())
    flats = sorted(set(dff["__Flat"].astype(str).fillna("").replace("", "UNKNOWN")), key=_flat_key)
    
    # Helper function to build floor-level sheet (for RCC, Aluform, Handover)
    def build_floor_status_sheet(tmp_data: pd.DataFrame) -> pd.DataFrame:
        """Build floor-level status list for RCC/Aluform/Handover checklists (not flat-wise)."""
        if tmp_data.empty:
            return pd.DataFrame()
        
        tmp = tmp_data.copy()
        tmp["__ChecklistKey"] = tmp["__Checklist"].astype(str).str.strip().str.lower()
        tmp["__DateOrd"] = pd.to_datetime(tmp["__Date"]).fillna(pd.Timestamp.min)
        
        # CRITICAL: Group by checklist+building+floor+pour to preserve ALL unique inspections
        # Don't deduplicate across different pours/locations - each is a separate inspection!
        group_cols = ["__ChecklistKey", "__Building", "__Floor", "__Pour"]
        
        # Check for duplicates before grouping
        print(f"\n[DEBUG build_floor_status_sheet] Before grouping:")
        print(f"  Total input rows: {len(tmp_data)}")
        print(f"  Unique combinations (Checklist+Building+Floor+Pour): {tmp.groupby(group_cols, dropna=False).ngroups}")
        
        # Count how many groups have multiple entries (duplicates to be consolidated)
        group_sizes = tmp.groupby(group_cols, dropna=False).size()
        duplicates = group_sizes[group_sizes > 1]
        if len(duplicates) > 0:
            print(f"  Found {len(duplicates)} groups with multiple inspections (will keep latest):")
            for idx, (group_key, count) in enumerate(duplicates.items()):
                if idx < 5:  # Show first 5
                    checklist, bldg, flr, pour = group_key
                    print(f"    • '{checklist}' at {bldg}/{flr}/Pour:{pour} - {count} inspections")
            if len(duplicates) > 5:
                print(f"    ... and {len(duplicates) - 5} more duplicate groups")
        
        tmp = tmp.sort_values("__DateOrd").groupby(group_cols, as_index=False, dropna=False).tail(1)
        
        print(f"  After grouping: {len(tmp)} rows (consolidated {len(tmp_data) - len(tmp)} duplicate inspections)")

        # Build list with checklist details
        def status_symbol(status: str) -> str:
            up = (status or "").strip().upper()
            # Be more lenient with PASSED detection
            if "PASS" in up and "FAIL" not in up:
                return "✓ Passed"
            if "PROGRESS" in up or "IN_PROGRESS" in up:
                return "– In Progress"
            if "REDO" in up:
                return "↻ Redo"
            if "FAIL" in up:
                return "✗ Failed"
            return "– Pending"
        
        def get_category(checklist: str) -> str:
            """Determine category for a checklist."""
            if is_rcc_checklist(checklist):
                return "RCC / Structural"
            elif is_aluform_checklist(checklist):
                return "Aluform"
            elif is_handover_checklist(checklist):
                return "Internal Handover"
            return "Other"
        
        def get_eqc_stage_status(stage_status: str) -> str:
            """Format EQC Stage Status for readability."""
            s = str(stage_status or "").strip()
            if not s or s == "-":
                return "–"
            # Make it more readable
            return s.replace("_", " ").title()

        grid = []
        for _, row in tmp.iterrows():
            cl_raw = str(row["__Checklist"])
            cl = base_name(cl_raw)
            category = get_category(cl_raw)
            status = status_symbol(str(row.get("Status", "")))
            stage = str(row.get("__Stage", ""))
            date = str(row.get("__Date", ""))
            pour = str(row.get("__Pour", "")).strip() if "__Pour" in row and row.get("__Pour") else ""
            
            # Add more detailed information
            inspector = str(row.get("Inspector", "")).strip() if "Inspector" in row else "–"
            team = str(row.get("Team", "")).strip() if "Team" in row else "–"
            approver = str(row.get("Approver", "-")).strip()
            if not approver or approver == "-":
                approver = "Pending"
            eqc_stage = get_eqc_stage_status(row.get("EQC Stage Status", ""))
            
            # Build full location for context
            loc_parts = []
            for loc_col in ["Location L1", "Location L2", "Location Variable", "Location L4"]:
                if loc_col in row:
                    val = str(row.get(loc_col, "")).strip()
                    if val:
                        loc_parts.append(val)
            full_location = " / ".join(loc_parts) if loc_parts else "–"
            
            grid.append({
                "Category": category,
                "Checklist": cl,
                "Location": full_location,
                "Pour": pour if pour else "–",
                "Stage": stage,
                "Status": status,
                "EQC Stage": eqc_stage,
                "Date": date,
                "Inspector": inspector,
                "Team": team,
                "Approver": approver,
            })

        return pd.DataFrame(grid, columns=["Category", "Checklist", "Location", "Pour", "Stage", "Status", "EQC Stage", "Date", "Inspector", "Team", "Approver"]) if grid else pd.DataFrame()
    
    # Helper function to build flat-wise sheet (for Finishing checklists)
    def build_status_sheet(tmp_data: pd.DataFrame) -> pd.DataFrame:
        """Build status matrix for flat-level checklists (Finishing)."""
        if tmp_data.empty:
            return pd.DataFrame()
        
        tmp = tmp_data.copy()
        tmp["__ChecklistKey"] = tmp["__Checklist"].astype(str).str.strip().str.lower()
        tmp["__FlatKey"] = tmp["__Flat"].astype(str).fillna("").replace("", "UNKNOWN")
        tmp["__DateOrd"] = pd.to_datetime(tmp["__Date"]).fillna(pd.Timestamp.min)
        tmp = tmp.sort_values("__DateOrd").groupby(["__FlatKey", "__ChecklistKey"], as_index=False).tail(1)

        # Determine checklist columns as union across this floor (use readable names)
        checklist_cols = list(dict.fromkeys(tmp["__Checklist"].astype(str).tolist()))

        # Build status matrix
        def status_symbol(status: str) -> str:
            up = (status or "").strip().upper()
            # Be more lenient with PASSED detection
            if "PASS" in up and "FAIL" not in up:
                return "✓"
            if "PROGRESS" in up or "IN_PROGRESS" in up:
                return "–"
            return "–" if up else "–"

        grid = []
        for flt in flats:
            row = {"Flat": flt}
            sub = tmp[tmp["__FlatKey"].eq(flt)]
            present = set(sub["__Checklist"].astype(str))
            any_present = bool(present)
            all_complete = True
            for col in checklist_cols:
                rec =sub[sub["__Checklist"].astype(str).eq(col)]
                if rec.empty:
                    row[col] = "✗"
                    all_complete = False
                else:
                    sym = status_symbol(str(rec.iloc[0].get("Status", "")))
                    row[col] = sym
                    if sym != "✓":
                        all_complete = False
            # Overall status per flat
            row["Overall"] = "✓" if all_complete and any_present else ("–" if any_present else "✗")
            grid.append(row)

        return pd.DataFrame(grid, columns=["Flat", "Overall"] + checklist_cols) if grid else pd.DataFrame()
    
    # Separate data by checklist type
    # Combine RCC, Aluform, and Handover into one category (all floor-level)
    
    # Detailed classification debugging
    print(f"\n[DEBUG CLASSIFICATION] Starting checklist classification...")
    unique_checklists = dff["__Checklist"].unique()
    print(f"[DEBUG CLASSIFICATION] Total unique checklist types: {len(unique_checklists)}")
    
    # Classify each unique checklist and log
    structural_checklists = []
    finishing_checklists = []
    
    for cl in unique_checklists:
        cl_str = str(cl).strip()
        is_structural = is_rcc_checklist(cl_str) or is_aluform_checklist(cl_str) or is_handover_checklist(cl_str)
        if is_structural:
            structural_checklists.append(cl_str)
        else:
            finishing_checklists.append(cl_str)
    
    print(f"\n[DEBUG CLASSIFICATION] Structural checklist types ({len(structural_checklists)}):")
    for cl in sorted(structural_checklists):
        print(f"  ✓ {cl}")
    
    print(f"\n[DEBUG CLASSIFICATION] Finishing checklist types ({len(finishing_checklists)}):")
    for cl in sorted(finishing_checklists)[:10]:  # Show first 10
        print(f"  • {cl}")
    if len(finishing_checklists) > 10:
        print(f"  ... and {len(finishing_checklists) - 10} more")
    
    # Check for potential misclassifications - checklists with "RCC" or "Aluform" etc in name but classified as finishing
    import re as _re_check
    potential_misclassifications = []
    for cl in finishing_checklists:
        cl_lower = cl.lower()
        if any(keyword in cl_lower for keyword in ['rcc', 'aluform', 'reinforc', 'shuttering', 'struct', 'footing', 'handover']):
            potential_misclassifications.append(cl)
    
    if potential_misclassifications:
        print(f"\n[WARNING] Found {len(potential_misclassifications)} potentially misclassified checklists:")
        for cl in potential_misclassifications:
            print(f"  ⚠ '{cl}' - classified as FINISHING but contains structural keywords!")
    
    structural_data = dff[dff["__Checklist"].apply(lambda x: is_rcc_checklist(x) or is_aluform_checklist(x) or is_handover_checklist(x))]
    finishing_data = dff[~dff["__Checklist"].apply(lambda x: is_rcc_checklist(x) or is_aluform_checklist(x) or is_handover_checklist(x))]
    
    # Data validation: ensure no checklists are lost
    total_checklists = len(dff)
    structural_count = len(structural_data)
    finishing_count = len(finishing_data)
    data_validation_passed = (structural_count + finishing_count == total_checklists)
    
    # Debug logging
    print(f"\n[DEBUG ROW COUNTS] Total rows after filter: {total_checklists}")
    print(f"[DEBUG ROW COUNTS] Structural rows: {structural_count}")
    print(f"[DEBUG ROW COUNTS] Finishing rows: {finishing_count}")
    print(f"[DEBUG ROW COUNTS] Validation: {structural_count} + {finishing_count} = {structural_count + finishing_count} (Expected: {total_checklists})")
    
    if structural_count + finishing_count != total_checklists:
        print(f"[ERROR] MISMATCH! {abs(total_checklists - structural_count - finishing_count)} rows unaccounted for!")
    
    # Build status DataFrames for each category
    # Structural (RCC, Aluform, Handover combined) - floor-level (not flat-wise)
    structural_df = build_floor_status_sheet(structural_data)
    # Finishing is flat-wise
    finishing_df = build_status_sheet(finishing_data)
    
    # Count ACTUAL output rows (after processing)
    structural_output_count = len(structural_df)
    finishing_output_count = len(finishing_df)
    total_output_rows = structural_output_count + finishing_output_count
    
    # Count unique finishing checklist types (columns in matrix)
    finishing_checklist_types = 0
    if not finishing_df.empty and len(finishing_df.columns) > 2:
        # Columns: Flat, Overall, ...checklist types...
        finishing_checklist_types = len(finishing_df.columns) - 2
    
    print(f"\n[DEBUG OUTPUT COUNTS]")
    print(f"  Structural sheet rows: {structural_output_count} (each row = 1 inspection)")
    print(f"  Finishing sheet rows: {finishing_output_count} (each row = 1 flat, columns = checklists)")
    print(f"  Finishing checklist types (columns): {finishing_checklist_types}")
    print(f"  Total output rows: {total_output_rows}")
    print(f"\n[IMPORTANT NOTE]")
    print(f"  Structural: INPUT {structural_count} inspections → OUTPUT {structural_output_count} rows")
    if structural_count != structural_output_count:
        print(f"    ⚠ Consolidated {structural_count - structural_output_count} duplicate inspections")
    print(f"  Finishing: INPUT {finishing_count} checklist entries → OUTPUT {finishing_output_count} flat rows (matrix format)")
    print(f"    (Finishing sheet is a MATRIX: {finishing_output_count} flats × {finishing_checklist_types} checklist types)")
    
    # Count unique flats to validate finishing matrix
    if not finishing_data.empty:
        unique_flats_in_finishing = finishing_data["__Flat"].nunique()
        print(f"    Unique flats in finishing data: {unique_flats_in_finishing}")
        if unique_flats_in_finishing != finishing_output_count:
            print(f"    ⚠ Flat count mismatch: {unique_flats_in_finishing} unique flats but {finishing_output_count} matrix rows!")

    # Write Excel with separate sheets
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import PatternFill, Font, Border, Side
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove the default sheet
    
    # Add Summary sheet first for data validation
    summary_ws = wb.create_sheet(title="Summary", index=0)
    summary_ws.cell(row=1, column=1, value="Data Export Summary")
    summary_ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    summary_ws.cell(row=3, column=1, value="Project:")
    summary_ws.cell(row=3, column=2, value=proj)
    summary_ws.cell(row=4, column=1, value="Building:")
    summary_ws.cell(row=4, column=2, value=building)
    summary_ws.cell(row=5, column=1, value="Floor:")
    summary_ws.cell(row=5, column=2, value=(floor or 'ALL'))
    summary_ws.cell(row=6, column=1, value="Export Date:")
    summary_ws.cell(row=6, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    
    summary_ws.cell(row=8, column=1, value="Data Validation:")
    summary_ws.cell(row=8, column=1).font = Font(bold=True)
    summary_ws.cell(row=9, column=1, value="CSV Raw Rows (Input):")
    summary_ws.cell(row=9, column=2, value=total_checklists)
    summary_ws.cell(row=10, column=1, value="  - Structural/Handover:")
    summary_ws.cell(row=10, column=2, value=structural_count)
    summary_ws.cell(row=11, column=1, value="  - Finishing:")
    summary_ws.cell(row=11, column=2, value=finishing_count)
    
    summary_ws.cell(row=13, column=1, value="Excel Rows (Output):")
    summary_ws.cell(row=13, column=1).font = Font(bold=True)
    summary_ws.cell(row=14, column=1, value="  - Structural/Handover:")
    summary_ws.cell(row=14, column=2, value=f"{structural_output_count} inspection records")
    summary_ws.cell(row=15, column=1, value="  - Finishing:")
    summary_ws.cell(row=15, column=2, value=f"{finishing_output_count} flat rows × {finishing_checklist_types} checklist types (matrix)")
    summary_ws.cell(row=16, column=1, value="  - Total Rows:")
    summary_ws.cell(row=16, column=2, value=f"{total_output_rows} (not directly comparable to input)")
    summary_ws.cell(row=16, column=1).font = Font(bold=True)
    
    summary_ws.cell(row=18, column=1, value="Validation Status:")
    if total_checklists == 0:
        summary_ws.cell(row=18, column=2, value="⚠ WARNING: No data found for selected filters!")
        summary_ws.cell(row=18, column=2).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    elif data_validation_passed:
        summary_ws.cell(row=18, column=2, value="✓ All input data categorized (No missing checklists)")
        summary_ws.cell(row=18, column=2).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    else:
        summary_ws.cell(row=18, column=2, value=f"⚠ Warning: {total_checklists - structural_count - finishing_count} rows uncategorized!")
        summary_ws.cell(row=18, column=2).fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    
    summary_ws.cell(row=20, column=1, value="Sheet Organization:")
    summary_ws.cell(row=20, column=1).font = Font(bold=True)
    summary_ws.cell(row=21, column=1, value="• Structural & Handover:")
    summary_ws.cell(row=21, column=2, value="Floor-level inspections (RCC, Aluform, Internal Handover) - One row per unique inspection")
    summary_ws.cell(row=22, column=1, value="• Finishing:")
    summary_ws.cell(row=22, column=2, value="Flat-wise matrix (Masonry, Plaster, Tiling, Painting, etc.)")
    
    summary_ws.cell(row=24, column=1, value="Note:")
    summary_ws.cell(row=24, column=1).font = Font(bold=True, color="FF0000")
    summary_ws.cell(row=25, column=1, value="⚠ IMPORTANT: Output row counts measure different things!")
    summary_ws.cell(row=25, column=1).font = Font(bold=True)
    summary_ws.cell(row=26, column=1, value="• Structural rows = Number of INSPECTIONS (one row per inspection)")
    summary_ws.cell(row=27, column=1, value="• Finishing rows = Number of FLATS (matrix: flat × checklist)")
    summary_ws.cell(row=28, column=1, value="")
    summary_ws.cell(row=29, column=1, value="Why output ≠ input:")
    summary_ws.cell(row=30, column=1, value="• Structural: Duplicate inspections consolidated (same checklist+floor+pour)")
    summary_ws.cell(row=31, column=1, value="• Finishing: Converts to flat-wise matrix (rows=flats, columns=checklists)")
    summary_ws.cell(row=32, column=1, value="• Example: 500 finishing inspections across 50 flats = 50 matrix rows")
    
    # Apply borders to summary
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in summary_ws.iter_rows(min_row=1, max_row=32, min_col=1, max_col=2):
        for cell in row:
            cell.border = border
    
    # Auto-adjust column widths in summary
    summary_ws.column_dimensions['A'].width = 30
    summary_ws.column_dimensions['B'].width = 80
    
    # Helper to add floor-level sheet (for RCC, Aluform, Handover)
    def add_floor_level_sheet(workbook, sheet_name: str, data_df: pd.DataFrame) -> None:
        """Add a floor-level sheet (not flat-wise breakdown)."""
        if data_df.empty:
            # Create sheet even if empty to show that no data exists
            ws = workbook.create_sheet(title=sheet_name)
            ws.cell(row=1, column=1, value="Building:")
            ws.cell(row=1, column=2, value=building)
            ws.cell(row=1, column=4, value="Floor:")
            ws.cell(row=1, column=5, value=(floor or 'ALL'))
            ws.cell(row=3, column=1, value="No data found for this category on the selected floor.")
            return
        
        ws = workbook.create_sheet(title=sheet_name)
        # Add building name header for clarity
        ws.cell(row=1, column=1, value="Building:")
        ws.cell(row=1, column=2, value=building)
        ws.cell(row=1, column=4, value="Floor:")
        ws.cell(row=1, column=5, value=(floor or 'ALL'))
        
        # Add note explaining floor-level nature
        ws.cell(row=1, column=7, value="Note:")
        ws.cell(row=1, column=8, value="These checklists apply at Floor/Pour level, not per flat")
        
        # Add data count for validation
        ws.cell(row=1, column=11, value="Total Rows:")
        ws.cell(row=1, column=12, value=len(data_df))
        
        # Leave row 2 blank, start data at row 3
        ws.append([])
        start_row = 3
        for r in dataframe_to_rows(data_df, index=False, header=True):
            ws.append(r)
        
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        red = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
        gray = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        # Apply coloring to Status column (6th column: Category, Checklist, Location, Pour, Stage, Status, ...)
        status_col_idx = 6
        for row in ws.iter_rows(min_row=start_row+1, min_col=status_col_idx, max_row=ws.max_row, max_col=status_col_idx):
            for cell in row:
                val = str(cell.value or "")
                if "✓" in val or "Passed" in val:
                    cell.fill = green
                elif "–" in val or "Progress" in val or "Pending" in val:
                    cell.fill = yellow
                elif "✗" in val or "Failed" in val:
                    cell.fill = red
                elif "↻" in val or "Redo" in val:
                    cell.fill = gray
        
        # Apply borders and wrap text
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        align = Alignment(wrap_text=True, vertical="center")
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                try:
                    cell.border = border
                    cell.alignment = align
                except Exception:
                    pass
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
    
    # Helper to add flat-level sheet (for Finishing)
    def add_flat_level_sheet(workbook, sheet_name: str, data_df: pd.DataFrame) -> None:
        """Add a flat-wise sheet (for Finishing checklists)."""
        if data_df.empty:
            # Create sheet even if empty to show that no data exists
            ws = workbook.create_sheet(title=sheet_name)
            ws.cell(row=1, column=1, value="Building:")
            ws.cell(row=1, column=2, value=building)
            ws.cell(row=1, column=4, value="Floor:")
            ws.cell(row=1, column=5, value=(floor or 'ALL'))
            ws.cell(row=3, column=1, value="No data found for this category on the selected floor.")
            return
        
        ws = workbook.create_sheet(title=sheet_name)
        # Add building name header for clarity
        ws.cell(row=1, column=1, value="Building:")
        ws.cell(row=1, column=2, value=building)
        ws.cell(row=1, column=4, value="Floor:")
        ws.cell(row=1, column=5, value=(floor or 'ALL'))
        
        # Add legend for symbols
        ws.cell(row=1, column=7, value="Legend:")
        ws.cell(row=1, column=8, value="✓ = Passed")
        ws.cell(row=1, column=9, value="– = In Progress")
        ws.cell(row=1, column=10, value="✗ = Not Done")
        
        # Add data count for validation (number of flats shown)
        ws.cell(row=1, column=12, value="Flats:")
        ws.cell(row=1, column=13, value=len(data_df))
        
        # Leave row 2 blank, start data at row 3
        ws.append([])
        start_row = 3
        for r in dataframe_to_rows(data_df, index=False, header=True):
            ws.append(r)
        
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        red = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
        
        # Apply coloring only over data rows (excluding header at start_row)
        for row in ws.iter_rows(min_row=start_row+1, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                val = str(cell.value or "")
                if val == "✓":
                    cell.fill = green
                elif val == "–":
                    cell.fill = yellow
                elif val == "✗":
                    cell.fill = red
        
        # Apply borders and wrap text
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        align = Alignment(wrap_text=True, vertical="center")
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                try:
                    cell.border = border
                    cell.alignment = align
                except Exception:
                    pass
    
    # Add sheets in order
    # Structural (RCC, Aluform, Handover combined) - floor-level (not flat-wise)
    add_floor_level_sheet(wb, "Structural & Handover", structural_df)
    # Finishing is flat-wise
    add_flat_level_sheet(wb, "Finishing", finishing_df)
    
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    
    # Sanitize building name for filename (extract key identifier)
    import re as _re_fname
    building_clean = building or "All"
    # Try to extract just the letter/number (e.g., "A wing" -> "A", "Building B" -> "B")
    m = _re_fname.search(r'\b([A-Z]\d?|\d+[A-Z]?)\b', building_clean)
    if m:
        building_clean = m.group(1)
    else:
        # Just remove spaces and special chars
        building_clean = _re_fname.sub(r'[^\w]', '_', building_clean)
    
    floor_clean = (floor or 'ALL').replace(" ", "_")
    fname = f"Flat_Status_{proj}_{building_clean}_{floor_clean}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    # Remove any remaining problematic characters
    fname = fname.replace("/", "-").replace("\\", "-")
    
    return send_file(bio, as_attachment=True, download_name=fname, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/flat-report/export-project")
@login_required
def flat_report_export_project():
    """Export flat-wise report for ALL buildings in the selected project"""
    path = request.args.get("path") or _get_session_eqc_path() or "Combined_EQC.csv"
    if not os.path.exists(path):
        return f"EQC file not found: {path}", 404
    df = _prepare_frame(_robust_read_eqc_path(path))
    
    projects = sorted([p for p in df["__Project"].astype(str).str.strip().unique() if p])
    proj = request.args.get("project") or (projects[0] if projects else None)
    
    # Filter to selected project
    dfp = df[df["__Project"].astype(str).str.strip().eq(proj)] if proj else df.iloc[0:0]
    if len(dfp) == 0 and proj:
        dfp = df[df["__Project"].astype(str).str.strip().str.lower().eq(proj.lower())]
    if len(dfp) == 0 and proj:
        dfp = df[df["__Project"].astype(str).str.strip().str.contains(proj, case=False, na=False)]
    
    if len(dfp) == 0:
        return f"No data found for project: {proj}", 404
    
    # Get all buildings in this project
    all_buildings = sorted([b for b in dfp['__Building'].unique() if str(b).strip()])
    
    if not all_buildings:
        return f"No buildings found for project: {proj}", 404
    
    print(f"[DEBUG] Exporting project '{proj}' with {len(all_buildings)} buildings: {all_buildings}")
    
    # Create workbook
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # Add summary sheet
    summary_ws = wb.create_sheet(title="Summary", index=0)
    summary_ws.cell(row=1, column=1, value="Project-Wide Export Summary")
    summary_ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    summary_ws.cell(row=3, column=1, value="Project:")
    summary_ws.cell(row=3, column=2, value=proj)
    summary_ws.cell(row=4, column=1, value="Buildings:")
    summary_ws.cell(row=4, column=2, value=len(all_buildings))
    summary_ws.cell(row=5, column=1, value="Export Date:")
    summary_ws.cell(row=5, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    
    summary_ws.cell(row=7, column=1, value="Buildings Included:")
    summary_ws.cell(row=7, column=1).font = Font(bold=True)
    for idx, bldg in enumerate(all_buildings, start=8):
        summary_ws.cell(row=idx, column=1, value=f"{idx-7}.")
        summary_ws.cell(row=idx, column=2, value=bldg)
    
    # Apply borders to summary
    thin = Side(border_style="thin", color="000000")
    border_style = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in summary_ws.iter_rows(min_row=1, max_row=7+len(all_buildings), min_col=1, max_col=2):
        for cell in row:
            cell.border = border_style
    
    summary_ws.column_dimensions['A'].width = 20
    summary_ws.column_dimensions['B'].width = 40
    
    # Helper function to build floor-level status sheet (Structural/Handover)
    def build_floor_status_sheet_for_building(tmp_data: pd.DataFrame) -> pd.DataFrame:
        """Build floor-level status list for RCC/Aluform/Handover checklists."""
        if tmp_data.empty:
            return pd.DataFrame()
        
        tmp = tmp_data.copy()
        tmp["__ChecklistKey"] = tmp["__Checklist"].astype(str).str.strip().str.lower()
        tmp["__DateOrd"] = pd.to_datetime(tmp["__Date"]).fillna(pd.Timestamp.min)
        
        group_cols = ["__ChecklistKey", "__Building", "__Floor", "__Pour"]
        tmp = tmp.sort_values("__DateOrd").groupby(group_cols, as_index=False, dropna=False).tail(1)
        
        # Debug: Check status values
        if "Status" in tmp.columns:
            status_values = tmp["Status"].value_counts()
            print(f"[DEBUG] Status values in structural data: {status_values.to_dict()}")
        
        def status_symbol(status: str) -> str:
            up = (status or "").strip().upper()
            # Be more lenient with PASSED detection
            if "PASS" in up and "FAIL" not in up:
                return "✓ Passed"
            if "PROGRESS" in up or "IN_PROGRESS" in up:
                return "– In Progress"
            if "REDO" in up:
                return "↻ Redo"
            if "FAIL" in up:
                return "✗ Failed"
            return "– Pending"
        
        def get_category(checklist: str) -> str:
            if is_rcc_checklist(checklist):
                return "RCC / Structural"
            elif is_aluform_checklist(checklist):
                return "Aluform"
            elif is_handover_checklist(checklist):
                return "Internal Handover"
            return "Other"
        
        def get_eqc_stage_status(stage_status: str) -> str:
            s = str(stage_status or "").strip()
            if not s or s == "-":
                return "–"
            return s.replace("_", " ").title()

        grid = []
        for _, row in tmp.iterrows():
            cl_raw = str(row["__Checklist"])
            cl = base_name(cl_raw)
            category = get_category(cl_raw)
            status = status_symbol(str(row.get("Status", "")))
            stage = str(row.get("__Stage", ""))
            date = str(row.get("__Date", ""))
            pour = str(row.get("__Pour", "")).strip() if "__Pour" in row and row.get("__Pour") else ""
            
            inspector = str(row.get("Inspector", "")).strip() if "Inspector" in row else "–"
            team = str(row.get("Team", "")).strip() if "Team" in row else "–"
            approver = str(row.get("Approver", "-")).strip()
            if not approver or approver == "-":
                approver = "Pending"
            eqc_stage = get_eqc_stage_status(row.get("EQC Stage Status", ""))
            
            loc_parts = []
            for loc_col in ["Location L1", "Location L2", "Location Variable", "Location L4"]:
                if loc_col in row:
                    val = str(row.get(loc_col, "")).strip()
                    if val:
                        loc_parts.append(val)
            full_location = " / ".join(loc_parts) if loc_parts else "–"
            
            grid.append({
                "Category": category,
                "Checklist": cl,
                "Location": full_location,
                "Pour": pour if pour else "–",
                "Stage": stage,
                "Status": status,
                "EQC Stage": eqc_stage,
                "Date": date,
                "Inspector": inspector,
                "Team": team,
                "Approver": approver,
            })

        return pd.DataFrame(grid, columns=["Category", "Checklist", "Location", "Pour", "Stage", "Status", "EQC Stage", "Date", "Inspector", "Team", "Approver"]) if grid else pd.DataFrame()
    
    # Helper function to build flat-wise status matrix (Finishing)
    def build_flat_status_matrix_for_building(tmp_data: pd.DataFrame, building_flats: list) -> pd.DataFrame:
        """Build status matrix for flat-level checklists (Finishing)."""
        if tmp_data.empty:
            return pd.DataFrame()
        
        tmp = tmp_data.copy()
        tmp["__ChecklistKey"] = tmp["__Checklist"].astype(str).str.strip().str.lower()
        tmp["__FlatKey"] = tmp["__Flat"].astype(str).fillna("").replace("", "UNKNOWN")
        tmp["__DateOrd"] = pd.to_datetime(tmp["__Date"]).fillna(pd.Timestamp.min)
        tmp = tmp.sort_values("__DateOrd").groupby(["__FlatKey", "__ChecklistKey"], as_index=False).tail(1)

        # Debug: Check status values
        if "Status" in tmp.columns:
            status_values = tmp["Status"].value_counts()
            print(f"[DEBUG] Status values in finishing data: {status_values.to_dict()}")

        # Determine checklist columns
        checklist_cols = list(dict.fromkeys(tmp["__Checklist"].astype(str).tolist()))

        def status_symbol(status: str) -> str:
            up = (status or "").strip().upper()
            # Be more lenient with PASSED detection
            if "PASS" in up and "FAIL" not in up:
                return "✓"
            if "PROGRESS" in up or "IN_PROGRESS" in up:
                return "–"
            return "–" if up else "–"

        grid = []
        for flt in building_flats:
            row = {"Flat": flt}
            sub = tmp[tmp["__FlatKey"].eq(flt)]
            present = set(sub["__Checklist"].astype(str))
            any_present = bool(present)
            all_complete = True
            for col in checklist_cols:
                rec = sub[sub["__Checklist"].astype(str).eq(col)]
                if rec.empty:
                    row[col] = "✗"
                    all_complete = False
                else:
                    sym = status_symbol(str(rec.iloc[0].get("Status", "")))
                    row[col] = sym
                    if sym != "✓":
                        all_complete = False
            row["Overall"] = "✓" if all_complete and any_present else ("–" if any_present else "✗")
            grid.append(row)

        return pd.DataFrame(grid, columns=["Flat", "Overall"] + checklist_cols) if grid else pd.DataFrame()
    
    # Helper to add floor-level sheet (Structural/Handover)
    def add_floor_level_sheet_project(workbook, sheet_name: str, building_name: str, data_df: pd.DataFrame):
        if data_df.empty:
            ws = workbook.create_sheet(title=sheet_name)
            ws.cell(row=1, column=1, value=f"Building: {building_name}")
            ws.cell(row=3, column=1, value="No structural/handover data found for this building.")
            return
        
        ws = workbook.create_sheet(title=sheet_name)
        ws.cell(row=1, column=1, value="Building:")
        ws.cell(row=1, column=2, value=building_name)
        ws.cell(row=1, column=7, value="Note:")
        ws.cell(row=1, column=8, value="Floor/Pour level checklists")
        ws.cell(row=1, column=11, value="Total:")
        ws.cell(row=1, column=12, value=len(data_df))
        
        ws.append([])
        start_row = 3
        for r in dataframe_to_rows(data_df, index=False, header=True):
            ws.append(r)
        
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        red = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
        gray = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        status_col_idx = 6  # Status column
        for row in ws.iter_rows(min_row=start_row+1, min_col=status_col_idx, max_row=ws.max_row, max_col=status_col_idx):
            for cell in row:
                val = str(cell.value or "")
                if "✓" in val or "Passed" in val:
                    cell.fill = green
                elif "–" in val or "Progress" in val or "Pending" in val:
                    cell.fill = yellow
                elif "✗" in val or "Failed" in val:
                    cell.fill = red
                elif "↻" in val or "Redo" in val:
                    cell.fill = gray
        
        align = Alignment(wrap_text=True, vertical="center")
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border_style
                cell.alignment = align
        
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    # Helper to add flat-level sheet (Finishing)
    def add_flat_level_sheet_project(workbook, sheet_name: str, building_name: str, data_df: pd.DataFrame):
        if data_df.empty:
            ws = workbook.create_sheet(title=sheet_name)
            ws.cell(row=1, column=1, value=f"Building: {building_name}")
            ws.cell(row=3, column=1, value="No finishing data found for this building.")
            return
        
        ws = workbook.create_sheet(title=sheet_name)
        ws.cell(row=1, column=1, value="Building:")
        ws.cell(row=1, column=2, value=building_name)
        ws.cell(row=1, column=7, value="Legend:")
        ws.cell(row=1, column=8, value="✓ = Passed")
        ws.cell(row=1, column=9, value="– = In Progress")
        ws.cell(row=1, column=10, value="✗ = Not Done")
        ws.cell(row=1, column=12, value="Flats:")
        ws.cell(row=1, column=13, value=len(data_df))
        
        ws.append([])
        start_row = 3
        for r in dataframe_to_rows(data_df, index=False, header=True):
            ws.append(r)
        
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        red = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
        
        for row in ws.iter_rows(min_row=start_row+1, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                val = str(cell.value or "")
                if val == "✓":
                    cell.fill = green
                elif val == "–":
                    cell.fill = yellow
                elif val == "✗":
                    cell.fill = red
        
        align = Alignment(wrap_text=True, vertical="center")
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border_style
                cell.alignment = align
    
    # Process each building
    for building in all_buildings:
        building_data = dfp[dfp["__Building"].astype(str).fillna("").replace("", "UNKNOWN").eq(building)]
        if building_data.empty:
            continue
        
        # Classify data
        structural_data = building_data[building_data["__Checklist"].apply(
            lambda x: is_rcc_checklist(x) or is_aluform_checklist(x) or is_handover_checklist(x)
        )]
        finishing_data = building_data[~building_data["__Checklist"].apply(
            lambda x: is_rcc_checklist(x) or is_aluform_checklist(x) or is_handover_checklist(x)
        )]
        
        # Get flats for this building
        def _flat_key(v: str):
            import re as _re
            s = str(v or "").strip()
            m = _re.search(r"(shop|flat|unit|apt|apartment)\s*([A-Z]?\d+)([A-Z]?)", s, _re.I)
            if m:
                num = int(_re.sub(r"\D", "", m.group(2))) if _re.sub(r"\D", "", m.group(2)) else 0
                suf = m.group(3).upper() if m.group(3) else ""
                return (0, num, suf)
            m2 = _re.search(r"(\d+)", s)
            if m2:
                return (1, int(m2.group(1)), "")
            return (9, 0, s.lower())
        
        building_flats = sorted(set(building_data["__Flat"].astype(str).fillna("").replace("", "UNKNOWN")), key=_flat_key)
        
        # Build DataFrames
        structural_df = build_floor_status_sheet_for_building(structural_data)
        finishing_df = build_flat_status_matrix_for_building(finishing_data, building_flats)
        
        # Create sheet names (limit to 31 chars for Excel)
        import re as _re_sheet
        building_clean = _re_sheet.sub(r'[^\w]', '_', building)[:15]
        
        # Add sheets for this building
        add_floor_level_sheet_project(wb, f"{building_clean}_Struct"[:31], building, structural_df)
        add_flat_level_sheet_project(wb, f"{building_clean}_Finish"[:31], building, finishing_df)
    
    # Save to BytesIO
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    
    # Generate filename
    import re as _re_fname
    proj_clean = _re_fname.sub(r'[^\w]', '_', proj)
    fname = f"Flat_Status_Project_{proj_clean}_AllBuildings_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    fname = fname.replace("/", "-").replace("\\", "-")
    
    return send_file(bio, as_attachment=True, download_name=fname, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---------------- Building-wise cumulative dashboard ----------------

@app.get("/building-dashboard")
@login_required
def building_dashboard():
    # Default to session CSV if present; allow override via ?path=
    path = request.args.get("path") or _get_session_eqc_path() or "Combined_EQC.csv"
    if not os.path.exists(path):
        return f"EQC file not found: {path}", 404
    # Reuse robust parsing and inference for project and building
    df = _prepare_frame(_robust_read_eqc_path(path))
    # Compute raw counts per normalized stage using __Stage from _prepare_frame (original/correct logic)
    def counts_by_stage(frame: pd.DataFrame) -> Dict[str, int]:
        if frame is None or frame.empty:
            return {"Pre": 0, "During": 0, "Post": 0}
        
        # Use __Stage which is already normalized (Pre, During, Post, Other)
        # Note: Reinforcement/Shuttering are mapped to 'Pre' by EQC.normalize_stage
        c = frame["__Stage"].value_counts()
        n_pre = int(c.get("Pre", 0))
        n_during = int(c.get("During", 0))
        n_post = int(c.get("Post", 0))
        n_other = int(c.get("Other", 0))

        # Cumulative Logic (matching Weekly_report.py / analysis_eqc.py):
        # Pre = Total (Pre + During + Post + Other)
        # During = During + Post + Other
        # Post = Post + Other
        
        return {
            "Pre": n_pre + n_during + n_post + n_other,
            "During": n_during + n_post + n_other,
            "Post": n_post + n_other,
        }

    data: Dict[str, Dict[str, Dict[str, int]]] = {}
    if not df.empty:
        for proj in sorted(df["__Project"].astype(str).str.strip().unique()):
            sub_p = df[df["__Project"].astype(str).str.strip().eq(proj)]
            # Prefer excluding UNKNOWN building from charts if there are known ones
            buildings = sorted([b for b in sub_p["__Building"].astype(str).replace("", "UNKNOWN").unique()])
            has_known = any(str(b).upper() != "UNKNOWN" for b in buildings)
            if has_known:
                buildings = [b for b in buildings if str(b).upper() != "UNKNOWN"]
            proj_map: Dict[str, Dict[str, int]] = {}
            for b in buildings:
                sub_b = sub_p[sub_p["__Building"].astype(str).fillna("").replace("", "UNKNOWN").eq(b)]
                proj_map[b] = counts_by_stage(sub_b)
            data[proj] = proj_map

    # Prepare a lightweight structure for Chart.js
    # { project: { labels: [buildings], pre: [], during: [], post: [] } }
    chart_data: Dict[str, Dict[str, List[int] | List[str]]] = {}
    for proj, bmap in data.items():
        labels = list(bmap.keys())
        pre = [bmap[b]["Pre"] for b in labels]
        dur = [bmap[b]["During"] for b in labels]
        post = [bmap[b]["Post"] for b in labels]
        chart_data[proj] = {"labels": labels, "pre": pre, "during": dur, "post": post}

    return render_template("building_dashboard.html", chart_data=chart_data, path=path)


# ---------------- Building-wise Issues dashboard ----------------

@app.get("/issues-building-dashboard")
@login_required
def issues_building_dashboard():
    # Determine issues file: ?path=, session, or autodetect
    path = request.args.get("path") or _get_session_issues_path()
    if not path or not os.path.exists(path):
        # Autodetect most recent Instruction CSV in workspace
        candidates: list[str] = []
        try:
            for root, _dirs, files in os.walk(os.getcwd()):
                for f in files:
                    fl = f.lower()
                    if fl.endswith('.csv') and ('instruction' in fl or fl.startswith('csv-instruction-latest-report')):
                        candidates.append(os.path.join(root, f))
        except Exception:
            candidates = []
        if candidates:
            path = max(candidates, key=lambda p: os.path.getmtime(p))
        else:
            path = ""
    if not path or not os.path.exists(path):
        return render_template("issues_building_dashboard.html", chart_data={}, path=""), 200

    # Read and filter issues
    df = _robust_read_issues_path(path)
    # DEMO filter
    for col in ("Project", "Project Name", "Location L0"):
        if col in df.columns:
            mask_demo = df[col].astype(str).str.contains("DEMO", case=False, na=False)
            df = df[~mask_demo]
    # Quality filter (exclude Safety)
    df = ISS._filter_quality(df)

    # Reuse frame prep to infer project and building where possible
    try:
        df = _prepare_frame(df)
    except Exception:
        # Minimal fallback: project/building from L0/L1
        df = df.copy()
        df["__Project"] = df.get("Location L0", df.get("Project Name", "")).astype(str)
        df["__Building"] = df.get("Location L1", "").astype(str)

    # Optional date range filter via query params (start/end)
    # Parse dates from "Raised On Date" using ISS._parse_date_safe
    dates = df.get("Raised On Date")
    if dates is not None:
        dates = dates.map(ISS._parse_date_safe)
    else:
        dates = pd.Series([None] * len(df), index=df.index)
    start_q = request.args.get("start", "").strip()
    end_q = request.args.get("end", "").strip()
    start_d = None
    end_d = None
    if start_q:
        # accept ISO (YYYY-MM-DD) or dd-mm-YYYY (handled by parser)
        start_d = ISS._parse_date_safe(start_q)
    if end_q:
        end_d = ISS._parse_date_safe(end_q)
    if start_d or end_d:
        mask = pd.Series([True] * len(df), index=df.index)
        if start_d:
            mask &= dates.map(lambda d: (d is not None) and (d >= start_d))
        if end_d:
            mask &= dates.map(lambda d: (d is not None) and (d <= end_d))
        df = df[mask]

    # Tag External/Internal per user-configured name
    ext_name = (session.get("issues_external_name", DEFAULT_ISSUES_EXTERNAL) or "").strip()
    has_external = bool(ext_name) and ext_name.strip().lower() != "none"
    df = ISS._tag_external(df, ext_name)

    # Count issues per (project, building, category): Total/Open/Closed
    def _counts(frame: pd.DataFrame) -> Dict[str, int]:
        c = ISS._count_frame(frame)
        return {"Total": int(c.total), "Open": int(c.open), "Closed": int(c.closed)}

    data: Dict[str, Dict[str, Dict[str, int]]] = {}
    if not df.empty:
        for proj in sorted([p for p in df["__Project"].astype(str).str.strip().unique() if p]):
            sub_p = df[df["__Project"].astype(str).str.strip().eq(proj)]
            buildings = sorted([b for b in sub_p["__Building"].astype(str).replace("", "UNKNOWN").unique()])
            has_known = any(str(b).upper() != "UNKNOWN" for b in buildings)
            if has_known:
                buildings = [b for b in buildings if str(b).upper() != "UNKNOWN"]
            proj_map: Dict[str, Dict[str, Dict[str, int]]] = {}
            for b in buildings:
                sub_b = sub_p[sub_p["__Building"].astype(str).fillna("").replace("", "UNKNOWN").eq(b)]
                cat_map: Dict[str, Dict[str, int]] = {}
                for cat, g in sub_b.groupby("__Category"):
                    cat_map[str(cat)] = _counts(g)
                # Ensure both categories present
                for missing in ("External", "Internal"):
                    if missing not in cat_map:
                        cat_map[missing] = {"Total": 0, "Open": 0, "Closed": 0}
                proj_map[b] = cat_map
            data[proj] = proj_map

    # Chart payload per project: labels, total, open, closed
    chart_data: Dict[str, Dict[str, List[int] | List[str]]] = {}
    for proj, bmap in data.items():
        labels = [b for b in bmap.keys() if str(b).strip()]
        labels.sort()
        tot_ext = [bmap[b]["External"]["Total"] for b in labels]
        open_ext = [bmap[b]["External"]["Open"] for b in labels]
        closed_ext = [bmap[b]["External"]["Closed"] for b in labels]
        tot_int = [bmap[b]["Internal"]["Total"] for b in labels]
        open_int = [bmap[b]["Internal"]["Open"] for b in labels]
        closed_int = [bmap[b]["Internal"]["Closed"] for b in labels]
        chart_data[proj] = {
            "labels": labels,
            "ext_total": tot_ext, "ext_open": open_ext, "ext_closed": closed_ext,
            "int_total": tot_int, "int_open": open_int, "int_closed": closed_int,
        }

    # Preserve ISO strings for date inputs
    start_iso = start_d.isoformat() if start_d else ""
    end_iso = end_d.isoformat() if end_d else ""

    return render_template("issues_building_dashboard.html", chart_data=chart_data, path=path, start=start_iso, end=end_iso, external_name=ext_name, has_external=has_external)


# ---------------- Issues Daily Dashboard (per project) ----------------

def _issues_summary_by_project(df: pd.DataFrame, target: date) -> Dict[str, Dict[str, ISS.Counts]]:
    """Aggregate Quality issues (Safety excluded) by project for today/month/all.
    Returns { project: { 'today': Counts, 'month': Counts, 'all': Counts } }"""
    out: Dict[str, Dict[str, ISS.Counts]] = {}
    if df is None or df.empty:
        return out
    df = df.copy()
    # DEMO filter
    for col in ("Project", "Project Name", "Location L0"):
        if col in df.columns:
            mask_demo = df[col].astype(str).str.contains("DEMO", case=False, na=False)
            df = df[~mask_demo]
    # Quality-only
    df = ISS._filter_quality(df)
    # Canonical project key
    df["__ProjectKey"] = df.apply(canonical_project_from_row, axis=1)
    # Dates
    dates = df.get("Raised On Date").map(ISS._parse_date_safe) if "Raised On Date" in df.columns else pd.Series([None] * len(df), index=df.index)
    masks = ISS._timeframe_masks(dates, target)
    # Close masks: use "Current Status Updated On (Date)" for closed-in-period counts
    _upd_col = ISS.UPDATED_DATE_COL
    close_dates = df[_upd_col].map(ISS._parse_date_safe) if _upd_col in df.columns else pd.Series([None] * len(df), index=df.index)
    close_masks = ISS._timeframe_masks(close_dates, target)
    for proj, sub in df.groupby("__ProjectKey"):
        res: Dict[str, ISS.Counts] = {}
        for tf, m in masks.items():
            mask_aligned = m.reindex(sub.index, fill_value=False)
            if tf != "all":
                close_mask_aligned = close_masks[tf].reindex(sub.index, fill_value=False)
                res[tf] = ISS._count_frame(sub[mask_aligned], parent_df=sub, close_mask=close_mask_aligned)
            else:
                res[tf] = ISS._count_frame(sub[mask_aligned])
        out[str(proj)] = res
    return out


@app.get("/issues-daily-dashboard")
@login_required
def issues_daily_dashboard():
    # Determine issues file from session or autodetect
    path = request.args.get("path") or _get_session_issues_path()
    if not path or not os.path.exists(path):
        # Autodetect most recent Instruction CSV in workspace
        candidates: list[str] = []
        try:
            for root, _dirs, files in os.walk(os.getcwd()):
                for f in files:
                    fl = f.lower()
                    if fl.endswith('.csv') and ('instruction' in fl or fl.startswith('csv-instruction-latest-report')):
                        candidates.append(os.path.join(root, f))
        except Exception:
            candidates = []
        if candidates:
            path = max(candidates, key=lambda p: os.path.getmtime(p))
        else:
            path = ""
    if not path or not os.path.exists(path):
        return render_template("issues_dashboard.html", data={}, target=date.today().strftime("%d-%m-%Y")), 200
    # Target date (optional ?date= dd-mm-YYYY or ISO)
    date_q = (request.args.get("date") or "").strip()
    if date_q:
        # try our safe parser; fall back to today
        t = ISS._parse_date_safe(date_q)
        target = t or date.today()
    else:
        target = date.today()
    # Read and compute
    df = _robust_read_issues_path(path)
    # Tag External/Internal per configured name
    ext_name = (session.get("issues_external_name", DEFAULT_ISSUES_EXTERNAL) or "").strip()
    df = ISS._filter_quality(df)
    df = ISS._tag_external(df, ext_name)
    # Canonical project key
    df["__ProjectKey"] = df.apply(canonical_project_from_row, axis=1)
    dates = df.get("Raised On Date").map(ISS._parse_date_safe) if "Raised On Date" in df.columns else pd.Series([None] * len(df), index=df.index)
    masks = ISS._timeframe_masks(dates, target)
    # Close masks: use "Current Status Updated On (Date)" for closed-in-period counts
    _upd_col = ISS.UPDATED_DATE_COL
    close_dates_dd = df[_upd_col].map(ISS._parse_date_safe) if _upd_col in df.columns else pd.Series([None] * len(df), index=df.index)
    close_masks_dd = ISS._timeframe_masks(close_dates_dd, target)
    # Build nested: { project: { 'External': {tf: Counts}, 'Internal': {tf: Counts} } }
    data: Dict[str, Dict[str, Dict[str, ISS.Counts]]] = {}
    for proj, sub in df.groupby("__ProjectKey"):
        pc: Dict[str, Dict[str, ISS.Counts]] = {}
        for cat, g in sub.groupby("__Category"):
            res: Dict[str, ISS.Counts] = {}
            for tf, m in masks.items():
                mask_aligned = m.reindex(g.index, fill_value=False)
                if tf != "all":
                    close_mask_aligned = close_masks_dd[tf].reindex(g.index, fill_value=False)
                    res[tf] = ISS._count_frame(g[mask_aligned], parent_df=g, close_mask=close_mask_aligned)
                else:
                    res[tf] = ISS._count_frame(g[mask_aligned])
            pc[str(cat)] = res
        for missing in ("External", "Internal"):
            if missing not in pc:
                pc[missing] = {tf: ISS.Counts(0, 0, 0) for tf in ("today", "month", "all")}
        data[str(proj)] = pc
    data_sorted = {k: data[k] for k in sorted(data.keys())}
    return render_template("issues_dashboard.html", data=data_sorted, target=target.strftime("%d-%m-%Y"), external_name=ext_name)


# ---------------- External Issues Report (External-only focus) ----------------

@app.get("/external-report")
@login_required
def external_report():
    """External Issues Report (site-wise), styled like the sample CSV:
    - Header with Site and date
    - Summary: Previous, Today's, Total (Total/Open/Closed/% Closed)
    - Details: grouped by 'Raised On Date' sections in descending order
    """
    # Load issues file (session or autodetect)
    path = request.args.get("path") or _get_session_issues_path()
    if not path or not os.path.exists(path):
        candidates: list[str] = []
        try:
            for root, _dirs, files in os.walk(os.getcwd()):
                for f in files:
                    fl = f.lower()
                    if fl.endswith('.csv') and ('instruction' in fl or fl.startswith('csv-instruction-latest-report')):
                        candidates.append(os.path.join(root, f))
        except Exception:
            candidates = []
        if candidates:
            path = max(candidates, key=lambda p: os.path.getmtime(p))
        else:
            path = ""
    ext_name = (session.get("issues_external_name", DEFAULT_ISSUES_EXTERNAL) or "").strip()
    # Early render if no file
    if not path or not os.path.exists(path):
        return render_template(
            "external_report.html",
            projects=[], selected_project="", external_name=ext_name,
            header_date=date.today().strftime("%d-%m-%Y"),
            summary=None, sections=[],
        ), 200

    # Target date (for 'Today's' bucket)
    date_q = (request.args.get("date") or "").strip()
    t = ISS._parse_date_safe(date_q) if date_q else None
    target_d = t or date.today()

    df = _robust_read_issues_path(path)
    df = ISS._filter_quality(df)
    df = ISS._tag_external(df, ext_name)
    df = df[df.get("__Category") == "External"]
    df["__ProjectKey"] = df.apply(canonical_project_from_row, axis=1)

    # Site selection
    projects = sorted([p for p in df["__ProjectKey"].astype(str).str.strip().unique() if p])
    sel = (request.args.get("project") or "").strip()
    if not sel and projects:
        sel = projects[0]
    if sel:
        df = df[df["__ProjectKey"].astype(str).str.strip().eq(sel)]

    # Parse dates safely
    dates = df.get("Raised On Date").map(ISS._parse_date_safe) if "Raised On Date" in df.columns else pd.Series([None] * len(df), index=df.index)
    masks = ISS._timeframe_masks(dates, target_d)
    # Close masks: use "Current Status Updated On (Date)" for closed-in-period counts
    _upd_col = ISS.UPDATED_DATE_COL
    close_dates_er = df[_upd_col].map(ISS._parse_date_safe) if _upd_col in df.columns else pd.Series([None] * len(df), index=df.index)
    close_masks_er = ISS._timeframe_masks(close_dates_er, target_d)

    # Build summary: Previous = All - Today
    _all_mask = masks["all"].reindex(df.index, fill_value=False)
    _today_mask = masks["today"].reindex(df.index, fill_value=False)
    _today_close_mask = close_masks_er["today"].reindex(df.index, fill_value=False)
    all_counts = ISS._count_frame(df[_all_mask]) if not df.empty else ISS.Counts(0,0,0)
    today_counts = ISS._count_frame(df[_today_mask], parent_df=df, close_mask=_today_close_mask) if not df.empty else ISS.Counts(0,0,0)
    prev_total = max(0, all_counts.total - today_counts.total)
    prev_open = max(0, all_counts.open - today_counts.open)
    prev_closed = max(0, all_counts.closed - today_counts.closed)
    def pct_closed(total:int, closed:int) -> float:
        return round((closed/total*100.0), 2) if total else 0.0
    summary = {
        "previous": {"total": prev_total, "open": prev_open, "closed": prev_closed, "pct": pct_closed(prev_total, prev_closed)},
        "today": {"total": today_counts.total, "open": today_counts.open, "closed": today_counts.closed, "pct": pct_closed(today_counts.total, today_counts.closed)},
        "total": {"total": all_counts.total, "open": all_counts.open, "closed": all_counts.closed, "pct": pct_closed(all_counts.total, all_counts.closed)},
    }

    # Build details grouped by date desc
    def _first_nonempty(row: pd.Series, cols: List[str]) -> str:
        for c in cols:
            if c in row and str(row[c]).strip():
                return str(row[c]).strip()
        return ""
    # Compose Location/Reference from L1..L4
    def _compose_location(row: pd.Series) -> str:
        parts = []
        for c in ("Location L1","Location L2","Location L3","Location L4"):
            v = str(row.get(c, "") or "").strip()
            if v:
                parts.append(v)
        return "/".join(parts)

    # Prepare ordering by date (newest first) and then time within date
    df = df.copy()
    df["__RaisedDate"] = dates
    # Try to keep a consistent time ordering if time column exists
    if "Raised On Time" in df.columns:
        try:
            df["__RaisedTimeOrd"] = pd.to_datetime(df["Raised On Time"], errors="coerce")
        except Exception:
            df["__RaisedTimeOrd"] = pd.NaT
    else:
        df["__RaisedTimeOrd"] = pd.NaT

    sections = []
    if not df.empty:
        groups = {}
        for idx, row in df.iterrows():
            d = row.get("__RaisedDate")
            if isinstance(d, date):
                key = d
            else:
                key = None
            groups.setdefault(key, []).append(row)
        # Sort by date desc, with None last
        ordered_keys = sorted(groups.keys(), key=lambda x: (x is None, x), reverse=True)
        for k in ordered_keys:
            rows = groups[k]
            # Sort by time desc if available, else stable
            try:
                rows_sorted = sorted(rows, key=lambda r: (r.get("__RaisedTimeOrd") if hasattr(r, "get") else None), reverse=True)
            except Exception:
                rows_sorted = rows
            sect_rows = []
            for i, r in enumerate(rows_sorted, start=1):
                sect_rows.append({
                    "sr": i,
                    "project": _first_nonempty(r, ["Project Name","Project","Location L0"]),
                    "location": _compose_location(r),
                    "description": _first_nonempty(r, ["Description","Title","Observation"]),
                    "recommendation": str(r.get("Recommendation", "") or "").strip(),
                    "raised_by": str(r.get("Raised By", "") or "").strip(),
                    "raised_on_date": (r.get("Raised On Date") or ""),
                    "raised_on_time": (r.get("Raised On Time") or ""),
                    "deadline_date": (r.get("Deadline Date") or ""),
                    "deadline_time": (r.get("Deadline Time") or ""),
                    "type": _first_nonempty(r, ["Type L0","Type","Category"]),
                    "tag1": str(r.get("Tag 1", "") or "").strip(),
                    "assigned_team": str(r.get("Assigned Team", "") or "").strip(),
                    "assigned_user": _first_nonempty(r, ["Assigned Team User","Assigned User","Assignee"]),
                    "status": _first_nonempty(r, ["Current Status","Status"]).upper(),
                })
            label = k.strftime("%d-%m-%Y") if isinstance(k, date) else ""
            sections.append({"date": label, "rows": sect_rows})

    return render_template(
        "external_report.html",
        projects=projects, selected_project=sel, external_name=ext_name,
        header_date=target_d.strftime("%d-%m-%Y"),
        summary=summary,
        sections=sections,
    )


@app.get("/external-report/export")
@login_required
def external_report_export():
    # Build a simple Excel: one sheet with External-only counts per project
    path = request.args.get("path") or _get_session_issues_path()
    if not path or not os.path.exists(path):
        return "Issues file not found", 404
    df = _robust_read_issues_path(path)
    df = ISS._filter_quality(df)
    ext_name = (session.get("issues_external_name", DEFAULT_ISSUES_EXTERNAL) or "").strip()
    df = ISS._tag_external(df, ext_name)
    df = df[df.get("__Category") == "External"]
    df["__ProjectKey"] = df.apply(canonical_project_from_row, axis=1)
    sel = (request.args.get("project") or "").strip()
    if sel:
        df = df[df["__ProjectKey"].astype(str).str.strip().eq(sel)]
    today = date.today()
    dates = df.get("Raised On Date").map(ISS._parse_date_safe) if "Raised On Date" in df.columns else pd.Series([None] * len(df), index=df.index)
    masks = ISS._timeframe_masks(dates, today)
    # Close masks: use "Current Status Updated On (Date)" for closed-in-period counts
    _upd_col = ISS.UPDATED_DATE_COL
    close_dates_exp = df[_upd_col].map(ISS._parse_date_safe) if _upd_col in df.columns else pd.Series([None] * len(df), index=df.index)
    close_masks_exp = ISS._timeframe_masks(close_dates_exp, today)
    rec: Dict[str, str | int] = {"Project": sel or "(All)"}
    for tf, m in masks.items():
        if tf != "all":
            close_mask_exp = close_masks_exp[tf].reindex(df.index, fill_value=False)
            c = ISS._count_frame(df[m.reindex(df.index, fill_value=False)], parent_df=df, close_mask=close_mask_exp)
        else:
            c = ISS._count_frame(df[m.reindex(df.index, fill_value=False)])
        rec[f"{tf.title()} Total"] = c.total
        rec[f"{tf.title()} Open"] = c.open
        rec[f"{tf.title()} Closed"] = c.closed
    out_df = pd.DataFrame([rec])
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    wb = Workbook(); ws = wb.active; ws.title = "External Issues"
    for r in dataframe_to_rows(out_df, index=False, header=True):
        ws.append(r)
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    suffix = (sel or "All").replace("/","-")
    fname = f"External_Issues_{suffix}_{today.strftime('%Y-%m-%d')}.xlsx"
    # Save to temp file to reuse the same formatting pipeline
    tmp_path = os.path.join(tempfile.gettempdir(), f"ext_export_{uuid.uuid4().hex}.xlsx")
    with open(tmp_path, "wb") as fh:
        fh.write(bio.getvalue())
    formatted = _format_xlsx_from_path(tmp_path)
    try:
        os.remove(tmp_path)
    except Exception:
        pass
    return send_file(formatted, as_attachment=True, download_name=fname, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---------------- Integrate digiqc_Inst (Combined report only) ----------------

def _inst_base_dir() -> str:
    # The repo is cloned in workspace at ./digiqc_Inst
    return os.path.join(os.getcwd(), "digiqc_Inst")


# ---------------- digiqc_Inst Combined report (Excel) ----------------

def _inst_paths() -> dict:
    base = _inst_base_dir()
    return {
        "base": base,
        "main": os.path.join(base, "main.py"),
        "combined_report": os.path.join(base, "Combined_report.xlsx"),
        "combined_summary_xlsx": os.path.join(base, "Combined_Summary.xlsx"),
    }


@app.get("/inst-report")
@login_required
def inst_report_page():
    p = _inst_paths()
    exists = {
        "main": os.path.exists(p["main"]),
        "report": os.path.exists(p["combined_report"]) or os.path.exists(p["combined_summary_xlsx"]),
    }
    return render_template("inst_report.html", exists=exists)


@app.get("/inst-report/build")
@login_required
def inst_report_build():
    p = _inst_paths()
    if not os.path.exists(p["main"]):
        return "digiqc_Inst/main.py not found", 404
    # Allow optional REPORT_DATE passthrough (?date=YYYY-MM-DD)
    env = os.environ.copy()
    report_date = (request.args.get("date") or "").strip()
    if report_date:
        env["REPORT_DATE"] = report_date
    proc = subprocess.run([os.sys.executable, p["main"]], cwd=p["base"], capture_output=True, text=True, env=env)
    if proc.returncode != 0:
        return f"Combined report pipeline failed.\nSTDOUT:\n{proc.stdout}\nSTDERR:\n{proc.stderr}", 500
    return redirect(url_for("inst_report_page"))


@app.get("/inst-report/download")
@login_required
def inst_report_download():
    p = _inst_paths()
    path = p["combined_report"]
    alt = p["combined_summary_xlsx"]
    today_str = date.today().strftime("%d-%m-%Y")
    dl_name = f"Activity wise stage wise instructions details {today_str}.xlsx"
    if not os.path.exists(path):
        if os.path.exists(alt):
            content = _format_xlsx_from_path(alt)
            return send_file(content, as_attachment=True, download_name=dl_name, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        return "Combined report not found. Run build first.", 404
    content = _format_xlsx_from_path(path)
    return send_file(content, as_attachment=True, download_name=dl_name, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ========== AI ANALYSIS ROUTES ==========

from ai_analysis import AIAnalyzer, analyze_eqc_file, analyze_issues_file, standardize_checklists_file

@app.get("/ai-analysis")
@login_required
def ai_analysis_page():
    """AI Analysis dashboard page."""
    eqc_file = session.get("eqc_file_name", "")
    issues_file = session.get("issues_file_name", "")
    return render_template("ai_analysis.html", 
                           eqc_file=eqc_file, 
                           issues_file=issues_file)


@app.post("/api/ai/eqc-completion")
@login_required
def api_ai_eqc_completion():
    """API endpoint for EQC completion analysis."""
    path = _get_session_eqc_path()
    if not path:
        return jsonify({"error": "No EQC file uploaded. Please upload a file first."}), 400
    
    try:
        result = analyze_eqc_file(path, use_ai=True)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.post("/api/ai/issues-patterns")
@login_required
def api_ai_issues_patterns():
    """API endpoint for Issues pattern analysis."""
    path = _get_session_issues_path()
    if not path:
        return jsonify({"error": "No Issues file uploaded. Please upload a file first."}), 400
    
    try:
        # Use threading timeout for cross-platform compatibility (Windows doesn't have signal.alarm)
        import threading
        import queue
        
        result_queue = queue.Queue()
        error_queue = queue.Queue()
        
        def analysis_worker():
            try:
                result = analyze_issues_file(path, use_ai=False)
                # DEBUG: Log the result structure
                print(f"[DEBUG] analyze_issues_file returned: {type(result)} with keys: {result.keys() if isinstance(result, dict) else 'N/A'}", file=sys.stderr)
                result_queue.put(result)
            except Exception as e:
                print(f"[DEBUG] Exception in worker: {str(e)}", file=sys.stderr)
                error_queue.put(str(e))
        
        # Start analysis in a separate thread
        worker_thread = threading.Thread(target=analysis_worker, daemon=True)
        worker_thread.start()
        
        # Wait up to 30 seconds for the result
        worker_thread.join(timeout=30)
        
        # Check if thread is still running (timeout occurred)
        if worker_thread.is_alive():
            print("[DEBUG] Worker thread timeout after 30 seconds", file=sys.stderr)
            return jsonify({
                "error": "Analysis took too long (exceeded 30 seconds).",
                "note": "The file may be too large or the AI model is slow. Try: 1) Upload a smaller file, 2) Simplify the data format, or 3) Try again later"
            }), 504
        
        # Check for errors
        if not error_queue.empty():
            error_msg = error_queue.get()
            print(f"[DEBUG] Error from queue: {error_msg}", file=sys.stderr)
            return jsonify({"error": error_msg}), 500
        
        # Get the result
        if not result_queue.empty():
            result = result_queue.get()
            num_teams = len(result.get('teams', []))
            print(f"[DEBUG] Returning result with {num_teams} teams", file=sys.stderr)
            return jsonify(result)
        
        print("[DEBUG] Result queue is empty!", file=sys.stderr)
        return jsonify({"error": "Unknown error: no result produced"}), 500
        
    except Exception as e:
        print(f"[DEBUG] Top-level exception: {str(e)}", file=sys.stderr)
        return jsonify({"error": str(e)}), 500


@app.post("/api/ai/standardize-checklists")
@login_required
def api_ai_standardize_checklists():
    """API endpoint for checklist name standardization."""
    path = _get_session_eqc_path()
    if not path:
        return jsonify({"error": "No EQC file uploaded. Please upload a file first."}), 400
    
    try:
        result = standardize_checklists_file(path, use_ai=True)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ========== ADVANCED ANALYTICS ROUTES ==========

@app.get("/advanced-analytics")
@login_required
def advanced_analytics_page():
    """Render the advanced analytics dashboard page."""
    eqc_file = os.path.basename(_get_session_eqc_path() or "") or None
    issues_file = os.path.basename(_get_session_issues_path() or "") or None
    return render_template("advanced_analytics.html", eqc_file=eqc_file, issues_file=issues_file)


@app.post("/api/analytics/overdue-issues")
@login_required
def api_overdue_issues():
    """API endpoint for overdue issues analysis."""
    path = _get_session_issues_path()
    if not path:
        return jsonify({"error": "No Issues file uploaded. Please upload a file first."}), 400
    
    try:
        from advanced_analytics import AdvancedAnalytics
        import pandas as pd
        df = pd.read_csv(path)
        analyzer = AdvancedAnalytics()
        result = analyzer.analyze_overdue_issues(df)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# Floor Heatmap - Removed (Future Scope)
# @app.post("/api/analytics/floor-heatmap")
# def api_floor_heatmap(): pass


@app.post("/api/analytics/repeat-offenders")
@login_required
def api_repeat_offenders():
    """API endpoint for repeat offenders analysis."""
    path = _get_session_issues_path()
    if not path:
        return jsonify({"error": "No Issues file uploaded. Please upload a file first."}), 400
    
    try:
        from advanced_analytics import AdvancedAnalytics
        import pandas as pd
        df = pd.read_csv(path)
        analyzer = AdvancedAnalytics()
        result = analyzer.analyze_repeat_offenders(df)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.post("/api/analytics/post-approval-issues")
@login_required
def api_post_approval_issues():
    """API endpoint for post-approval issue detection."""
    eqc_path = _get_session_eqc_path()
    issues_path = _get_session_issues_path()
    
    if not eqc_path:
        return jsonify({"error": "No EQC file uploaded. Please upload a file first."}), 400
    if not issues_path:
        return jsonify({"error": "No Issues file uploaded. Please upload a file first."}), 400
    
    try:
        from advanced_analytics import AdvancedAnalytics
        import pandas as pd
        eqc_df = pd.read_csv(eqc_path)
        issues_df = pd.read_csv(issues_path)
        analyzer = AdvancedAnalytics()
        result = analyzer.detect_post_approval_issues(eqc_df, issues_df)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ========== COMBINED QUALITY DASHBOARD ==========

@app.get("/quality-dashboard")
@login_required
def quality_dashboard_page():
    """Render the combined quality dashboard page."""
    from datetime import date
    import pandas as pd
    
    eqc_path = _get_session_eqc_path()
    issues_path = _get_session_issues_path()
    
    if not eqc_path or not issues_path:
        return render_template("quality_dashboard.html",
                             error="Please upload both EQC and Issues CSV files first.",
                             projects=[], dashboard_date=date.today().strftime('%d/%m/%Y'))
    
    try:
        from quality_dashboard import generate_quality_dashboard
        eqc_df = pd.read_csv(eqc_path)
        issues_df = pd.read_csv(issues_path)
        data = generate_quality_dashboard(eqc_df, issues_df)
        return render_template("quality_dashboard.html",
                             projects=data['projects'],
                             dashboard_date=data['date'],
                             error=None)
    except Exception as e:
        return render_template("quality_dashboard.html",
                             error=str(e),
                             projects=[], dashboard_date=date.today().strftime('%d/%m/%Y'))


# ========== QUALITY EXPERT ASSISTANT ==========

@app.get("/quality-expert")
@login_required
def quality_expert_page():
    """Feature retired to reduce latency."""
    return redirect(url_for("index"))


@app.post("/api/quality-expert/ask")
@login_required
def api_quality_expert_ask():
    """API retired."""
    return jsonify({"error": "Quality Expert is disabled"}), 410


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=True)




